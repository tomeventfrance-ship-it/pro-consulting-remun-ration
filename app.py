import base64
import json
import hashlib
import re
import unicodedata

import pandas as pd
import psycopg
import streamlit as st
from streamlit import runtime
from concurrent.futures import ThreadPoolExecutor, as_completed
from cryptography.hazmat.primitives import serialization
from cryptography.hazmat.primitives.asymmetric import ec
from datetime import date, datetime
from html import escape
from io import BytesIO, StringIO
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from urllib.request import Request, urlopen
from xml.etree import ElementTree
from zoneinfo import ZoneInfo
from pywebpush import WebPushException, webpush

from utils import (
    calculate_consultant_rewards,
    calculate_creator_rewards,
    prepare_backstage_data,
)
from tournaments import (
    add_participants,
    create_tournament,
    delete_tournament,
    finalize_draw,
    initialize_tournament_database,
    list_tournaments,
    load_tournament,
    parse_participant_batch,
    remove_participant,
    reopen_registrations,
    set_match_winner,
    tournament_tables,
    update_match_schedule,
    update_tournament_options,
    validate_prepared_draw,
)


st.set_page_config(
    page_title="Pro Consulting",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="expanded",
)


st.markdown(
    """
    <style>
    :root {
        --pc-lagoon: #00d7c8;
        --pc-lagoon-light: #73fff0;
        --pc-teal: #007c86;
        --pc-teal-deep: #071c2b;
        --pc-sand: #e7b76a;
        --pc-gold: #ffc857;
        --pc-cream: #fff8ea;
        --pc-ink: #071b25;
        --pc-panel: rgba(8, 38, 52, 0.9);
        --pc-border: rgba(255, 200, 87, 0.38);
    }

    /* L'application utilise son propre menu : le menu multipage natif est un doublon. */
    [data-testid="stSidebarNav"] {
        display: none !important;
    }

    [data-testid="stAppViewContainer"] {
        background:
            radial-gradient(circle at 86% 8%, rgba(0, 215, 200, 0.22), transparent 30%),
            radial-gradient(circle at 10% 92%, rgba(255, 200, 87, 0.13), transparent 28%),
            linear-gradient(145deg, #071927 0%, #0a2636 48%, #0d3442 100%);
        color: var(--pc-cream);
    }

    [data-testid="stHeader"] {
        background: transparent;
    }

    [data-testid="stDecoration"] {
        background: linear-gradient(90deg, var(--pc-lagoon), var(--pc-gold));
    }

    .block-container {
        max-width: 1480px;
        padding-top: 2rem;
        padding-bottom: 4rem;
    }

    h1, h2, h3 {
        color: var(--pc-cream) !important;
        letter-spacing: -0.025em;
    }

    h1 {
        font-weight: 780 !important;
    }

    h2, h3 {
        font-weight: 700 !important;
    }

    a {
        color: var(--pc-lagoon-light) !important;
    }

    [data-testid="stSidebar"] {
        background:
            radial-gradient(circle at 20% 5%, rgba(0, 215, 200, 0.18), transparent 27%),
            linear-gradient(180deg, #071a28 0%, #0b2938 58%, #0d3a46 100%);
        border-right: 1px solid rgba(255, 200, 87, 0.34);
    }

    [data-testid="stSidebar"] [data-testid="stSidebarContent"] {
        padding-top: 1.15rem;
    }

    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] span {
        color: var(--pc-cream) !important;
    }

    [data-testid="stSidebar"] [role="radiogroup"] label {
        border: 1px solid transparent;
        border-radius: 12px;
        margin: 0.18rem 0;
        padding: 0.46rem 0.6rem;
        transition: all 0.2s ease;
    }

    [data-testid="stSidebar"] [role="radiogroup"] label:hover {
        background: rgba(0, 215, 200, 0.12);
        border-color: rgba(115, 255, 240, 0.42);
        transform: translateX(2px);
    }

    [data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) {
        background: linear-gradient(90deg, rgba(0, 215, 200, 0.3), rgba(255, 200, 87, 0.17));
        border-color: rgba(255, 200, 87, 0.56);
        box-shadow: inset 3px 0 0 var(--pc-gold);
    }

    [data-testid="stMetric"] {
        height: 100%;
        padding: 1.05rem 1.15rem;
        border: 1px solid var(--pc-border);
        border-radius: 17px;
        background: linear-gradient(145deg, rgba(12, 54, 70, 0.95), rgba(7, 29, 42, 0.88));
        box-shadow: 0 12px 30px rgba(0, 8, 18, 0.32);
        backdrop-filter: blur(14px);
    }

    [data-testid="stMetricLabel"] {
        color: var(--pc-lagoon-light) !important;
        font-weight: 650;
    }

    [data-testid="stMetricValue"] {
        color: var(--pc-gold) !important;
        font-weight: 760;
    }

    [data-testid="stForm"],
    [data-testid="stExpander"] {
        border: 1px solid var(--pc-border) !important;
        border-radius: 18px !important;
        background: linear-gradient(145deg, rgba(12, 48, 62, 0.94), rgba(8, 32, 45, 0.88));
        box-shadow: 0 14px 34px rgba(0, 8, 18, 0.3);
    }

    [data-testid="stForm"] {
        padding: 1.25rem 1.35rem;
    }

    [data-testid="stFileUploaderDropzone"] {
        border: 1px dashed rgba(115, 255, 240, 0.62);
        border-radius: 16px;
        background: rgba(0, 124, 134, 0.2);
    }

    [data-baseweb="input"] > div,
    [data-baseweb="select"] > div,
    [data-baseweb="textarea"] > div {
        border-color: rgba(115, 255, 240, 0.34) !important;
        border-radius: 11px !important;
        background-color: rgba(7, 29, 42, 0.82) !important;
    }

    [data-baseweb="input"] input,
    [data-baseweb="select"] input,
    [data-baseweb="textarea"] textarea {
        color: var(--pc-cream) !important;
    }

    [data-testid="stTextArea"] [data-baseweb="textarea"],
    [data-testid="stTextArea"] [data-baseweb="textarea"] > div,
    [data-testid="stTextArea"] textarea {
        border-color: rgba(115, 255, 240, 0.48) !important;
        background: #071d2b !important;
        color: #fff8ea !important;
    }

    [data-testid="stTextArea"] textarea {
        caret-color: var(--pc-gold) !important;
        -webkit-text-fill-color: #fff8ea !important;
    }

    [data-testid="stTextArea"] textarea::placeholder {
        color: rgba(255, 248, 234, 0.58) !important;
        -webkit-text-fill-color: rgba(255, 248, 234, 0.58) !important;
        opacity: 1 !important;
    }

    [data-testid="stTextArea"] label,
    [data-testid="stTextArea"] label p {
        color: var(--pc-cream) !important;
    }

    [data-testid="stNumberInput"] button {
        color: var(--pc-gold) !important;
        border-color: rgba(239, 206, 136, 0.2) !important;
        background: rgba(255, 255, 255, 0.035) !important;
    }

    .stButton > button,
    .stDownloadButton > button,
    [data-testid="stFormSubmitButton"] > button {
        min-height: 2.75rem;
        border: 1px solid rgba(255, 200, 87, 0.78) !important;
        border-radius: 12px !important;
        background: linear-gradient(105deg, #f4b942 0%, #ffd166 52%, #00d7c8 145%) !important;
        color: var(--pc-ink) !important;
        font-weight: 760 !important;
        box-shadow: 0 8px 22px rgba(0, 32, 36, 0.22);
        transition: transform 0.18s ease, box-shadow 0.18s ease;
    }

    [data-testid="stSidebar"] .stButton button p,
    [data-testid="stSidebar"] .stButton button span {
        color: var(--pc-ink) !important;
        font-weight: 800 !important;
    }

    .stButton > button:hover,
    .stDownloadButton > button:hover,
    [data-testid="stFormSubmitButton"] > button:hover {
        border-color: var(--pc-cream) !important;
        transform: translateY(-1px);
        box-shadow: 0 11px 26px rgba(0, 32, 36, 0.28);
    }

    .stButton > button:disabled,
    [data-testid="stFormSubmitButton"] > button:disabled {
        opacity: 0.55;
        transform: none;
    }

    [data-testid="stAlert"] {
        border-radius: 14px;
        border: 1px solid rgba(255, 255, 255, 0.12);
        box-shadow: 0 8px 22px rgba(0, 28, 32, 0.12);
    }

    [data-testid="stDataFrame"],
    [data-testid="stDataEditor"] {
        overflow: hidden;
        border: 1px solid var(--pc-border);
        border-radius: 16px;
        box-shadow: 0 14px 34px rgba(0, 25, 30, 0.18);
    }

    hr {
        border-color: rgba(239, 206, 136, 0.18) !important;
    }

    .pc-hero {
        position: relative;
        overflow: hidden;
        margin: 0 0 1.55rem;
        padding: 2.25rem 2.4rem;
        border: 1px solid rgba(255, 200, 87, 0.58);
        border-radius: 24px;
        background:
            radial-gradient(circle at 90% 18%, rgba(255, 248, 234, 0.2), transparent 24%),
            linear-gradient(118deg, #071e2c 0%, #006f78 58%, #e7ad49 145%);
        box-shadow: 0 24px 52px rgba(0, 7, 18, 0.4);
    }

    .pc-hero::after {
        content: "◆";
        position: absolute;
        right: 3.2rem;
        top: 50%;
        color: rgba(247, 241, 229, 0.16);
        font-size: 7rem;
        line-height: 1;
        transform: translateY(-50%) rotate(45deg);
    }

    .pc-hero-kicker {
        margin-bottom: 0.55rem;
        color: var(--pc-gold);
        font-size: 0.78rem;
        font-weight: 800;
        letter-spacing: 0.17em;
        text-transform: uppercase;
    }

    .pc-hero h1 {
        position: relative;
        z-index: 1;
        margin: 0 0 0.55rem;
        max-width: 820px;
        color: #fffaf0 !important;
        font-size: clamp(2rem, 4vw, 3.3rem);
        line-height: 1.05;
    }

    .pc-hero p {
        position: relative;
        z-index: 1;
        max-width: 780px;
        margin: 0;
        color: rgba(247, 241, 229, 0.88);
        font-size: 1.04rem;
        line-height: 1.6;
    }

    .pc-sidebar-brand {
        margin: 0.2rem 0 1rem;
        padding: 1rem 1.05rem;
        border: 1px solid rgba(255, 200, 87, 0.46);
        border-radius: 17px;
        background: linear-gradient(145deg, rgba(12, 58, 72, 0.96), rgba(7, 29, 42, 0.9));
        box-shadow: 0 12px 28px rgba(0, 7, 18, 0.34);
    }

    .pc-sidebar-logo {
        color: #fffaf0;
        font-size: 1.3rem;
        font-weight: 820;
        letter-spacing: -0.025em;
    }

    .pc-sidebar-logo span {
        color: var(--pc-gold) !important;
    }

    .pc-sidebar-tagline {
        margin-top: 0.25rem;
        color: rgba(247, 241, 229, 0.7) !important;
        font-size: 0.76rem;
        letter-spacing: 0.04em;
    }

    .pc-user-card {
        margin: 0 0 0.8rem;
        padding: 0.85rem 0.95rem;
        border-left: 3px solid var(--pc-gold);
        border-radius: 10px;
        background: linear-gradient(100deg, rgba(0, 124, 134, 0.22), rgba(7, 29, 42, 0.62));
    }

    .pc-user-card strong {
        display: block;
        color: #fffaf0;
        font-size: 0.98rem;
    }

    .pc-user-card span {
        display: block;
        margin-top: 0.15rem;
        color: rgba(247, 241, 229, 0.7) !important;
        font-size: 0.78rem;
    }

    .pc-user-status {
        margin-bottom: 0.4rem;
        color: var(--pc-lagoon-light) !important;
        font-size: 0.72rem;
        font-weight: 760;
        letter-spacing: 0.08em;
        text-transform: uppercase;
    }

    .pc-chat-message {
        margin: 0.55rem 0;
        padding: 0.85rem 1rem;
        border: 1px solid rgba(115, 255, 240, 0.24);
        border-radius: 14px;
        background: linear-gradient(
            145deg,
            rgba(12, 54, 70, 0.94),
            rgba(7, 29, 42, 0.88)
        );
    }

    .pc-chat-meta {
        margin-bottom: 0.35rem;
        color: var(--pc-gold);
        font-size: 0.78rem;
        font-weight: 760;
    }

    .pc-chat-body {
        color: var(--pc-cream);
        line-height: 1.45;
        overflow-wrap: anywhere;
    }

    @media (max-width: 760px) {
        .block-container {
            padding: 1rem 0.85rem 3rem;
        }

        .pc-hero {
            padding: 1.5rem 1.3rem;
            border-radius: 19px;
        }

        .pc-hero::after {
            right: 1rem;
            font-size: 4.4rem;
        }

        .pc-hero p {
            padding-right: 1.4rem;
            font-size: 0.94rem;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def render_brand_hero(title, subtitle, kicker="PRO CONSULTING"):
    st.markdown(
        f"""
        <section class="pc-hero">
            <div class="pc-hero-kicker">{escape(kicker)}</div>
            <h1>{escape(title)}</h1>
            <p>{escape(subtitle)}</p>
        </section>
        """,
        unsafe_allow_html=True,
    )


def safe_export_name(value):
    """Construit un fragment de nom de fichier lisible et portable."""
    ascii_value = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = ascii_value.encode("ascii", "ignore").decode("ascii")
    cleaned_value = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_value)
    return cleaned_value.strip("_").lower() or "export"


def safe_sheet_name(value):
    """Respecte les contraintes de nommage des feuilles Excel."""
    cleaned_value = re.sub(r"[\\/*?:\[\]]", "-", str(value or "Données"))
    return cleaned_value[:31] or "Données"


def prepare_excel_dataframe(dataframe):
    """Prépare une copie exportable sans modifier le tableau affiché."""
    export_dataframe = dataframe.copy().reset_index(drop=True)

    for column in export_dataframe.select_dtypes(include=["datetimetz"]):
        export_dataframe[column] = export_dataframe[column].dt.tz_localize(
            None
        )

    return export_dataframe


def write_excel_sheet(writer, dataframe, sheet_name):
    """Ajoute une feuille lisible avec en-tête figé et filtres actifs."""
    export_dataframe = prepare_excel_dataframe(dataframe)
    safe_name = safe_sheet_name(sheet_name)
    export_dataframe.to_excel(writer, index=False, sheet_name=safe_name)

    worksheet = writer.sheets[safe_name]
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_index, column_name in enumerate(
        export_dataframe.columns,
        start=1,
    ):
        displayed_values = export_dataframe[column_name].head(500).tolist()
        maximum_length = max(
            [len(str(column_name))]
            + [len(str(value)) for value in displayed_values]
        )
        worksheet.column_dimensions[
            worksheet.cell(row=1, column=column_index).column_letter
        ].width = min(maximum_length + 2, 45)


@st.cache_data(show_spinner=False)
def dataframe_to_excel(dataframe, sheet_name):
    """Transforme un tableau en classeur Excel téléchargeable."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        write_excel_sheet(writer, dataframe, sheet_name)
    return output.getvalue()


@st.cache_data(show_spinner=False)
def dataframes_to_excel(workbook_tables):
    """Crée un classeur Excel regroupant plusieurs tableaux."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, dataframe in workbook_tables:
            write_excel_sheet(writer, dataframe, sheet_name)
    return output.getvalue()


@st.cache_data(show_spinner=False)
def reward_tracking_to_excel(dataframe):
    """Crée le suivi avec formules et choix Live/Match compatibles Sheets."""
    export_dataframe = prepare_excel_dataframe(dataframe)
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_dataframe.to_excel(
            writer,
            index=False,
            sheet_name="Suivi récompenses",
        )
        worksheet = writer.sheets["Suivi récompenses"]
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False

        header_fill = PatternFill("solid", fgColor="007C86")
        total_fill = PatternFill("solid", fgColor="FFC857")
        sent_fill = PatternFill("solid", fgColor="B7F7D0")
        refused_fill = PatternFill("solid", fgColor="F4B7B7")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        worksheet.row_dimensions[1].height = 34

        data_last_row = len(export_dataframe) + 1
        if len(export_dataframe):
            worksheet.auto_filter.ref = f"A1:J{data_last_row}"
            event_validation = DataValidation(
                type="list",
                formula1='"Live,Match"',
                allow_blank=True,
            )
            event_validation.error = "Choisissez Live ou Match."
            event_validation.errorTitle = "Type d’événement incorrect"
            worksheet.add_data_validation(event_validation)
            event_validation.add(f"C2:C{data_last_row}")

            for row_number in range(2, data_last_row + 1):
                worksheet[f"J{row_number}"] = (
                    f"=H{row_number}+I{row_number}"
                )
                for column_letter in ("H", "I", "J"):
                    worksheet[f"{column_letter}{row_number}"].number_format = (
                        "#,##0"
                    )
                reward_is_refused = bool(
                    export_dataframe.iloc[row_number - 2].get(
                        "Récompense refusée",
                        False,
                    )
                )
                reward_is_validated = bool(
                    export_dataframe.iloc[row_number - 2].get(
                        "Récompense validée",
                        False,
                    )
                )
                if reward_is_refused:
                    for cell in worksheet[row_number]:
                        cell.fill = refused_fill
                        cell.font = Font(color="6B0F1A", bold=True)
                elif reward_is_validated:
                    for cell in worksheet[row_number]:
                        cell.fill = sent_fill
                        cell.font = Font(color="073B24", bold=True)

            total_row = data_last_row + 1
            worksheet[f"G{total_row}"] = "TOTAL GÉNÉRAL"
            for column_letter in ("H", "I", "J"):
                worksheet[f"{column_letter}{total_row}"] = (
                    f"=SUM({column_letter}2:{column_letter}{data_last_row})"
                )
                worksheet[f"{column_letter}{total_row}"].number_format = (
                    "#,##0"
                )
            for cell in worksheet[total_row]:
                cell.fill = total_fill
                cell.font = Font(color="071B25", bold=True)

        column_widths = {
            "A": 14,
            "B": 11,
            "C": 18,
            "D": 10,
            "E": 10,
            "F": 28,
            "G": 24,
            "H": 25,
            "I": 38,
            "J": 23,
        }
        for column_letter, width in column_widths.items():
            worksheet.column_dimensions[column_letter].width = width

        writer.book.calculation.fullCalcOnLoad = True
        writer.book.calculation.forceFullCalc = True

    return output.getvalue()


def export_filename(table_name):
    """Ajoute automatiquement la direction et le mois au nom du fichier."""
    direction = globals().get("current_user_direction", "direction")
    month = st.session_state.get("month", "mois")
    return "_".join(
        [
            "pro_consulting",
            safe_export_name(direction),
            safe_export_name(month),
            safe_export_name(table_name),
        ]
    ) + ".xlsx"


def show_excel_download(
    dataframe,
    table_name,
    sheet_name,
    key,
    label="⬇️ Télécharger ce tableau (Excel)",
):
    """Affiche un bouton d'export Excel pour tout administrateur/directeur."""
    st.download_button(
        label=label,
        data=dataframe_to_excel(dataframe, sheet_name),
        file_name=export_filename(table_name),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        key=key,
        use_container_width=True,
        on_click="ignore",
    )


def show_workbook_download(
    workbook_tables,
    table_name,
    key,
    label="⬇️ Télécharger le classeur complet (Excel)",
):
    """Affiche un bouton d'export regroupant plusieurs feuilles Excel."""
    st.download_button(
        label=label,
        data=dataframes_to_excel(tuple(workbook_tables)),
        file_name=export_filename(table_name),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        key=key,
        use_container_width=True,
        type="primary",
        on_click="ignore",
    )


DEFAULT_VALUES = {
    "backstage_data": None,
    "backstage_filename": None,
    "month": "Août 2026",
    "creator_level": 9,
    "consultant_level": 9,
    "manager_level": 9,
    "director_level": 8,
    "revenue_usd": 0.0,
    "usd_to_eur": 0.92,
    "other_expenses": 0.0,
    "coin_pack_price": 11.183,
    "invoice_rate": 0.0084,
    "charges_rate": 24.6,
    "exclusions": [],
    "director_branch_revenue_usd": 0.0,
    "director_branch_other_expenses": 0.0,
    "use_ecb_rate": True,
    "last_ecb_usd_to_eur": None,
    "last_ecb_rate_date": None,
}

GLOBAL_FINANCIAL_KEYS = (
    "coin_pack_price",
    "invoice_rate",
    "charges_rate",
)

GLOBAL_EXCHANGE_RATE_KEYS = (
    "usd_to_eur",
    "last_ecb_usd_to_eur",
    "last_ecb_rate_date",
)

BRANCH_PARAMETER_KEYS = (
    "month",
    "creator_level",
    "consultant_level",
    "manager_level",
    "director_level",
    "revenue_usd",
    "other_expenses",
    "director_branch_revenue_usd",
    "director_branch_other_expenses",
    "use_ecb_rate",
)

for key, default_value in DEFAULT_VALUES.items():
    if key not in st.session_state:
        st.session_state[key] = default_value


def get_database_url():
    """Retourne l'URL PostgreSQL stockée dans les Secrets Streamlit."""
    try:
        if "database" not in st.secrets:
            return None

        database = st.secrets["database"]
        database_url = str(database.get("url", "")).strip()
        return database_url or None
    except (FileNotFoundError, KeyError, TypeError):
        return None


@st.cache_resource(show_spinner=False)
def initialize_settings_database(database_url):
    """Crée les tables persistantes nécessaires si elles n'existent pas."""
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS pro_consulting_settings (
                    scope TEXT PRIMARY KEY,
                    payload JSONB NOT NULL,
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_by TEXT NOT NULL
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS pro_consulting_chat_messages (
                    id BIGSERIAL PRIMARY KEY,
                    channel TEXT NOT NULL DEFAULT 'general',
                    author_email TEXT NOT NULL,
                    author_name TEXT NOT NULL,
                    author_role TEXT NOT NULL,
                    message TEXT NOT NULL,
                    created_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    expires_at TIMESTAMPTZ NOT NULL
                )
                """
            )
            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    pro_consulting_chat_active_messages_idx
                ON pro_consulting_chat_messages (channel, expires_at)
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS pro_consulting_chat_reads (
                    user_email TEXT PRIMARY KEY,
                    last_seen_message_id BIGINT NOT NULL DEFAULT 0,
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS
                    pro_consulting_push_subscriptions (
                        endpoint TEXT PRIMARY KEY,
                        user_email TEXT NOT NULL,
                        subscription JSONB NOT NULL,
                        active BOOLEAN NOT NULL DEFAULT TRUE,
                        created_at TIMESTAMPTZ NOT NULL
                            DEFAULT CURRENT_TIMESTAMP,
                        updated_at TIMESTAMPTZ NOT NULL
                            DEFAULT CURRENT_TIMESTAMP
                    )
                """
            )
            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    pro_consulting_push_subscriptions_user_idx
                ON pro_consulting_push_subscriptions (
                    user_email,
                    active
                )
                """
            )
        connection.commit()

    initialize_tournament_database(database_url)

    return True


def load_persistent_scope(database_url, scope):
    """Charge un bloc de paramètres depuis PostgreSQL."""
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT payload
                FROM pro_consulting_settings
                WHERE scope = %s
                """,
                (scope,),
            )
            row = cursor.fetchone()

    if row is None:
        return {}

    payload = row[0]
    if isinstance(payload, str):
        return json.loads(payload)
    return dict(payload)


def load_persistent_scopes(database_url, scopes):
    """Charge plusieurs blocs persistants avec une seule connexion."""
    requested_scopes = tuple(dict.fromkeys(str(scope) for scope in scopes))
    if not requested_scopes:
        return {}

    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT scope, payload
                FROM pro_consulting_settings
                WHERE scope = ANY(%s)
                """,
                (list(requested_scopes),),
            )
            rows = cursor.fetchall()

    loaded_scopes = {scope: {} for scope in requested_scopes}
    for scope, payload in rows:
        if isinstance(payload, str):
            payload = json.loads(payload)
        loaded_scopes[str(scope)] = dict(payload)
    return loaded_scopes


def save_persistent_scopes(database_url, scopes, updated_by):
    """Enregistre plusieurs blocs de paramètres dans une transaction."""
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            for scope, payload in scopes.items():
                cursor.execute(
                    """
                    INSERT INTO pro_consulting_settings (
                        scope,
                        payload,
                        updated_by
                    )
                    VALUES (%s, %s::jsonb, %s)
                    ON CONFLICT (scope)
                    DO UPDATE SET
                        payload = EXCLUDED.payload,
                        updated_at = CURRENT_TIMESTAMP,
                        updated_by = EXCLUDED.updated_by
                    """,
                    (
                        scope,
                        json.dumps(payload, ensure_ascii=False),
                        updated_by,
                    ),
                )
        connection.commit()

    # Les accès sont consultés très souvent par Streamlit. Leur cache court
    # est invalidé après toute écriture afin qu'une modification faite par
    # l'administrateur soit visible immédiatement dans ce processus.
    authorization_scopes = {
        "admin:director_management",
        "admin:collaborator_access",
    }
    if authorization_scopes.intersection(scopes):
        authorization_loader = globals().get("load_authorization_payloads")
        if authorization_loader is not None:
            authorization_loader.clear()


def delete_persistent_scopes(database_url, scopes):
    """Supprime uniquement les blocs persistants explicitement demandés."""
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            for scope in scopes:
                cursor.execute(
                    """
                    DELETE FROM pro_consulting_settings
                    WHERE scope = %s
                    """,
                    (scope,),
                )
        connection.commit()


def branch_settings_scope(user_email):
    """Chaque direction conserve ses propres paramètres mensuels."""
    return f"branch:{normalize_email(user_email)}"


def exclusions_scope(user_email):
    """Chaque direction conserve sa propre liste d'exclusions."""
    return f"exclusions:{normalize_email(user_email)}"


def backstage_import_scope(user_email):
    """Chaque compte conserve uniquement son dernier import validé."""
    return f"backstage:{normalize_email(user_email)}"


def director_management_scope():
    """Configuration privée des quatre directions, gérée par le fondateur."""
    return "admin:director_management"


def collaborator_access_scope():
    """Accès ajoutés et gérés uniquement depuis le compte administrateur."""
    return "admin:collaborator_access"


@st.cache_data(ttl=15, show_spinner=False)
def load_authorization_payloads(database_url):
    """Mémorise brièvement les accès pour accélérer les reruns de widgets."""
    scope_names = (
        director_management_scope(),
        collaborator_access_scope(),
    )
    return load_persistent_scopes(database_url, scope_names)


def reward_tracking_scope():
    """Un registre actif unique est partagé par les cinq comptes."""
    return "shared:reward_tracking:current"


def legacy_reward_tracking_scope(month):
    """Clé utilisée avant la mise en place du registre collectif unique."""
    return ":".join(
        [
            "shared",
            "reward_tracking",
            safe_export_name(month),
        ]
    )


def reward_tracking_archive_scope(month):
    """Archive immuable d'un mois collectif termine."""
    return ":".join(
        [
            "shared",
            "reward_tracking",
            "archive",
            safe_export_name(month),
        ]
    )


CHAT_RETENTION_HOURS = 48
CHAT_MAX_MESSAGE_LENGTH = 2000
WEB_PUSH_KEYS_SCOPE = "system:web_push_keys:v1"
CHAT_PUSH_FRONTEND_VERSION = "v4_20260826"


CHAT_PUSH_SERVICE_WORKER_JS = r"""
self.addEventListener("push", (event) => {
    let payload = {};
    try {
        payload = event.data ? event.data.json() : {};
    } catch (error) {
        payload = {
            title: "Nouveau message Pro Consulting",
            body: event.data ? event.data.text() : "Ouvrez le chat collectif.",
        };
    }

    const notificationTitle = payload.title || "Nouveau message Pro Consulting";
    const notificationOptions = {
        body: payload.body || "Ouvrez le chat collectif.",
        data: {
            url: payload.url || "/?page=chat",
        },
        tag: payload.tag || "pro-consulting-chat",
        renotify: true,
        vibrate: [180, 80, 180],
    };

    event.waitUntil(
        self.registration.showNotification(
            notificationTitle,
            notificationOptions,
        ),
    );
});

self.addEventListener("notificationclick", (event) => {
    event.notification.close();
    const targetUrl = new URL(
        event.notification.data?.url || "/?page=chat",
        self.location.origin,
    ).href;

    event.waitUntil(
        self.clients
            .matchAll({ type: "window", includeUncontrolled: true })
            .then(async (clientList) => {
                for (const client of clientList) {
                    if ("navigate" in client) {
                        await client.navigate(targetUrl);
                    }
                    if ("focus" in client) {
                        return client.focus();
                    }
                }
                if (self.clients.openWindow) {
                    return self.clients.openWindow(targetUrl);
                }
                return undefined;
            }),
    );
});
"""


CHAT_PUSH_MANIFEST_JSON = """
{
  "name": "Pro Consulting",
  "short_name": "Pro Consulting",
  "description": "Pilotage et chat collectif Pro Consulting",
  "start_url": "/?page=chat",
  "display": "standalone",
  "background_color": "#071927",
  "theme_color": "#007c86"
}
"""


def register_chat_push_assets():
    """Publie les fichiers Web Push via les routes média de Streamlit."""
    if not runtime.exists():
        return "", ""
    media_manager = runtime.get_instance().media_file_mgr
    worker_url = media_manager.add(
        CHAT_PUSH_SERVICE_WORKER_JS.encode("utf-8"),
        "text/javascript",
        "chat.push.worker.v4",
        file_name="chat-push-sw.js",
    )
    manifest_url = media_manager.add(
        CHAT_PUSH_MANIFEST_JSON.encode("utf-8"),
        "application/manifest+json",
        "chat.push.manifest.v4",
        file_name="manifest.webmanifest",
    )
    return worker_url.lstrip("/"), manifest_url.lstrip("/")


CHAT_PUSH_WORKER_URL, CHAT_PUSH_MANIFEST_URL = (
    register_chat_push_assets()
)


CHAT_PUSH_COMPONENT_HTML = """
<div class="pc-push-control">
    <button type="button" data-pc-push-button>
        🔔 Activer les notifications mobiles
    </button>
    <div data-pc-push-status class="pc-push-status">
        Recevez une alerte lorsqu’un nouveau message est publié.
    </div>
</div>
"""


CHAT_PUSH_COMPONENT_CSS = """
.pc-push-control {
    display: grid;
    gap: 0.55rem;
    width: 100%;
}

.pc-push-control button {
    min-height: 2.75rem;
    width: 100%;
    padding: 0.65rem 1rem;
    border: 1px solid rgba(255, 200, 87, 0.78);
    border-radius: 12px;
    background: linear-gradient(105deg, #f4b942 0%, #ffd166 52%, #00d7c8 145%);
    color: #071b25;
    font: inherit;
    font-weight: 760;
    cursor: pointer;
}

.pc-push-control button:disabled {
    cursor: default;
    opacity: 0.78;
}

.pc-push-status {
    color: rgba(255, 248, 234, 0.78);
    font-size: 0.83rem;
    line-height: 1.4;
}
"""


CHAT_PUSH_COMPONENT_JS = """
export default function(component) {
    const { data, parentElement, setStateValue } = component;
    const button = parentElement.querySelector('[data-pc-push-button]');
    const status = parentElement.querySelector('[data-pc-push-status]');
    const publicKey = data.public_key;
    // Streamlit Cloud exécute l’application sous un chemin interne (/~/+/).
    // Une URL relative conserve ce préfixe indispensable aux composants.
    const workerUrl = new URL(data.worker_url, window.location.href);
    const manifestUrl = new URL(data.manifest_url, window.location.href);
    const workerScope = new URL('./', workerUrl).pathname;

    const publishState = (permission, subscription, message) => {
        setStateValue('push_state', JSON.stringify({
            permission,
            subscription,
            message,
        }));
    };

    const setStatus = (message, enabled = true) => {
        status.textContent = message;
        button.disabled = !enabled;
    };

    const toUint8Array = (base64String) => {
        const padding = '='.repeat((4 - base64String.length % 4) % 4);
        const base64 = (base64String + padding)
            .replace(/-/g, '+')
            .replace(/_/g, '/');
        const rawData = window.atob(base64);
        return Uint8Array.from([...rawData].map((char) => char.charCodeAt(0)));
    };

    const ensureManifest = () => {
        if (document.querySelector('link[data-pc-chat-manifest]')) return;
        const manifest = document.createElement('link');
        manifest.rel = 'manifest';
        manifest.href = manifestUrl.href;
        manifest.dataset.pcChatManifest = 'true';
        document.head.appendChild(manifest);
    };

    const getRegistration = async () => {
        let registration = await navigator.serviceWorker.getRegistration(
            workerScope
        );
        if (!registration) {
            registration = await navigator.serviceWorker.register(
                workerUrl.href,
                { scope: workerScope }
            );
        }
        return registration;
    };

    const inspectSubscription = async () => {
        ensureManifest();
        if (!('serviceWorker' in navigator) || !('PushManager' in window)) {
            setStatus(
                'Ce navigateur ne prend pas en charge les notifications mobiles.',
                false
            );
            publishState('unsupported', null, 'Navigateur non compatible');
            return;
        }

        if (!('Notification' in window)) {
            setStatus(
                'Les notifications ne sont pas disponibles dans ce navigateur.',
                false
            );
            publishState('unsupported', null, 'Notifications indisponibles');
            return;
        }

        const registration = await navigator.serviceWorker.getRegistration(
            workerScope
        );
        const subscription = registration
            ? await registration.pushManager.getSubscription()
            : null;
        if (Notification.permission === 'granted' && subscription) {
            button.textContent = '✅ Notifications mobiles activées';
            setStatus(
                'Ce téléphone recevra les nouveaux messages du chat.',
                false
            );
            publishState(
                'granted',
                subscription.toJSON(),
                'Notifications activées'
            );
        } else if (Notification.permission === 'denied') {
            button.textContent = '🔕 Notifications bloquées';
            setStatus(
                'Autorisez les notifications dans les réglages du navigateur.',
                false
            );
            publishState('denied', null, 'Autorisation refusée');
        }
    };

    button.onclick = async () => {
        try {
            const isIos = /iphone|ipad|ipod/i.test(navigator.userAgent);
            const isStandalone = window.matchMedia(
                '(display-mode: standalone)'
            ).matches || window.navigator.standalone === true;
            if (isIos && !isStandalone) {
                setStatus(
                    'Sur iPhone, ajoutez d’abord l’application à l’écran d’accueil, puis ouvrez-la depuis son icône.'
                );
                publishState(
                    'ios_home_screen_required',
                    null,
                    'Ajout à l’écran d’accueil requis'
                );
                return;
            }

            const permission = await Notification.requestPermission();
            if (permission !== 'granted') {
                button.textContent = '🔕 Notifications non autorisées';
                setStatus(
                    'Vous pourrez les autoriser depuis les réglages du navigateur.',
                    false
                );
                publishState(permission, null, 'Autorisation non accordée');
                return;
            }

            setStatus('Activation en cours…', false);
            const registration = await getRegistration();
            let subscription = await registration.pushManager.getSubscription();
            if (!subscription) {
                subscription = await registration.pushManager.subscribe({
                    userVisibleOnly: true,
                    applicationServerKey: toUint8Array(publicKey),
                });
            }
            button.textContent = '✅ Notifications mobiles activées';
            setStatus(
                'Ce téléphone recevra les nouveaux messages du chat.',
                false
            );
            publishState(
                'granted',
                subscription.toJSON(),
                'Notifications activées'
            );
        } catch (error) {
            button.textContent = '🔔 Réessayer l’activation';
            const errorMessage = error?.message || String(error);
            setStatus(
                'Activation impossible (module v4) : ' + errorMessage
            );
            publishState('error', null, errorMessage);
        }
    };

    inspectSubscription().catch((error) => {
        setStatus('Impossible de vérifier les notifications sur ce téléphone.');
        publishState('error', null, String(error));
    });

    return () => {
        button.onclick = null;
    };
}
"""


chat_push_component = st.components.v2.component(
    f"pro_consulting_chat_push_{CHAT_PUSH_FRONTEND_VERSION}",
    html=CHAT_PUSH_COMPONENT_HTML,
    css=CHAT_PUSH_COMPONENT_CSS,
    js=CHAT_PUSH_COMPONENT_JS,
)


def base64url_without_padding(value):
    """Encode les clés Web Push au format URL-safe attendu."""
    return base64.urlsafe_b64encode(value).rstrip(b"=").decode("ascii")


@st.cache_data(ttl=3600, show_spinner=False)
def get_or_create_web_push_keys(database_url):
    """Crée une seule paire VAPID et la conserve uniquement côté serveur."""
    initialize_settings_database(database_url)
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT payload
                FROM pro_consulting_settings
                WHERE scope = %s
                """,
                (WEB_PUSH_KEYS_SCOPE,),
            )
            row = cursor.fetchone()
            if row is None:
                private_key = ec.generate_private_key(ec.SECP256R1())
                private_der = private_key.private_bytes(
                    encoding=serialization.Encoding.DER,
                    format=serialization.PrivateFormat.PKCS8,
                    encryption_algorithm=serialization.NoEncryption(),
                )
                public_point = private_key.public_key().public_bytes(
                    encoding=serialization.Encoding.X962,
                    format=serialization.PublicFormat.UncompressedPoint,
                )
                generated_payload = {
                    "private_key": base64url_without_padding(private_der),
                    "public_key": base64url_without_padding(public_point),
                }
                cursor.execute(
                    """
                    INSERT INTO pro_consulting_settings (
                        scope,
                        payload,
                        updated_by
                    )
                    VALUES (%s, %s::jsonb, 'system:web_push')
                    ON CONFLICT (scope) DO NOTHING
                    """,
                    (
                        WEB_PUSH_KEYS_SCOPE,
                        json.dumps(generated_payload),
                    ),
                )
                connection.commit()
                cursor.execute(
                    """
                    SELECT payload
                    FROM pro_consulting_settings
                    WHERE scope = %s
                    """,
                    (WEB_PUSH_KEYS_SCOPE,),
                )
                row = cursor.fetchone()

    if row is None:
        raise RuntimeError("Les clés de notification n’ont pas pu être créées.")
    payload = row[0] if isinstance(row[0], dict) else json.loads(row[0])
    private_key = str(payload.get("private_key", "")).strip()
    public_key = str(payload.get("public_key", "")).strip()
    if not private_key or not public_key:
        raise RuntimeError("Les clés de notification enregistrées sont invalides.")
    return {"private_key": private_key, "public_key": public_key}


def save_chat_push_subscription(database_url, user_email, subscription):
    """Enregistre ou réactive l’abonnement du navigateur connecté."""
    if not isinstance(subscription, dict):
        raise ValueError("Abonnement mobile invalide.")
    endpoint = str(subscription.get("endpoint", "")).strip()
    subscription_keys = subscription.get("keys", {})
    p256dh = str(subscription_keys.get("p256dh", "")).strip()
    auth = str(subscription_keys.get("auth", "")).strip()
    if not endpoint.startswith("https://") or not p256dh or not auth:
        raise ValueError("Abonnement mobile incomplet.")

    cleaned_subscription = {
        "endpoint": endpoint,
        "expirationTime": subscription.get("expirationTime"),
        "keys": {"p256dh": p256dh, "auth": auth},
    }
    initialize_settings_database(database_url)
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO pro_consulting_push_subscriptions (
                    endpoint,
                    user_email,
                    subscription,
                    active
                )
                VALUES (%s, %s, %s::jsonb, TRUE)
                ON CONFLICT (endpoint)
                DO UPDATE SET
                    user_email = EXCLUDED.user_email,
                    subscription = EXCLUDED.subscription,
                    active = TRUE,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (
                    endpoint,
                    normalize_email(user_email),
                    json.dumps(cleaned_subscription),
                ),
            )
        connection.commit()
    return endpoint


def load_chat_notification_summary(database_url, user_email):
    """Retourne le nombre de messages collectifs non encore consultés."""
    normalized_email = normalize_email(user_email)
    initialize_settings_database(database_url)
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT last_seen_message_id
                FROM pro_consulting_chat_reads
                WHERE user_email = %s
                """,
                (normalized_email,),
            )
            read_row = cursor.fetchone()
            last_seen_message_id = int(read_row[0]) if read_row else 0
            cursor.execute(
                """
                SELECT
                    COUNT(*),
                    COALESCE(MAX(id), 0)
                FROM pro_consulting_chat_messages
                WHERE channel = 'general'
                  AND expires_at > CURRENT_TIMESTAMP
                  AND id > %s
                  AND author_email <> %s
                """,
                (last_seen_message_id, normalized_email),
            )
            unread_count, latest_unread_id = cursor.fetchone()
            latest_message = None
            if int(unread_count):
                cursor.execute(
                    """
                    SELECT author_name, message
                    FROM pro_consulting_chat_messages
                    WHERE id = %s
                    """,
                    (int(latest_unread_id),),
                )
                message_row = cursor.fetchone()
                if message_row:
                    latest_message = {
                        "author_name": str(message_row[0]),
                        "message": str(message_row[1]),
                    }

    return {
        "unread_count": int(unread_count),
        "latest_unread_id": int(latest_unread_id),
        "latest_message": latest_message,
    }


def mark_collective_chat_read(database_url, user_email, message_id):
    """Mémorise le dernier message affiché pour ce collaborateur."""
    if not message_id:
        return
    initialize_settings_database(database_url)
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO pro_consulting_chat_reads (
                    user_email,
                    last_seen_message_id
                )
                VALUES (%s, %s)
                ON CONFLICT (user_email)
                DO UPDATE SET
                    last_seen_message_id = GREATEST(
                        pro_consulting_chat_reads.last_seen_message_id,
                        EXCLUDED.last_seen_message_id
                    ),
                    updated_at = CURRENT_TIMESTAMP
                """,
                (normalize_email(user_email), int(message_id)),
            )
        connection.commit()


def deactivate_chat_push_subscription(database_url, endpoint):
    """Désactive une destination refusée définitivement par le navigateur."""
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                UPDATE pro_consulting_push_subscriptions
                SET active = FALSE,
                    updated_at = CURRENT_TIMESTAMP
                WHERE endpoint = %s
                """,
                (endpoint,),
            )
        connection.commit()


def send_collective_chat_push_notifications(
    database_url,
    author_email,
    author_name,
    message,
    message_id,
):
    """Envoie le nouveau message aux téléphones abonnés, sauf à l’auteur."""
    web_push_keys = get_or_create_web_push_keys(database_url)
    normalized_author = normalize_email(author_email)
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT endpoint, subscription
                FROM pro_consulting_push_subscriptions
                WHERE active = TRUE
                  AND user_email <> %s
                """,
                (normalized_author,),
            )
            subscriptions = cursor.fetchall()

    cleaned_message = " ".join(str(message or "").split())
    if len(cleaned_message) > 160:
        cleaned_message = cleaned_message[:157].rstrip() + "…"
    payload = json.dumps(
        {
            "title": f"Nouveau message de {author_name}",
            "body": cleaned_message,
            "url": f"/?page=chat&message={int(message_id)}",
            "tag": "pro-consulting-chat",
        },
        ensure_ascii=False,
    )
    def send_to_subscription(endpoint, subscription):
        subscription_payload = (
            subscription
            if isinstance(subscription, dict)
            else json.loads(subscription)
        )
        try:
            webpush(
                subscription_info=subscription_payload,
                data=payload,
                vapid_private_key=web_push_keys["private_key"],
                vapid_claims={
                    "sub": "mailto:tomeventfrance@gmail.com",
                },
                timeout=6,
            )
            return True
        except WebPushException as error:
            response = getattr(error, "response", None)
            if response is not None and response.status_code in (404, 410):
                deactivate_chat_push_subscription(database_url, endpoint)
        except Exception:
            pass
        return False

    sent_count = 0
    if subscriptions:
        with ThreadPoolExecutor(
            max_workers=min(8, len(subscriptions))
        ) as executor:
            futures = [
                executor.submit(send_to_subscription, endpoint, subscription)
                for endpoint, subscription in subscriptions
            ]
            for future in as_completed(futures):
                try:
                    sent_count += int(bool(future.result()))
                except Exception:
                    continue
    return sent_count


def load_collective_chat_messages(database_url, limit=250):
    """Charge uniquement les messages collectifs encore actifs."""
    initialize_settings_database(database_url)
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    id,
                    author_email,
                    author_name,
                    author_role,
                    message,
                    created_at,
                    expires_at
                FROM pro_consulting_chat_messages
                WHERE channel = 'general'
                  AND expires_at > CURRENT_TIMESTAMP
                ORDER BY created_at DESC, id DESC
                LIMIT %s
                """,
                (max(1, min(int(limit), 500)),),
            )
            rows = cursor.fetchall()

    return [
        {
            "id": row[0],
            "author_email": row[1],
            "author_name": row[2],
            "author_role": row[3],
            "message": row[4],
            "created_at": row[5],
            "expires_at": row[6],
        }
        for row in reversed(rows)
    ]


def save_collective_chat_message(
    database_url,
    author_email,
    author_name,
    author_role,
    message,
):
    """Enregistre un message collectif avec une expiration serveur à 48 h."""
    cleaned_message = str(message or "").strip()
    if not cleaned_message:
        raise ValueError("Le message est vide.")
    if len(cleaned_message) > CHAT_MAX_MESSAGE_LENGTH:
        raise ValueError(
            f"Le message dépasse {CHAT_MAX_MESSAGE_LENGTH} caractères."
        )

    initialize_settings_database(database_url)
    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                DELETE FROM pro_consulting_chat_messages
                WHERE expires_at <= CURRENT_TIMESTAMP
                """
            )
            cursor.execute(
                """
                INSERT INTO pro_consulting_chat_messages (
                    channel,
                    author_email,
                    author_name,
                    author_role,
                    message,
                    expires_at
                )
                VALUES (
                    'general',
                    %s,
                    %s,
                    %s,
                    %s,
                    CURRENT_TIMESTAMP + INTERVAL '48 hours'
                )
                RETURNING id
                """,
                (
                    normalize_email(author_email),
                    str(author_name or "Utilisateur").strip(),
                    str(author_role or "collaborator").strip(),
                    cleaned_message,
                ),
            )
            inserted_message_id = cursor.fetchone()[0]
        connection.commit()
    return int(inserted_message_id)


def format_chat_datetime(value):
    """Affiche les horaires du chat dans le fuseau français."""
    if not isinstance(value, datetime):
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=ZoneInfo("UTC"))
    return value.astimezone(ZoneInfo("Europe/Paris")).strftime(
        "%d/%m/%Y à %H:%M"
    )


PERSISTED_BACKSTAGE_COLUMNS = [
    "Pseudo",
    "Groupe",
    "Agent",
    "Diamants",
    "Heures LIVE",
    "Jours valides",
    "Statut évolution",
    "Statut échelon",
]


def serialize_backstage_dataframe(dataframe):
    """Sérialise le tableau préparé dans un format typé et portable."""
    return dataframe[PERSISTED_BACKSTAGE_COLUMNS].to_json(
        orient="table",
        index=False,
        date_format="iso",
        force_ascii=False,
    )


def deserialize_backstage_dataframe(serialized_dataframe):
    """Restaure et revalide les types du dernier import Backstage."""
    if not isinstance(serialized_dataframe, str) or not serialized_dataframe:
        raise ValueError("Le fichier Backstage enregistré est vide.")

    dataframe = pd.read_json(
        StringIO(serialized_dataframe),
        orient="table",
    )
    missing_columns = [
        column
        for column in PERSISTED_BACKSTAGE_COLUMNS
        if column not in dataframe.columns
    ]
    if missing_columns:
        raise ValueError(
            "Colonnes absentes de l'import enregistré : "
            + ", ".join(missing_columns)
        )

    dataframe = dataframe[PERSISTED_BACKSTAGE_COLUMNS].copy()
    for column in (
        "Pseudo",
        "Groupe",
        "Agent",
        "Statut évolution",
        "Statut échelon",
    ):
        dataframe[column] = dataframe[column].fillna("").astype(str).str.strip()

    dataframe["Diamants"] = pd.to_numeric(
        dataframe["Diamants"],
        errors="coerce",
    ).fillna(0).round(0).astype(int)
    dataframe["Heures LIVE"] = pd.to_numeric(
        dataframe["Heures LIVE"],
        errors="coerce",
    ).fillna(0.0).round(2)
    dataframe["Jours valides"] = pd.to_numeric(
        dataframe["Jours valides"],
        errors="coerce",
    ).fillna(0).round(0).astype(int)
    return dataframe.reset_index(drop=True)


def apply_persistent_settings(database_url, user_email):
    """Recharge les réglages communs et ceux de la direction connectée."""
    initialize_settings_database(database_url)

    scope_names = (
        "global:financial",
        "global:exchange_rate",
        branch_settings_scope(user_email),
        exclusions_scope(user_email),
    )
    loaded_scopes = load_persistent_scopes(database_url, scope_names)
    global_settings = loaded_scopes["global:financial"]
    exchange_rate_settings = loaded_scopes["global:exchange_rate"]
    branch_settings = loaded_scopes[branch_settings_scope(user_email)]
    exclusions_settings = loaded_scopes[exclusions_scope(user_email)]

    for key in GLOBAL_FINANCIAL_KEYS:
        if key in global_settings:
            st.session_state[key] = global_settings[key]

    for key in GLOBAL_EXCHANGE_RATE_KEYS:
        if key in exchange_rate_settings:
            st.session_state[key] = exchange_rate_settings[key]

    for key in BRANCH_PARAMETER_KEYS:
        if key in branch_settings:
            st.session_state[key] = branch_settings[key]

    # Une absence de liste persistante doit vider les exclusions afin
    # qu'un changement de compte ne puisse jamais réutiliser celles d'une
    # autre direction.
    st.session_state.exclusions = clean_exclusions(
        exclusions_settings.get("rows", [])
    )


def normalize_email(value):
    return str(value or "").strip().lower()


def show_estimation_notice():
    st.caption(
        "* Les rémunérations affichées sont des estimations calculées "
        "à partir des données importées et des paramètres enregistrés. "
        "Elles peuvent différer des montants définitifs réellement versés."
    )


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_ecb_usd_to_eur_rate():
    url = (
        "https://www.ecb.europa.eu/stats/eurofxref/"
        "eurofxref-daily.xml"
    )
    request = Request(url, headers={"User-Agent": "ProConsulting/1.0"})
    with urlopen(request, timeout=10) as response:
        xml_content = response.read()

    root = ElementTree.fromstring(xml_content)
    rate_date = None
    usd_per_eur = None

    for element in root.iter():
        if "time" in element.attrib:
            rate_date = element.attrib["time"]
        if element.attrib.get("currency") == "USD":
            usd_per_eur = float(element.attrib["rate"])

    if not rate_date or not usd_per_eur:
        raise ValueError("Le taux USD de la BCE est introuvable.")

    return 1 / usd_per_eur, rate_date


def resolve_daily_usd_to_eur_rate(
    database_url=None,
    updated_by=None,
    force_refresh=False,
):
    """Retourne le taux BCE quotidien et mémorise un secours partagé."""
    if force_refresh:
        fetch_ecb_usd_to_eur_rate.clear()

    previous_rate_date = st.session_state.get("last_ecb_rate_date")

    try:
        usd_to_eur, rate_date = fetch_ecb_usd_to_eur_rate()
        st.session_state.usd_to_eur = float(usd_to_eur)
        st.session_state.last_ecb_usd_to_eur = float(usd_to_eur)
        st.session_state.last_ecb_rate_date = rate_date

        persistence_warning = None
        if database_url and rate_date != previous_rate_date:
            try:
                save_persistent_scopes(
                    database_url,
                    {
                        "global:exchange_rate": {
                            "usd_to_eur": float(usd_to_eur),
                            "last_ecb_usd_to_eur": float(usd_to_eur),
                            "last_ecb_rate_date": rate_date,
                        }
                    },
                    updated_by or "automatic-rate-update",
                )
            except Exception:
                persistence_warning = (
                    "Le taux BCE est utilisable, mais sa copie de secours "
                    "n’a pas pu être enregistrée dans la base."
                )

        return {
            "rate": float(usd_to_eur),
            "date": rate_date,
            "source": "Banque centrale européenne",
            "warning": persistence_warning,
        }
    except Exception:
        remembered_rate = st.session_state.get("last_ecb_usd_to_eur")
        remembered_date = st.session_state.get("last_ecb_rate_date")

        if remembered_rate is not None:
            return {
                "rate": float(remembered_rate),
                "date": remembered_date or "date inconnue",
                "source": "Dernier taux BCE mémorisé",
                "warning": (
                    "La BCE est momentanément inaccessible. Le dernier "
                    "taux quotidien mémorisé est utilisé."
                ),
            }

        return {
            "rate": float(st.session_state.usd_to_eur),
            "date": "date inconnue",
            "source": "Taux de secours enregistré",
            "warning": (
                "La BCE est momentanément inaccessible et aucun taux BCE "
                "mémorisé n’est disponible."
            ),
        }


def clean_exclusions(rows):
    cleaned = []
    seen = set()

    for row in rows:
        email = normalize_email(row.get("Adresse e-mail", ""))
        if not email or email in seen:
            continue

        seen.add(email)
        cleaned.append(
            {
                "Adresse e-mail": email,
                "Exclure consultants": bool(
                    row.get("Exclure consultants", True)
                ),
                "Exclure responsables performance": bool(
                    row.get("Exclure responsables performance", True)
                ),
            }
        )

    return cleaned


REWARD_TRACKING_COLUMNS = [
    "Date",
    "Heure",
    "Type d’événement",
    "Récompense validée",
    "Récompense refusée",
    "Créateur",
    "Groupe",
    "Récompense créateur",
    "Rémunération consultant / responsable",
    "Total récompense",
]


def clean_reward_tracking_rows(rows):
    """Valide le suivi avant affichage, calcul ou sauvegarde."""
    cleaned_rows = []
    rows = rows if isinstance(rows, list) else []

    def clean_text(value):
        text = str(value or "").strip()
        return "" if text.lower() == "nan" else text

    def clean_reward(value):
        try:
            return max(0, int(round(float(value or 0))))
        except (TypeError, ValueError):
            return 0

    def clean_checkbox(value):
        if isinstance(value, bool):
            return value
        return str(value or "").strip().casefold() in {
            "true",
            "1",
            "oui",
            "yes",
        }

    for row in rows:
        if not isinstance(row, dict):
            continue

        creator = clean_text(row.get("Créateur", ""))
        if not creator:
            continue

        event_type = clean_text(row.get("Type d’événement", ""))
        if event_type not in {"Live", "Match"}:
            event_type = ""

        creator_reward = clean_reward(
            row.get("Récompense créateur", 0)
        )
        hierarchy_reward = clean_reward(
            row.get("Rémunération consultant / responsable", 0)
        )
        reward_is_validated = clean_checkbox(
            row.get(
                "Récompense validée",
                row.get("Récompense envoyée", False),
            )
        )
        reward_is_refused = clean_checkbox(
            row.get("Récompense refusée", False)
        )
        if reward_is_validated and reward_is_refused:
            reward_is_validated = False
            reward_is_refused = False
        cleaned_rows.append(
            {
                "Date": clean_text(row.get("Date", "")),
                "Heure": clean_text(row.get("Heure", "")),
                "Type d’événement": event_type,
                "Récompense validée": reward_is_validated,
                "Récompense refusée": reward_is_refused,
                "Créateur": creator,
                "Groupe": clean_text(row.get("Groupe", "")),
                "Récompense créateur": creator_reward,
                "Rémunération consultant / responsable": hierarchy_reward,
                "Total récompense": creator_reward + hierarchy_reward,
            }
        )

    return cleaned_rows


def build_reward_tracking_table(creator_results, saved_rows=None):
    """Fusionne les créateurs calculés avec les champs manuels sauvegardés."""
    saved_by_creator = {}
    for saved_row in clean_reward_tracking_rows(saved_rows or []):
        saved_by_creator.setdefault(saved_row["Créateur"], []).append(
            saved_row
        )

    tracking_rows = []
    for _, creator_row in creator_results.iterrows():
        creator = str(creator_row.get("Pseudo", "") or "").strip()
        if not creator or creator.lower() == "nan":
            continue

        saved_candidates = saved_by_creator.get(creator, [])
        saved_row = saved_candidates.pop(0) if saved_candidates else {}
        payment_mode = str(
            creator_row.get("Mode paiement", "Diamants") or "Diamants"
        ).strip()
        creator_group = str(
            creator_row.get("Groupe", saved_row.get("Groupe", "")) or ""
        ).strip()
        if payment_mode == "Facture €":
            creator_reward = 0
        else:
            try:
                creator_reward = max(
                    0,
                    int(
                        round(
                            float(creator_row.get("Rémunération 💎", 0))
                        )
                    ),
                )
            except (TypeError, ValueError):
                creator_reward = 0

        hierarchy_reward = clean_reward_tracking_rows(
            [
                {
                    "Créateur": creator,
                    "Rémunération consultant / responsable": saved_row.get(
                        "Rémunération consultant / responsable",
                        0,
                    ),
                }
            ]
        )[0]["Rémunération consultant / responsable"]

        tracking_rows.append(
            {
                "Date": saved_row.get("Date", ""),
                "Heure": saved_row.get("Heure", ""),
                "Type d’événement": saved_row.get(
                    "Type d’événement",
                    "",
                ),
                "Récompense validée": bool(
                    saved_row.get("Récompense validée", False)
                ),
                "Récompense refusée": bool(
                    saved_row.get("Récompense refusée", False)
                ),
                "Créateur": creator,
                "Groupe": creator_group,
                "Récompense créateur": creator_reward,
                "Rémunération consultant / responsable": hierarchy_reward,
                "Total récompense": creator_reward + hierarchy_reward,
            }
        )

    # Les lignes des autres directions restent visibles dans le tableau
    # collectif même si leur créateur n’est pas présent dans l’import local.
    for remaining_rows in saved_by_creator.values():
        tracking_rows.extend(remaining_rows)

    return pd.DataFrame(tracking_rows, columns=REWARD_TRACKING_COLUMNS)


def reward_creator_key(value):
    """Construit une clé stable sans modifier le pseudo affiché."""
    return " ".join(str(value or "").strip().casefold().split())


def merge_collective_reward_tracking_rows(
    remote_rows,
    local_rows,
    baseline_rows,
    local_creators,
    can_update_reward_status=False,
    editable_groups=None,
    allow_new_rows=True,
):
    """Fusionne une sauvegarde sans effacer le travail d'un autre compte."""
    remote_rows = clean_reward_tracking_rows(remote_rows)
    local_rows = clean_reward_tracking_rows(local_rows)
    baseline_rows = clean_reward_tracking_rows(baseline_rows)

    def rows_by_creator(rows):
        return {
            reward_creator_key(row["Créateur"]): row
            for row in rows
            if reward_creator_key(row["Créateur"])
        }

    remote_by_creator = rows_by_creator(remote_rows)
    local_by_creator = rows_by_creator(local_rows)
    baseline_by_creator = rows_by_creator(baseline_rows)
    local_creator_keys = {
        reward_creator_key(creator)
        for creator in local_creators
        if reward_creator_key(creator)
    }
    editable_group_keys = None
    if editable_groups is not None:
        editable_group_keys = {
            normalize_group_name(group)
            for group in editable_groups
            if normalize_group_name(group)
        }

    ordered_keys = list(remote_by_creator)
    ordered_keys.extend(
        key for key in local_by_creator if key not in remote_by_creator
    )
    manual_columns = (
        "Date",
        "Heure",
        "Type d’événement",
        "Rémunération consultant / responsable",
    )
    if can_update_reward_status:
        manual_columns += (
            "Récompense validée",
            "Récompense refusée",
        )
    merged_rows = []

    for creator_key in ordered_keys:
        remote_row = remote_by_creator.get(creator_key)
        local_row = local_by_creator.get(creator_key)
        baseline_row = baseline_by_creator.get(creator_key)

        if local_row is None:
            merged_rows.append(remote_row)
            continue
        if remote_row is None:
            if allow_new_rows:
                merged_rows.append(local_row)
            continue

        row_is_editable = (
            editable_group_keys is None
            or normalize_group_name(remote_row.get("Groupe", ""))
            in editable_group_keys
        )
        if not row_is_editable:
            merged_rows.append(remote_row)
            continue

        merged_row = dict(remote_row)
        if creator_key in local_creator_keys:
            merged_row["Groupe"] = local_row.get(
                "Groupe",
                remote_row.get("Groupe", ""),
            )
            merged_row["Récompense créateur"] = local_row[
                "Récompense créateur"
            ]

        for column in manual_columns:
            local_value = local_row.get(column)
            baseline_value = (
                baseline_row.get(column)
                if baseline_row is not None
                else None
            )
            if baseline_row is None or local_value != baseline_value:
                merged_row[column] = local_value

        merged_row["Total récompense"] = (
            merged_row["Récompense créateur"]
            + merged_row["Rémunération consultant / responsable"]
        )
        merged_rows.append(merged_row)

    return clean_reward_tracking_rows(merged_rows)


def save_collective_reward_tracking(
    database_url,
    scope,
    local_rows,
    baseline_rows,
    local_creators,
    requested_month,
    updated_by,
    can_update_reward_status=False,
    editable_groups=None,
    allow_new_rows=True,
):
    """Fusionne et confirme une sauvegarde collective dans une transaction.

    Le verrou PostgreSQL empeche deux sessions ouvertes en meme temps de
    s'ecraser. Lors d'un changement de mois, le registre precedent est
    archive avant que le nouveau mois ne soit cree sans ses champs manuels.
    """
    requested_month = str(requested_month or "").strip()
    updated_by = normalize_email(updated_by)

    with psycopg.connect(database_url) as connection:
        with connection.cursor() as cursor:
            # Ce verrou fonctionne meme si la ligne du registre n'existe pas
            # encore, contrairement a un simple SELECT ... FOR UPDATE.
            cursor.execute(
                "SELECT pg_advisory_xact_lock(hashtext(%s))",
                (scope,),
            )
            cursor.execute(
                """
                SELECT payload
                FROM pro_consulting_settings
                WHERE scope = %s
                FOR UPDATE
                """,
                (scope,),
            )
            row = cursor.fetchone()
            remote_payload = row[0] if row else {}
            if isinstance(remote_payload, str):
                remote_payload = json.loads(remote_payload)
            remote_payload = dict(remote_payload or {})

            remote_month = str(remote_payload.get("month") or "").strip()
            starts_new_month = bool(
                remote_payload
                and requested_month
                and remote_month
                and remote_month.casefold() != requested_month.casefold()
            )

            if starts_new_month:
                archive_scope = reward_tracking_archive_scope(remote_month)
                cursor.execute(
                    """
                    INSERT INTO pro_consulting_settings (
                        scope,
                        payload,
                        updated_by
                    )
                    VALUES (%s, %s::jsonb, %s)
                    ON CONFLICT (scope)
                    DO UPDATE SET
                        payload = EXCLUDED.payload,
                        updated_at = CURRENT_TIMESTAMP,
                        updated_by = EXCLUDED.updated_by
                    """,
                    (
                        archive_scope,
                        json.dumps(remote_payload, ensure_ascii=False),
                        updated_by,
                    ),
                )
                remote_rows = []
                effective_baseline_rows = []
            else:
                remote_rows = remote_payload.get("rows", [])
                effective_baseline_rows = baseline_rows

            merged_rows = merge_collective_reward_tracking_rows(
                remote_rows=remote_rows,
                local_rows=local_rows,
                baseline_rows=effective_baseline_rows,
                local_creators=local_creators,
                can_update_reward_status=can_update_reward_status,
                editable_groups=editable_groups,
                allow_new_rows=allow_new_rows,
            )
            saved_moment = datetime.now(ZoneInfo("Europe/Paris"))
            saved_at = saved_moment.isoformat(timespec="seconds")
            save_token = hashlib.sha256(
                (
                    f"{scope}|{updated_by}|{saved_moment.isoformat()}|"
                    f"{len(merged_rows)}"
                ).encode("utf-8")
            ).hexdigest()[:20]
            saved_payload = {
                "month": requested_month or remote_month,
                "rows": merged_rows,
                "saved_at": saved_at,
                "saved_by": updated_by,
                "save_token": save_token,
            }
            cursor.execute(
                """
                INSERT INTO pro_consulting_settings (
                    scope,
                    payload,
                    updated_by
                )
                VALUES (%s, %s::jsonb, %s)
                ON CONFLICT (scope)
                DO UPDATE SET
                    payload = EXCLUDED.payload,
                    updated_at = CURRENT_TIMESTAMP,
                    updated_by = EXCLUDED.updated_by
                RETURNING payload
                """,
                (
                    scope,
                    json.dumps(saved_payload, ensure_ascii=False),
                    updated_by,
                ),
            )
            confirmed_payload = cursor.fetchone()[0]
            if isinstance(confirmed_payload, str):
                confirmed_payload = json.loads(confirmed_payload)
            confirmed_payload = dict(confirmed_payload or {})
            if (
                confirmed_payload.get("save_token") != save_token
                or normalize_email(confirmed_payload.get("saved_by"))
                != updated_by
            ):
                raise RuntimeError(
                    "La base n'a pas confirme la nouvelle version du suivi."
                )
        connection.commit()

    # Une seconde connexion prouve que la transaction est effectivement
    # visible pour les autres comptes, pas seulement dans la session courante.
    persisted_payload = load_persistent_scope(database_url, scope)
    if (
        persisted_payload.get("save_token") != save_token
        or normalize_email(persisted_payload.get("saved_by")) != updated_by
    ):
        raise RuntimeError(
            "La sauvegarde n'est pas visible depuis une nouvelle connexion."
        )

    return persisted_payload, starts_new_month


def style_reward_status_rows(row):
    """Colore toute la ligne selon la décision administrateur."""
    if bool(row.get("Récompense refusée", False)):
        return [
            (
                "background-color: #f4b7b7; color: #6b0f1a; "
                "font-weight: 800; border-color: #d6455d"
            )
            for _ in row
        ]
    if bool(row.get("Récompense validée", False)):
        return [
            (
                "background-color: #b7f7d0; color: #073b24; "
                "font-weight: 800; border-color: #22a06b"
            )
            for _ in row
        ]
    return ["" for _ in row]


def synchronize_reward_tracking_editor():
    """Recalcule le total dès qu’une cellule manuelle est modifiée."""
    current_table = st.session_state.get("reward_tracking_table")
    editor_state = st.session_state.get("reward_tracking_editor", {})
    if not isinstance(current_table, pd.DataFrame):
        return
    if not isinstance(editor_state, dict):
        return

    updated_table = current_table.copy().reset_index(drop=True)
    reset_editor_after_status_change = False
    editable_columns = {
        "Date",
        "Heure",
        "Type d’événement",
        "Rémunération consultant / responsable",
    }
    if globals().get("current_user_role") == "admin":
        editable_columns.update(
            {"Récompense validée", "Récompense refusée"}
        )
    for row_index, changes in editor_state.get("edited_rows", {}).items():
        try:
            row_number = int(row_index)
        except (TypeError, ValueError):
            continue
        if row_number < 0 or row_number >= len(updated_table):
            continue
        status_columns = {
            "Récompense validée",
            "Récompense refusée",
        }
        changed_status_columns = [
            column
            for column in status_columns
            if column in changes
            and bool(changes[column])
            != bool(updated_table.at[row_number, column])
        ]
        if changed_status_columns:
            reset_editor_after_status_change = True
        for column, value in changes.items():
            if column in editable_columns and column not in status_columns:
                updated_table.at[row_number, column] = value
        if globals().get("current_user_role") == "admin":
            activated_statuses = [
                column
                for column in changed_status_columns
                if bool(changes[column])
            ]
            if len(activated_statuses) == 1:
                activated_status = activated_statuses[0]
                other_status = (
                    "Récompense refusée"
                    if activated_status == "Récompense validée"
                    else "Récompense validée"
                )
                updated_table.at[row_number, activated_status] = True
                updated_table.at[row_number, other_status] = False
            elif not activated_statuses:
                for status_column in changed_status_columns:
                    updated_table.at[row_number, status_column] = False

    st.session_state.reward_tracking_table = pd.DataFrame(
        clean_reward_tracking_rows(updated_table.to_dict("records")),
        columns=REWARD_TRACKING_COLUMNS,
    )
    if reset_editor_after_status_change:
        st.session_state.pop("reward_tracking_editor", None)


def excluded_emails(scope):
    column = {
        "consultants": "Exclure consultants",
        "responsables": "Exclure responsables performance",
    }[scope]

    return {
        normalize_email(row.get("Adresse e-mail", ""))
        for row in st.session_state.exclusions
        if row.get(column, False)
    }


def reset_calculations():
    for key in (
        "creator_results",
        "creator_signature",
        "consultant_results",
        "consultant_signature",
        "responsable_results",
        "responsable_signature",
        "creator_payment_editor",
        "consultant_payment_editor",
        "responsable_payment_editor",
        "reward_tracking_table",
        "reward_tracking_signature",
        "reward_tracking_editor",
        "manager_reward_tracking_editor",
        "reward_tracking_loaded_scope",
        "reward_tracking_baseline_rows",
        "reward_tracking_active_month",
        "reward_tracking_last_saved_at",
        "reward_tracking_last_saved_by",
        "reward_tracking_pending_new_month",
        "reward_tracking_save_notice",
    ):
        st.session_state.pop(key, None)


def clear_backstage_session():
    """Retire l'import du compte courant sans toucher aux autres comptes."""
    st.session_state.backstage_data = None
    st.session_state.backstage_filename = None
    st.session_state.pop("backstage_raw_data", None)
    st.session_state.pop("backstage_file_digest", None)
    st.session_state.pop("detected_columns", None)
    reset_calculations()


def financial_columns(dataframe):
    result = dataframe.copy()
    coin_price = float(st.session_state.coin_pack_price) / 1000
    invoice_rate = float(st.session_state.invoice_rate)

    result["Facture €"] = result.apply(
        lambda row: int(row["Rémunération 💎"] * invoice_rate)
        if row["Mode paiement"] == "Facture €"
        else 0,
        axis=1,
    )
    result["Coût diamants €"] = result.apply(
        lambda row: round(row["Rémunération 💎"] * coin_price, 2)
        if row["Mode paiement"] == "Diamants"
        else 0.0,
        axis=1,
    )
    result["Total déduction €"] = (
        result["Facture €"] + result["Coût diamants €"]
    )
    return result


def combined_payment_totals(*dataframes):
    """Additionne les paiements en diamants et en factures."""
    diamond_total = 0.0
    diamond_cost = 0.0
    invoice_total = 0.0

    for dataframe in dataframes:
        if dataframe is None or dataframe.empty:
            continue

        diamond_rows = dataframe["Mode paiement"] == "Diamants"
        diamond_total += float(
            dataframe.loc[diamond_rows, "Rémunération 💎"].sum()
        )
        diamond_cost += float(dataframe["Coût diamants €"].sum())
        invoice_total += float(dataframe["Facture €"].sum())

    return {
        "diamonds": diamond_total,
        "diamond_cost": diamond_cost,
        "invoices": invoice_total,
    }


RESPONSABLE_RATES = {
    5: 0.010,
    7: 0.015,
    9: 0.018,
    11: 0.020,
    13: 0.025,
}

DIRECTOR_RATES = {
    4: 0.04,
    5: 0.05,
    7: 0.07,
    8: 0.09,
    10: 0.11,
    11: 0.13,
    13: 0.15,
}


def calculate_agency_profit(
    revenue_usd,
    usd_to_eur,
    charges_rate,
    creator_cost,
    consultant_cost,
    responsable_cost,
    other_expenses,
    director_rate,
):
    """Calcule le bénéfice agence après toutes les déductions."""
    revenue_eur = float(revenue_usd) * float(usd_to_eur)
    charges_eur = revenue_eur * float(charges_rate) / 100
    result_before_directors = (
        revenue_eur
        - charges_eur
        - float(creator_cost)
        - float(consultant_cost)
        - float(responsable_cost)
        - float(other_expenses)
    )
    directors_reward = (
        max(0.0, result_before_directors) * float(director_rate)
    )

    return {
        "revenue_eur": revenue_eur,
        "charges_eur": charges_eur,
        "result_before_directors": result_before_directors,
        "directors_reward": directors_reward,
        "agency_profit": result_before_directors - directors_reward,
    }


def floor_to_hundred(value):
    return int(float(value) // 100 * 100)


def calculate_responsable_rewards(
    creator_results,
    responsable_level,
    minimum_group_diamonds=600_000,
):
    required_columns = {
        "Groupe",
        "Diamants",
        "Compté hiérarchie",
    }
    missing_columns = required_columns.difference(creator_results.columns)
    if missing_columns:
        raise ValueError(
            "Colonnes manquantes pour calculer les responsables : "
            + ", ".join(sorted(missing_columns))
        )

    data = creator_results.copy()
    data["Groupe"] = data["Groupe"].fillna("").astype(str).str.strip()
    data = data[data["Groupe"].ne("") & data["Groupe"].ne("nan")]

    rate = RESPONSABLE_RATES.get(int(responsable_level), 0.0)
    rows = []

    for responsable, group in data.groupby("Groupe", dropna=False):
        eligible = group[group["Compté hiérarchie"] == "Oui"]
        eligible_diamonds = int(eligible["Diamants"].sum())
        threshold_reached = eligible_diamonds >= minimum_group_diamonds
        reward = (
            floor_to_hundred(eligible_diamonds * rate)
            if threshold_reached
            else 0
        )

        rows.append(
            {
                "Responsable performance": responsable,
                "Créateurs rattachés": len(group),
                "Créateurs comptés": len(eligible),
                "Diamants éligibles": eligible_diamonds,
                "Seuil atteint": "Oui" if threshold_reached else "Non",
                "Taux": rate if threshold_reached else 0.0,
                "Rémunération 💎": reward,
            }
        )

    result = pd.DataFrame(rows)
    if not result.empty:
        result.sort_values(
            "Diamants éligibles",
            ascending=False,
            inplace=True,
        )
        result.reset_index(drop=True, inplace=True)
    return result


def calculate_director_branch_finances(
    all_creator_results,
    selected_groups,
    revenue_usd,
    other_expenses,
    usd_to_eur,
):
    """Calcule une branche sans modifier les résultats des créateurs."""
    branch_creators = all_creator_results[
        all_creator_results["Groupe"].isin(selected_groups)
    ].copy()
    if "Mode paiement" not in branch_creators.columns:
        branch_creators["Mode paiement"] = "Diamants"
    branch_creators = financial_columns(branch_creators)

    branch_consultants = calculate_consultant_rewards(
        creator_results=branch_creators,
        consultant_level=int(st.session_state.consultant_level),
        minimum_team_diamonds=200_000,
    )
    if not branch_consultants.empty:
        branch_consultants["Mode paiement"] = "Diamants"
        if "consultant_results" in st.session_state:
            saved_consultant_modes = (
                st.session_state.consultant_results
                .drop_duplicates("Consultant")
                .set_index("Consultant")["Mode paiement"]
                .to_dict()
            )
            branch_consultants["Mode paiement"] = branch_consultants[
                "Consultant"
            ].map(saved_consultant_modes).fillna("Diamants")
        consultant_exclusions = excluded_emails("consultants")
        consultant_excluded_mask = branch_consultants["Consultant"].map(
            lambda value: normalize_email(value) in consultant_exclusions
        )
        branch_consultants.loc[
            consultant_excluded_mask,
            ["Taux", "Rémunération 💎"],
        ] = 0
        branch_consultants = financial_columns(branch_consultants)

    branch_responsables = calculate_responsable_rewards(
        creator_results=branch_creators,
        responsable_level=int(st.session_state.manager_level),
        minimum_group_diamonds=600_000,
    )
    if not branch_responsables.empty:
        branch_responsables["Mode paiement"] = "Diamants"
        if "responsable_results" in st.session_state:
            saved_responsable_modes = (
                st.session_state.responsable_results
                .drop_duplicates("Responsable performance")
                .set_index("Responsable performance")["Mode paiement"]
                .to_dict()
            )
            branch_responsables["Mode paiement"] = branch_responsables[
                "Responsable performance"
            ].map(saved_responsable_modes).fillna("Diamants")
        responsable_exclusions = excluded_emails("responsables")
        responsable_excluded_mask = branch_responsables[
            "Responsable performance"
        ].map(lambda value: normalize_email(value) in responsable_exclusions)
        branch_responsables.loc[
            responsable_excluded_mask,
            ["Taux", "Rémunération 💎"],
        ] = 0
        branch_responsables = financial_columns(branch_responsables)

    revenue_eur = float(revenue_usd) * float(usd_to_eur)
    charges_eur = (
        revenue_eur * float(st.session_state.charges_rate) / 100
    )
    creator_cost = float(branch_creators["Total déduction €"].sum())
    consultant_cost = (
        float(branch_consultants["Total déduction €"].sum())
        if not branch_consultants.empty
        else 0.0
    )
    responsable_cost = (
        float(branch_responsables["Total déduction €"].sum())
        if not branch_responsables.empty
        else 0.0
    )
    net_profit_before_director = max(
        0.0,
        revenue_eur
        - charges_eur
        - creator_cost
        - consultant_cost
        - responsable_cost
        - float(other_expenses),
    )
    director_rate = DIRECTOR_RATES.get(
        int(st.session_state.director_level),
        0.0,
    )
    director_reward = net_profit_before_director * director_rate

    return {
        "creators": branch_creators,
        "consultants": branch_consultants,
        "responsables": branch_responsables,
        "diamonds_generated": float(branch_creators["Diamants"].sum()),
        "revenue_eur": revenue_eur,
        "charges_eur": charges_eur,
        "creator_cost": creator_cost,
        "consultant_cost": consultant_cost,
        "responsable_cost": responsable_cost,
        "net_profit_before_director": net_profit_before_director,
        "director_rate": director_rate,
        "director_reward": director_reward,
        "remaining_after_director": (
            net_profit_before_director - director_reward
        ),
    }


def build_director_overview(
    all_creator_results,
    director_config,
    usd_to_eur,
):
    """Construit la synthèse financière des quatre directions."""
    director_rows = []

    for profile in DIRECTOR_MANAGEMENT_PROFILES:
        email = profile["email"]
        direction_config = director_config[email]
        selected_groups = direction_config["groups"]

        if selected_groups:
            branch_result = calculate_director_branch_finances(
                all_creator_results=all_creator_results,
                selected_groups=selected_groups,
                revenue_usd=direction_config["revenue_usd"],
                other_expenses=direction_config["other_expenses"],
                usd_to_eur=usd_to_eur,
            )
        else:
            branch_result = {
                "diamonds_generated": 0.0,
                "revenue_eur": 0.0,
                "charges_eur": 0.0,
                "creator_cost": 0.0,
                "consultant_cost": 0.0,
                "responsable_cost": 0.0,
                "net_profit_before_director": 0.0,
                "director_rate": DIRECTOR_RATES.get(
                    int(st.session_state.director_level),
                    0.0,
                ),
                "director_reward": 0.0,
                "remaining_after_director": 0.0,
            }

        director_rows.append(
            {
                "Directeur": profile["name"],
                "Direction": profile["direction"],
                "Groupes gérés": ", ".join(selected_groups) or "Aucun",
                "Diamants générés": branch_result["diamonds_generated"],
                "CA branche ($)": direction_config["revenue_usd"],
                "CA converti (€)": branch_result["revenue_eur"],
                "Charges (€)": branch_result["charges_eur"],
                "Coût créateurs (€)": branch_result["creator_cost"],
                "Coût consultants (€)": branch_result["consultant_cost"],
                "Coût responsables (€)": branch_result[
                    "responsable_cost"
                ],
                "Autres dépenses (€)": direction_config[
                    "other_expenses"
                ],
                "Bénéfice avant directeur (€)": branch_result[
                    "net_profit_before_director"
                ],
                "Taux directeur": branch_result["director_rate"] * 100,
                "Montant facture directeur (€)": branch_result[
                    "director_reward"
                ],
                "Bénéfice restant (€)": branch_result[
                    "remaining_after_director"
                ],
            }
        )

    return pd.DataFrame(director_rows)


def show_financial_summary(dataframe):
    st.subheader("Récapitulatif financier")
    diamond_total = dataframe.loc[
        dataframe["Mode paiement"] == "Diamants",
        "Rémunération 💎",
    ].sum()
    col1, col2, col3, col4 = st.columns(4)
    col1.metric(
        "Total à payer en diamants",
        f"{diamond_total:,.0f} 💎",
    )
    col2.metric("Total factures", f"{dataframe['Facture €'].sum():,.0f} €")
    col3.metric(
        "Coût des diamants",
        f"{dataframe['Coût diamants €'].sum():,.2f} €",
    )
    col4.metric(
        "Déduction totale",
        f"{dataframe['Total déduction €'].sum():,.2f} €",
    )


BASE_AUTHORIZED_USERS = {
    "tomeventfrance@gmail.com": {
        "name": "FONDATEUR ADMIN",
        "role": "admin",
        "direction": "Administration générale",
    },
    "a.stone.authorbusiness@gmail.com": {
        "name": "Biker",
        "role": "director",
        "direction": "Direction Biker",
    },
    "melvynschmidt2013@gmail.com": {
        "name": "Max",
        "role": "director",
        "direction": "Direction Max",
    },
    "moon441330@gmail.com": {
        "name": "Moon",
        "role": "director",
        "direction": "Direction Moon",
    },
    "vividirectrice@gmail.com": {
        "name": "Vivi",
        "role": "director",
        "direction": "Direction Vivi",
    },
}

COLLABORATOR_ROLE_LABELS = {
    "director": "Directeur",
    "performance_manager": "Responsable performance",
}


def clean_collaborator_access_records(rows, available_groups=None):
    """Valide les accès créés par l'administrateur avant utilisation."""
    rows = rows if isinstance(rows, list) else []
    available_group_set = (
        {str(group).strip() for group in available_groups}
        if available_groups is not None
        else None
    )
    cleaned = []
    seen_emails = set(BASE_AUTHORIZED_USERS)

    for row in rows:
        if not isinstance(row, dict):
            continue
        email = normalize_email(row.get("email"))
        name = str(row.get("name") or "").strip()
        role = str(row.get("role") or "").strip()
        groups = row.get("groups", [])
        groups = groups if isinstance(groups, list) else []
        active = bool(row.get("active", True))

        if (
            not email
            or not name
            or role not in COLLABORATOR_ROLE_LABELS
            or email in seen_emails
        ):
            continue

        unique_groups = []
        seen_groups = set()
        for group in groups:
            cleaned_group = str(group or "").strip()
            group_key = normalize_group_name(cleaned_group)
            if not cleaned_group or group_key in seen_groups:
                continue
            if (
                available_group_set is not None
                and cleaned_group not in available_group_set
            ):
                continue
            seen_groups.add(group_key)
            unique_groups.append(cleaned_group)

        seen_emails.add(email)
        cleaned.append(
            {
                "email": email,
                "name": name,
                "role": role,
                "groups": unique_groups,
                "active": active,
            }
        )

    return cleaned


def build_authorized_users(database_url):
    """Réunit les comptes protégés et les collaborateurs actifs."""
    authorized_users = {
        email: dict(access)
        for email, access in BASE_AUTHORIZED_USERS.items()
    }
    access_error = None

    if not database_url:
        return authorized_users, access_error

    try:
        initialize_settings_database(database_url)
        authorization_payloads = load_authorization_payloads(database_url)
        director_payload = authorization_payloads[
            director_management_scope()
        ]
        director_config = director_payload.get("directors", {})
        if isinstance(director_config, dict):
            for email, saved_row in director_config.items():
                normalized_email = normalize_email(email)
                if (
                    normalized_email in authorized_users
                    and isinstance(saved_row, dict)
                ):
                    authorized_users[normalized_email]["groups"] = list(
                        saved_row.get("groups", [])
                    )

        collaborator_payload = authorization_payloads[
            collaborator_access_scope()
        ]
        collaborator_rows = clean_collaborator_access_records(
            collaborator_payload.get("collaborators", [])
        )
        for collaborator in collaborator_rows:
            if not collaborator["active"]:
                continue
            role = collaborator["role"]
            authorized_users[collaborator["email"]] = {
                "name": collaborator["name"],
                "role": role,
                "direction": (
                    f"Direction {collaborator['name']}"
                    if role == "director"
                    else f"Responsable performance • {collaborator['name']}"
                ),
                "groups": list(collaborator["groups"]),
                "managed_access": True,
            }
    except Exception:
        access_error = (
            "La liste des collaborateurs ajoutés ne peut pas être chargée "
            "pour le moment. Les comptes administrateur et directeurs "
            "protégés restent accessibles."
        )

    return authorized_users, access_error

DIRECTOR_MANAGEMENT_PROFILES = (
    {
        "email": "a.stone.authorbusiness@gmail.com",
        "name": "Biker",
        "direction": "Direction Biker",
    },
    {
        "email": "melvynschmidt2013@gmail.com",
        "name": "Max",
        "direction": "Direction Max",
    },
    {
        "email": "moon441330@gmail.com",
        "name": "Moon",
        "direction": "Direction Moon",
    },
    {
        "email": "vividirectrice@gmail.com",
        "name": "Vivi",
        "direction": "Direction Vivi",
    },
)


def default_director_management_config():
    return {
        profile["email"]: {
            "groups": [],
            "revenue_usd": 0.0,
            "other_expenses": 0.0,
        }
        for profile in DIRECTOR_MANAGEMENT_PROFILES
    }


def clean_director_management_config(payload, available_groups):
    """Nettoie les affectations enregistrées avant leur utilisation."""
    def non_negative_number(value):
        try:
            return max(0.0, float(value or 0.0))
        except (TypeError, ValueError):
            return 0.0

    available_group_set = set(available_groups)
    payload = payload if isinstance(payload, dict) else {}
    cleaned_config = default_director_management_config()

    for profile in DIRECTOR_MANAGEMENT_PROFILES:
        email = profile["email"]
        saved_row = payload.get(email, {})
        saved_row = saved_row if isinstance(saved_row, dict) else {}
        saved_groups = saved_row.get("groups", [])
        if not isinstance(saved_groups, list):
            saved_groups = []

        cleaned_config[email] = {
            "groups": [
                str(group)
                for group in saved_groups
                if str(group) in available_group_set
            ],
            "revenue_usd": non_negative_number(
                saved_row.get("revenue_usd", 0.0)
            ),
            "other_expenses": non_negative_number(
                saved_row.get("other_expenses", 0.0)
            ),
        }

    return cleaned_config


def normalize_group_name(value):
    """Normalise un nom de groupe pour un contrôle fiable des imports."""
    return " ".join(str(value or "").strip().casefold().split())


def validate_director_import_groups(
    prepared_dataframe,
    user_email,
    database_url,
):
    """Compare l’import d’un directeur à ses groupes autorisés."""
    if not database_url:
        return {
            "valid": False,
            "error": (
                "La base permanente est indisponible : les groupes autorisés "
                "ne peuvent pas être contrôlés. L’import est bloqué."
            ),
            "missing_groups": [],
            "unexpected_groups": [],
        }

    normalized_user_email = normalize_email(user_email)
    user_access = globals().get("AUTHORIZED_USERS", {}).get(
        normalized_user_email
    )
    if not user_access or user_access.get("role") != "director":
        return {
            "valid": False,
            "error": "Aucune direction n’est associée à ce compte.",
            "missing_groups": [],
            "unexpected_groups": [],
        }

    allowed_groups = user_access.get("groups", [])
    allowed_groups = (
        allowed_groups if isinstance(allowed_groups, list) else []
    )
    allowed_by_key = {
        normalize_group_name(group): str(group).strip()
        for group in allowed_groups
        if normalize_group_name(group)
    }

    if not allowed_by_key:
        return {
            "valid": False,
            "error": (
                "Aucun groupe n’a encore été attribué à votre direction. "
                "Demandez à l’administrateur de configurer vos groupes avant "
                "d’importer le fichier."
            ),
            "missing_groups": [],
            "unexpected_groups": [],
        }

    imported_groups = (
        prepared_dataframe["Groupe"]
        .fillna("")
        .astype(str)
        .str.strip()
    )
    imported_by_key = {
        normalize_group_name(group): group
        for group in imported_groups.unique()
        if normalize_group_name(group)
    }
    ungrouped_creators = int(
        imported_groups.map(normalize_group_name).eq("").sum()
    )

    missing_keys = set(allowed_by_key).difference(imported_by_key)
    unexpected_keys = set(imported_by_key).difference(allowed_by_key)
    missing_groups = sorted(allowed_by_key[key] for key in missing_keys)
    unexpected_groups = sorted(
        imported_by_key[key] for key in unexpected_keys
    )

    if missing_groups or unexpected_groups or ungrouped_creators:
        return {
            "valid": False,
            "error": (
                "Le fichier ne correspond pas exactement aux groupes "
                "attribués à votre direction. L’import est inutilisable."
            ),
            "missing_groups": missing_groups,
            "unexpected_groups": unexpected_groups,
            "ungrouped_creators": ungrouped_creators,
        }

    return {
        "valid": True,
        "error": None,
        "missing_groups": [],
        "unexpected_groups": [],
        "ungrouped_creators": 0,
        "allowed_groups": sorted(allowed_by_key.values()),
    }


def authentication_is_configured():
    try:
        if "auth" not in st.secrets:
            return False

        auth = st.secrets["auth"]
        if "google" not in auth:
            return False

        google = auth["google"]
        return all(
            auth.get(key)
            for key in ("redirect_uri", "cookie_secret")
        ) and all(
            google.get(key)
            for key in (
                "client_id",
                "client_secret",
                "server_metadata_url",
            )
        )
    except (FileNotFoundError, KeyError):
        return False


def login_screen():
    render_brand_hero(
        "Connexion sécurisée",
        "Accédez à votre espace de calcul avec l’adresse professionnelle "
        "autorisée pour votre direction.",
        "PRO CONSULTING • ESPACE PRIVÉ",
    )
    st.button(
        "Continuer avec Google",
        on_click=st.login,
        args=["google"],
        use_container_width=True,
        type="primary",
    )


if not authentication_is_configured():
    st.error(
        "La connexion sécurisée n’est pas encore configurée dans les "
        "Secrets Streamlit."
    )
    st.info(
        "Ajoutez la configuration Google dans les Secrets de "
        "l’application, puis redémarrez-la."
    )
    st.stop()

if not st.user.is_logged_in:
    login_screen()
    st.stop()

user_claims = st.user.to_dict()
current_user_email = normalize_email(
    user_claims.get("email")
    or user_claims.get("preferred_username")
    or user_claims.get("upn")
)
database_url = get_database_url()
AUTHORIZED_USERS, collaborator_access_load_error = build_authorized_users(
    database_url
)
current_user_access = AUTHORIZED_USERS.get(current_user_email)

if current_user_access is None:
    st.error(
        "Accès refusé : cette adresse n’est pas autorisée à utiliser "
        "Pro Consulting."
    )
    if current_user_email:
        st.code(current_user_email)
    st.button(
        "Se déconnecter",
        on_click=st.logout,
        use_container_width=True,
    )
    st.stop()

current_user_role = current_user_access["role"]
current_user_name = current_user_access["name"]
current_user_direction = current_user_access["direction"]
current_user_groups = list(current_user_access.get("groups", []))

# Streamlit peut conserver la même session après une déconnexion. On efface
# donc toutes les données propres à la branche lors d'un changement de compte.
if st.session_state.get("active_user_email") != current_user_email:
    for key in BRANCH_PARAMETER_KEYS:
        st.session_state[key] = DEFAULT_VALUES[key]
    st.session_state.exclusions = []
    clear_backstage_session()
    st.session_state.pop("exclusions_editor", None)
    st.session_state.pop("persistent_settings_loaded_for", None)
    st.session_state.pop("persistent_backstage_loaded_for", None)
    st.session_state.pop("backstage_restore_notice", None)
    st.session_state.pop("backstage_restore_error", None)
    st.session_state.pop("admin_director_management_loaded", None)
    st.session_state.pop("admin_director_management_config", None)
    st.session_state.pop("admin_collaborator_access_loaded", None)
    st.session_state.pop("admin_collaborator_access_rows", None)
    st.session_state.pop("chat_push_saved_endpoint", None)
    st.session_state.pop("last_internal_chat_toast_id", None)
    st.session_state.pop("last_rendered_chat_message_id", None)
    st.session_state.pop("last_read_chat_message_id", None)
    st.session_state.pop("handled_chat_deeplink", None)
    st.session_state.pop("selected_tournament_id", None)
    st.session_state.pop("tournament_category_view", None)
    st.session_state.active_user_email = current_user_email

persistent_settings_available = database_url is not None
persistent_settings_error = None

if persistent_settings_available:
    try:
        settings_session_key = f"settings:{current_user_email}"
        if (
            st.session_state.get("persistent_settings_loaded_for")
            != settings_session_key
        ):
            apply_persistent_settings(
                database_url,
                current_user_email,
            )
            st.session_state.persistent_settings_loaded_for = (
                settings_session_key
            )
    except Exception:
        persistent_settings_available = False
        persistent_settings_error = (
            "La section [database] est présente dans les Secrets, mais la "
            "connexion PostgreSQL échoue actuellement. "
            "Les valeurs de cette session restent utilisables, mais elles "
            "ne seront pas sauvegardées après la déconnexion."
        )

if persistent_settings_available:
    backstage_session_key = backstage_import_scope(current_user_email)
    if (
        st.session_state.get("persistent_backstage_loaded_for")
        != backstage_session_key
    ):
        try:
            saved_backstage = load_persistent_scope(
                database_url,
                backstage_session_key,
            )
            serialized_backstage = saved_backstage.get(
                "prepared_dataframe_json"
            )
            if serialized_backstage:
                restored_backstage = deserialize_backstage_dataframe(
                    serialized_backstage
                )
                if current_user_role == "director":
                    restored_import_control = validate_director_import_groups(
                        prepared_dataframe=restored_backstage,
                        user_email=current_user_email,
                        database_url=database_url,
                    )
                    if not restored_import_control["valid"]:
                        raise ValueError(restored_import_control["error"])

                clear_backstage_session()
                st.session_state.backstage_data = restored_backstage
                st.session_state.backstage_filename = str(
                    saved_backstage.get("filename", "Export enregistré")
                )
                st.session_state.detected_columns = saved_backstage.get(
                    "detected_columns",
                    {},
                )
                st.session_state.backstage_file_digest = str(
                    saved_backstage.get("file_digest", "")
                ) or hashlib.sha256(
                    serialized_backstage.encode("utf-8")
                ).hexdigest()
                st.session_state.backstage_restore_notice = (
                    "Le dernier export Backstage enregistré a été restauré "
                    "automatiquement."
                )

            st.session_state.persistent_backstage_loaded_for = (
                backstage_session_key
            )
        except Exception as error:
            clear_backstage_session()
            st.session_state.backstage_restore_error = (
                "Le dernier export enregistré n'a pas pu être restauré : "
                f"{error}"
            )
            st.session_state.persistent_backstage_loaded_for = (
                backstage_session_key
            )


st.sidebar.markdown(
    """
    <div class="pc-sidebar-brand">
        <div class="pc-sidebar-logo"><span>◆</span> Pro Consulting</div>
        <div class="pc-sidebar-tagline">Pilotage & rémunérations</div>
    </div>
    """,
    unsafe_allow_html=True,
)
st.sidebar.markdown(
    f"""
    <div class="pc-user-card">
        <div class="pc-user-status">● Connecté</div>
        <strong>{escape(current_user_name)}</strong>
        <span>{escape(current_user_direction)}</span>
    </div>
    """,
    unsafe_allow_html=True,
)
st.sidebar.button(
    "🚪 Se déconnecter",
    on_click=st.logout,
    use_container_width=True,
)

if collaborator_access_load_error and current_user_role == "admin":
    st.sidebar.warning(collaborator_access_load_error)

admin_pages = [
    "🏠 Accueil",
    "💬 Chat collectif",
    "🏆 Tournois",
    "📥 Import Backstage",
    "⚙️ Paramètres",
    "🛡️ Administration",
    "🔐 Accès collaborateurs",
    "💎 Créateurs",
    "👥 Consultants",
    "📈 Responsables performance",
    "🎁 Suivi récompenses",
    "🏢 Directeur de branche",
    "💰 Bénéfice agence",
]

director_pages = [
    "🏠 Accueil",
    "💬 Chat collectif",
    "🏆 Tournois",
    "📥 Import Backstage",
    "⚙️ Paramètres",
    "🛡️ Administration",
    "💎 Créateurs",
    "👥 Consultants",
    "📈 Responsables performance",
    "🎁 Suivi récompenses",
    "🏢 Directeur de branche",
]

performance_manager_pages = [
    "💬 Chat collectif",
    "🏆 Tournois",
    "🎁 Suivi récompenses",
]

pages_by_role = {
    "admin": admin_pages,
    "director": director_pages,
    "performance_manager": performance_manager_pages,
}

CHAT_PAGE_LABEL = "💬 Chat collectif"
requested_page = str(st.query_params.get("page", "")).strip().casefold()
requested_message = str(st.query_params.get("message", "")).strip()
chat_deeplink_token = f"{requested_page}:{requested_message}"
if (
    requested_page == "chat"
    and st.session_state.get("handled_chat_deeplink")
    != chat_deeplink_token
):
    st.session_state.main_navigation_page = CHAT_PAGE_LABEL
    st.session_state.handled_chat_deeplink = chat_deeplink_token

page = st.sidebar.radio(
    "Navigation",
    pages_by_role.get(current_user_role, []),
    key="main_navigation_page",
)


def open_collective_chat_from_notification():
    """Ouvre le chat depuis l’alerte interne de la barre latérale."""
    st.session_state.main_navigation_page = CHAT_PAGE_LABEL
    st.rerun(scope="app")


if persistent_settings_available and page != CHAT_PAGE_LABEL:
    def render_internal_chat_notification():
        try:
            notification_summary = load_chat_notification_summary(
                database_url,
                current_user_email,
            )
        except Exception:
            return

        unread_count = notification_summary["unread_count"]
        latest_unread_id = notification_summary["latest_unread_id"]
        if not unread_count:
            return

        latest_message = notification_summary.get("latest_message") or {}
        latest_author = latest_message.get("author_name", "Un collaborateur")
        latest_body = " ".join(
            str(latest_message.get("message", "")).split()
        )
        if len(latest_body) > 90:
            latest_body = latest_body[:87].rstrip() + "…"

        if (
            st.session_state.get("last_internal_chat_toast_id")
            != latest_unread_id
        ):
            toast_message = f"Nouveau message de {latest_author}"
            if latest_body:
                toast_message += f" : {latest_body}"
            st.toast(toast_message, icon="💬")
            st.session_state.last_internal_chat_toast_id = latest_unread_id

        with st.sidebar:
            st.warning(
                f"🔔 {unread_count} nouveau"
                f"{'x' if unread_count > 1 else ''} message"
                f"{'s' if unread_count > 1 else ''} dans le chat."
            )
            st.button(
                "💬 Ouvrir le chat",
                key="open_collective_chat_notification",
                on_click=open_collective_chat_from_notification,
                use_container_width=True,
            )

    render_internal_chat_notification()

# Le suivi est partagé entre plusieurs sessions Streamlit. Lorsqu'un
# utilisateur revient sur cette page, on recharge systématiquement la copie
# PostgreSQL afin qu'il voie les enregistrements effectués entre-temps par une
# autre direction. Les reruns provoqués par l'éditeur ne déclenchent pas ce
# rechargement et conservent donc les modifications locales non enregistrées.
previous_navigation_page = st.session_state.get(
    "previous_navigation_page"
)
if (
    page == "🎁 Suivi récompenses"
    and previous_navigation_page != page
):
    st.session_state.pop("reward_tracking_loaded_scope", None)
    st.session_state.pop("reward_tracking_editor", None)
    st.session_state.pop("manager_reward_tracking_editor", None)
st.session_state.previous_navigation_page = page


if page == "🏠 Accueil":
    render_brand_hero(
        "Pilotez vos rémunérations",
        "Importez les données Backstage, contrôlez chaque niveau de "
        "rémunération et obtenez une vision claire de votre branche.",
        "PRO CONSULTING • TABLEAU DE BORD",
    )
    st.write(f"Bienvenue **{current_user_name}** — {current_user_direction}")

    if st.session_state.backstage_data is None:
        st.info("Aucun export Backstage n’a encore été importé.")
    else:
        st.success(f"Export chargé : {st.session_state.backstage_filename}")
        st.metric(
            "Nombre de créateurs importés",
            len(st.session_state.backstage_data),
        )

    st.write(
        "Utilisez le menu de gauche pour importer l’export et calculer "
        "les rémunérations."
    )


elif page == "💬 Chat collectif":
    render_brand_hero(
        "Chat collectif",
        "Échangez entre l’administration, les directeurs et les "
        "responsables performance. Chaque message disparaît "
        "automatiquement après 48 heures.",
        "PRO CONSULTING • ÉQUIPE",
    )

    if not persistent_settings_available:
        st.error(
            "Le chat collectif nécessite la connexion PostgreSQL. Aucun "
            "message local ne sera accepté afin d’éviter une fausse "
            "confirmation d’envoi."
        )
        if persistent_settings_error:
            st.warning(persistent_settings_error)
        st.stop()

    try:
        web_push_keys = get_or_create_web_push_keys(database_url)
        with st.expander(
            "🔔 Notifications mobiles du chat",
            expanded=not bool(
                st.session_state.get("chat_push_saved_endpoint")
            ),
        ):
            st.caption(
                "L’activation est propre à chaque téléphone et à chaque "
                "navigateur. Aucun collaborateur n’est obligé de l’accepter."
            )
            push_component_result = chat_push_component(
                data={
                    "public_key": web_push_keys["public_key"],
                    "worker_url": CHAT_PUSH_WORKER_URL,
                    "manifest_url": CHAT_PUSH_MANIFEST_URL,
                },
                key=(
                    f"chat_push_{CHAT_PUSH_FRONTEND_VERSION}_"
                    f"{safe_export_name(current_user_email)}"
                ),
                on_push_state_change=lambda: None,
            )
            push_state_value = getattr(
                push_component_result,
                "push_state",
                None,
            )
            if push_state_value:
                push_state = (
                    json.loads(push_state_value)
                    if isinstance(push_state_value, str)
                    else dict(push_state_value)
                )
                push_subscription = push_state.get("subscription")
                if push_subscription:
                    push_endpoint = str(
                        push_subscription.get("endpoint", "")
                    )
                    if (
                        push_endpoint
                        and st.session_state.get(
                            "chat_push_saved_endpoint"
                        )
                        != push_endpoint
                    ):
                        saved_endpoint = save_chat_push_subscription(
                            database_url,
                            current_user_email,
                            push_subscription,
                        )
                        st.session_state.chat_push_saved_endpoint = (
                            saved_endpoint
                        )
                        st.toast(
                            "Notifications mobiles activées sur ce téléphone.",
                            icon="🔔",
                        )
    except Exception:
        st.warning(
            "Les notifications mobiles ne peuvent pas être configurées "
            "pour le moment. Le chat et les alertes internes restent actifs."
        )

    st.caption(
        "🔄 Actualisation automatique toutes les 10 secondes • "
        "Conservation maximale : 48 heures"
    )

    fragment_decorator = (
        st.fragment(run_every="10s")
        if hasattr(st, "fragment")
        else (lambda function: function)
    )

    @fragment_decorator
    def render_collective_chat_messages():
        try:
            chat_messages = load_collective_chat_messages(database_url)
        except Exception:
            st.error(
                "Les messages ne peuvent pas être chargés pour le moment."
            )
            return

        if not chat_messages:
            st.info(
                "Aucun message actif. Le prochain message sera visible par "
                "tous les collaborateurs autorisés."
            )
            return

        latest_chat_message_id = int(chat_messages[-1]["id"])
        previous_rendered_message_id = int(
            st.session_state.get("last_rendered_chat_message_id", 0)
        )
        if (
            previous_rendered_message_id
            and latest_chat_message_id > previous_rendered_message_id
        ):
            new_external_messages = [
                message
                for message in chat_messages
                if int(message["id"]) > previous_rendered_message_id
                and normalize_email(message["author_email"])
                != normalize_email(current_user_email)
            ]
            if new_external_messages:
                newest_message = new_external_messages[-1]
                st.toast(
                    "Nouveau message de "
                    f"{newest_message['author_name']}.",
                    icon="💬",
                )
        st.session_state.last_rendered_chat_message_id = (
            latest_chat_message_id
        )

        if (
            int(st.session_state.get("last_read_chat_message_id", 0))
            < latest_chat_message_id
        ):
            try:
                mark_collective_chat_read(
                    database_url,
                    current_user_email,
                    latest_chat_message_id,
                )
                st.session_state.last_read_chat_message_id = (
                    latest_chat_message_id
                )
            except Exception:
                pass

        for chat_message in chat_messages:
            displayed_author_name = AUTHORIZED_USERS.get(
                normalize_email(chat_message["author_email"]),
                {},
            ).get("name", chat_message["author_name"])
            role_label = COLLABORATOR_ROLE_LABELS.get(
                chat_message["author_role"],
                "Administrateur"
                if chat_message["author_role"] == "admin"
                else chat_message["author_role"],
            )
            message_body = escape(
                str(chat_message["message"] or "")
            ).replace("\n", "<br>")
            st.markdown(
                f"""
                <div class="pc-chat-message">
                    <div class="pc-chat-meta">
                        {escape(str(displayed_author_name))}
                        • {escape(str(role_label))}
                        • {escape(format_chat_datetime(chat_message['created_at']))}
                    </div>
                    <div class="pc-chat-body">{message_body}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    refresh_chat_column, identity_chat_column = st.columns([1, 3])
    refresh_chat_column.button(
        "🔄 Actualiser",
        key="refresh_collective_chat",
        use_container_width=True,
    )
    identity_chat_column.info(
        f"Vous écrivez en tant que **{current_user_name}** — "
        f"{current_user_direction}."
    )
    render_collective_chat_messages()

    with st.form("collective_chat_form", clear_on_submit=True):
        chat_message_text = st.text_area(
            "Votre message",
            max_chars=CHAT_MAX_MESSAGE_LENGTH,
            height=100,
            placeholder="Écrivez votre message à l’équipe…",
        )
        send_chat_message = st.form_submit_button(
            "📨 Envoyer au chat collectif",
            type="primary",
            use_container_width=True,
        )

    if send_chat_message:
        if not str(chat_message_text or "").strip():
            st.warning("Écrivez un message avant de l’envoyer.")
        else:
            try:
                inserted_message_id = save_collective_chat_message(
                    database_url=database_url,
                    author_email=current_user_email,
                    author_name=current_user_name,
                    author_role=current_user_role,
                    message=chat_message_text,
                )
                try:
                    send_collective_chat_push_notifications(
                        database_url=database_url,
                        author_email=current_user_email,
                        author_name=current_user_name,
                        message=chat_message_text,
                        message_id=inserted_message_id,
                    )
                except Exception:
                    pass
                st.rerun()
            except ValueError as error:
                st.warning(str(error))
            except Exception:
                st.error(
                    "Le message n’a pas été enregistré. Vous pouvez le "
                    "copier puis réessayer sans perdre son contenu."
                )


elif page == "🏆 Tournois":
    render_brand_hero(
        "Tournois Pro Consulting",
        "Créez des tournois 1v1 ou 2v2, inscrivez les créateurs en groupe "
        "et suivez automatiquement les qualifications jusqu’à la finale.",
        "PRO CONSULTING • COMPÉTITIONS",
    )

    if not persistent_settings_available:
        st.error(
            "Le module Tournois nécessite PostgreSQL afin que les tirages "
            "et les résultats soient identiques pour tous les comptes."
        )
        st.stop()

    tournament_delete_notice = st.session_state.pop(
        "tournament_delete_notice",
        None,
    )
    if tournament_delete_notice:
        st.success(tournament_delete_notice)

    tournament_status_labels = {
        "registration": "📝 Inscriptions ouvertes",
        "draw_ready": "🎲 Tirage à valider",
        "validated": "✅ Tableau validé",
        "in_progress": "▶️ En cours",
        "finished": "🏆 Terminé",
        "archived": "📦 Archivé",
    }

    if current_user_role in {"admin", "director"}:
        with st.expander("➕ Créer un nouveau tournoi", expanded=False):
            st.caption(
                "Le tournoi sera général à la structure depuis le compte "
                "fondateur, et privé à votre branche depuis un compte "
                "directeur."
            )
            with st.form("create_tournament_form", clear_on_submit=True):
                new_tournament_title = st.text_input(
                    "Titre du tournoi",
                    placeholder="Exemple : Tournoi de septembre",
                )
                new_tournament_format = st.radio(
                    "Format",
                    ["1v1", "2v2"],
                    horizontal=True,
                )
                new_registration_deadline = st.date_input(
                    "Date limite d’inscription",
                    value=date.today(),
                    format="DD/MM/YYYY",
                )
                create_tournament_button = st.form_submit_button(
                    "Créer le tournoi",
                    type="primary",
                    use_container_width=True,
                )
            if create_tournament_button:
                try:
                    created_tournament_id = create_tournament(
                        database_url=database_url,
                        title=new_tournament_title,
                        tournament_format=new_tournament_format,
                        registration_deadline=new_registration_deadline,
                        creator_email=current_user_email,
                        creator_name=current_user_name,
                        creator_role=current_user_role,
                    )
                    st.session_state.selected_tournament_id = (
                        created_tournament_id
                    )
                    st.success("Le tournoi a été créé.")
                    st.rerun()
                except (ValueError, PermissionError) as error:
                    st.warning(str(error))
                except Exception:
                    st.error("La création du tournoi a échoué.")

    refresh_tournament_column, tournament_info_column = st.columns([1, 3])
    if refresh_tournament_column.button(
        "🔄 Actualiser",
        key="refresh_tournaments",
        use_container_width=True,
    ):
        st.rerun()
    tournament_info_column.caption(
        "Les tirages, résultats et droits sont relus depuis PostgreSQL."
    )

    try:
        available_tournaments = list_tournaments(
            database_url,
            current_user_email,
            current_user_role,
        )
    except Exception:
        st.error("Les tournois ne peuvent pas être chargés pour le moment.")
        st.stop()

    if current_user_role == "admin":
        tournament_view = st.radio(
            "Catégorie de tournois",
            [
                "🏢 Tournois de la structure",
                "👤 Tournois individuels des directeurs",
                "📚 Tous les tournois",
            ],
            horizontal=True,
            key="tournament_category_view",
        )
        if tournament_view.startswith("🏢"):
            displayed_tournaments = [
                row
                for row in available_tournaments
                if row["scope_type"] == "structure"
            ]
        elif tournament_view.startswith("👤"):
            displayed_tournaments = [
                row
                for row in available_tournaments
                if row["scope_type"] == "branch"
            ]
        else:
            displayed_tournaments = available_tournaments
    elif current_user_role == "director":
        tournament_view = st.radio(
            "Catégorie de tournois",
            [
                "🏢 Tournois de la structure",
                "👤 Mes tournois individuels",
            ],
            horizontal=True,
            key="tournament_category_view",
        )
        if tournament_view.startswith("🏢"):
            displayed_tournaments = [
                row
                for row in available_tournaments
                if row["scope_type"] == "structure"
            ]
        else:
            displayed_tournaments = [
                row
                for row in available_tournaments
                if row["scope_type"] == "branch"
                and normalize_email(row["owner_email"])
                == normalize_email(current_user_email)
            ]
    else:
        displayed_tournaments = available_tournaments

    if not displayed_tournaments:
        if current_user_role == "performance_manager":
            st.info(
                "Aucun tournoi ne vous est actuellement rendu visible par "
                "FONDATEUR ADMIN."
            )
        elif current_user_role == "admin" and tournament_view.startswith("👤"):
            st.info(
                "Aucun directeur n’a encore créé de tournoi individuel. "
                "Ils le feront depuis ce même onglet avec leur propre compte."
            )
        elif current_user_role == "director" and tournament_view.startswith("👤"):
            st.info(
                "Vous n’avez pas encore créé de tournoi individuel pour "
                "votre branche. Utilisez « Créer un nouveau tournoi » ci-dessus."
            )
        else:
            st.info("Aucun tournoi dans cette catégorie.")
        st.stop()

    tournament_by_id = {
        tournament["id"]: tournament
        for tournament in displayed_tournaments
    }
    tournament_ids = list(tournament_by_id)
    selected_tournament_id = st.session_state.get(
        "selected_tournament_id"
    )
    if selected_tournament_id not in tournament_by_id:
        selected_tournament_id = tournament_ids[0]

    selected_tournament_id = st.selectbox(
        "Tournoi affiché",
        tournament_ids,
        index=tournament_ids.index(selected_tournament_id),
        format_func=lambda tournament_id: (
            f"{tournament_by_id[tournament_id]['title']} • "
            f"{tournament_by_id[tournament_id]['format']} • "
            f"{tournament_status_labels.get(tournament_by_id[tournament_id]['status'], tournament_by_id[tournament_id]['status'])}"
        ),
        key="selected_tournament_id",
    )
    try:
        tournament = load_tournament(
            database_url,
            selected_tournament_id,
            current_user_email,
            current_user_role,
        )
    except (ValueError, PermissionError) as error:
        st.error(str(error))
        st.stop()

    is_tournament_operator = bool(
        current_user_role == "admin"
        or (
            current_user_role == "director"
            and (
                tournament["scope_type"] == "structure"
                or normalize_email(tournament["owner_email"])
                == normalize_email(current_user_email)
            )
        )
    )
    can_finalize_tournament = bool(
        current_user_role == "admin"
        or (
            current_user_role == "director"
            and tournament["scope_type"] == "branch"
            and normalize_email(tournament["owner_email"])
            == normalize_email(current_user_email)
        )
    )
    scope_label = (
        "Tournoi général de la structure"
        if tournament["scope_type"] == "structure"
        else f"Tournoi privé • {tournament['owner_name']}"
    )
    st.subheader(tournament["title"])
    st.caption(
        f"{scope_label} • Format {tournament['format']} • "
        f"{tournament_status_labels.get(tournament['status'], tournament['status'])}"
    )

    participant_rows, match_rows = tournament_tables(tournament)
    participant_count = len(participant_rows)
    duo_count = (
        len(tournament.get("competitors", []))
        if tournament["format"] == "2v2"
        else 0
    )
    decided_match_count = sum(
        bool(match.get("winner_id")) and not match.get("bye")
        for match in tournament.get("matches", [])
    )
    summary_columns = st.columns(4)
    summary_columns[0].metric("Inscrits", participant_count)
    summary_columns[1].metric(
        "Duos",
        duo_count if tournament["format"] == "2v2" else "—",
    )
    summary_columns[2].metric("Matchs générés", len(match_rows))
    summary_columns[3].metric("Résultats saisis", decided_match_count)

    if current_user_role == "admin":
        with st.expander("⚙️ Contrôle fondateur", expanded=False):
            deadline_value = tournament.get("registration_deadline")
            try:
                deadline_value = date.fromisoformat(deadline_value)
            except (TypeError, ValueError):
                deadline_value = date.today()
            with st.form(
                f"founder_tournament_options_{selected_tournament_id}"
            ):
                edited_tournament_title = st.text_input(
                    "Titre",
                    value=tournament["title"],
                )
                edited_tournament_deadline = st.date_input(
                    "Date limite d’inscription",
                    value=deadline_value,
                    format="DD/MM/YYYY",
                )
                manager_visibility = st.checkbox(
                    "Visible en lecture seule par les responsables performance",
                    value=bool(tournament["visible_to_managers"]),
                    disabled=tournament["status"] in {"registration", "draw_ready", "archived"},
                )
                archive_tournament = st.checkbox(
                    "Archiver et retirer le tournoi des responsables performance",
                    value=tournament["status"] == "archived",
                )
                save_tournament_options = st.form_submit_button(
                    "Enregistrer les réglages",
                    use_container_width=True,
                )
            if save_tournament_options:
                try:
                    update_tournament_options(
                        database_url,
                        selected_tournament_id,
                        current_user_email,
                        current_user_role,
                        title=edited_tournament_title,
                        registration_deadline=edited_tournament_deadline,
                        visible_to_managers=(
                            manager_visibility
                            if tournament["status"] not in {"registration", "draw_ready", "archived"}
                            else False
                        ),
                        archive=archive_tournament,
                    )
                    st.success("Les réglages du tournoi sont enregistrés.")
                    st.rerun()
                except Exception as error:
                    st.error(f"Réglages non enregistrés : {error}")

    elif (
        current_user_role == "director"
        and tournament["scope_type"] == "branch"
        and normalize_email(tournament["owner_email"])
        == normalize_email(current_user_email)
    ):
        with st.expander("⚙️ Réglages de mon tournoi", expanded=False):
            deadline_value = tournament.get("registration_deadline")
            try:
                deadline_value = date.fromisoformat(deadline_value)
            except (TypeError, ValueError):
                deadline_value = date.today()
            with st.form(f"director_tournament_options_{selected_tournament_id}"):
                edited_tournament_title = st.text_input(
                    "Titre", value=tournament["title"]
                )
                edited_tournament_deadline = st.date_input(
                    "Date limite d’inscription",
                    value=deadline_value,
                    format="DD/MM/YYYY",
                )
                save_tournament_options = st.form_submit_button(
                    "Enregistrer",
                    use_container_width=True,
                )
            if save_tournament_options:
                try:
                    update_tournament_options(
                        database_url,
                        selected_tournament_id,
                        current_user_email,
                        current_user_role,
                        title=edited_tournament_title,
                        registration_deadline=edited_tournament_deadline,
                    )
                    st.rerun()
                except Exception as error:
                    st.error(f"Réglages non enregistrés : {error}")

    if tournament["status"] == "registration" and is_tournament_operator:
        st.subheader("Inscriptions")
        st.caption(
            "Collez un pseudo par ligne. Les virgules, points-virgules et "
            "tabulations sont également reconnus ; les doublons sont bloqués."
        )
        with st.form(
            f"add_tournament_participants_{selected_tournament_id}",
            clear_on_submit=True,
        ):
            participant_batch = st.text_area(
                "Créateurs à inscrire",
                height=160,
                placeholder="createur_1\ncreateur_2\ncreateur_3",
            )
            add_tournament_participants = st.form_submit_button(
                "Ajouter les créateurs",
                type="primary",
                use_container_width=True,
            )
        if add_tournament_participants:
            try:
                parsed_participants = parse_participant_batch(
                    participant_batch
                )
                addition_result = add_participants(
                    database_url,
                    selected_tournament_id,
                    parsed_participants,
                    current_user_email,
                    current_user_name,
                    current_user_role,
                )
                if addition_result["duplicates"]:
                    st.warning(
                        f"{len(addition_result['duplicates'])} doublon(s) "
                        "ont été ignorés."
                    )
                st.success(
                    f"{len(addition_result['added'])} créateur(s) ajouté(s)."
                )
                st.rerun()
            except (ValueError, PermissionError) as error:
                st.warning(str(error))
            except Exception:
                st.error("Les inscriptions n’ont pas été enregistrées.")

        if participant_rows:
            participant_dataframe = pd.DataFrame(participant_rows)
            st.dataframe(
                participant_dataframe,
                use_container_width=True,
                hide_index=True,
            )
            participant_options = {
                row["id"]: row["name"]
                for row in tournament["participants"]
            }
            with st.form(
                f"remove_tournament_participant_{selected_tournament_id}"
            ):
                participant_to_remove = st.selectbox(
                    "Retirer un participant",
                    list(participant_options),
                    format_func=lambda participant_id: participant_options[participant_id],
                )
                confirm_participant_removal = st.checkbox(
                    "Je confirme le retrait de ce participant"
                )
                remove_tournament_participant_button = st.form_submit_button(
                    "Retirer",
                    disabled=not confirm_participant_removal,
                    use_container_width=True,
                )
            if remove_tournament_participant_button:
                try:
                    remove_participant(
                        database_url,
                        selected_tournament_id,
                        participant_to_remove,
                        current_user_email,
                        current_user_role,
                    )
                    st.rerun()
                except Exception as error:
                    st.error(f"Participant non retiré : {error}")

        minimum_participants = 2 if tournament["format"] == "1v1" else 3
        st.subheader("Préparer le tirage")
        if not can_finalize_tournament:
            st.info(
                "Les inscriptions restent ouvertes. Seul FONDATEUR ADMIN "
                "validera le tirage du tournoi général."
            )
        if tournament["format"] == "2v2":
            solo_policy_label = st.radio(
                "Si un créateur reste seul",
                [
                    "Le placer en attente d’un partenaire",
                    "L’autoriser à jouer seul contre un duo",
                ],
                key=f"solo_policy_{selected_tournament_id}",
            )
            solo_policy = (
                "waiting"
                if solo_policy_label.startswith("Le placer")
                else "solo"
            )
            if solo_policy == "waiting":
                minimum_participants = 4
        else:
            solo_policy = "solo"
            if participant_count % 2:
                st.info(
                    "Le tirage créera automatiquement un match à trois avec "
                    "un seul gagnant."
                )
        confirm_draw_creation = st.checkbox(
            "Je confirme la clôture des inscriptions et la création du tirage",
            key=f"confirm_draw_{selected_tournament_id}",
            disabled=not can_finalize_tournament,
        )
        if st.button(
            "🎲 Générer le tirage aléatoire",
            key=f"prepare_draw_{selected_tournament_id}",
            type="primary",
            use_container_width=True,
            disabled=(
                participant_count < minimum_participants
                or not confirm_draw_creation
                or not can_finalize_tournament
            ),
        ):
            try:
                finalize_draw(
                    database_url,
                    selected_tournament_id,
                    current_user_email,
                    current_user_role,
                    solo_policy=solo_policy,
                    preview_only=True,
                )
                st.rerun()
            except Exception as error:
                st.error(f"Le tirage n’a pas été généré : {error}")

    elif participant_rows:
        with st.expander("👥 Liste des participants", expanded=False):
            st.dataframe(
                pd.DataFrame(participant_rows),
                use_container_width=True,
                hide_index=True,
            )

    waiting_participant = tournament.get("waiting_participant")
    if waiting_participant:
        st.warning(
            "Créateur actuellement en attente d’un partenaire : "
            f"**{waiting_participant.get('name', '')}**"
        )

    if tournament.get("matches"):
        st.subheader("Tableau du tournoi")
        round_numbers = sorted(
            {
                int(match.get("round_number", 0))
                for match in tournament["matches"]
            }
        )
        round_tabs = st.tabs(
            [
                next(
                    match.get("round_label", f"Tour {round_number}")
                    for match in tournament["matches"]
                    if int(match.get("round_number", 0)) == round_number
                )
                for round_number in round_numbers
            ]
        )
        for round_tab, round_number in zip(round_tabs, round_numbers):
            with round_tab:
                round_matches = [
                    match
                    for match in tournament["matches"]
                    if int(match.get("round_number", 0)) == round_number
                ]
                for match in round_matches:
                    contestant_labels = [
                        contestant.get("label", "")
                        for contestant in match.get("contestants", [])
                    ]
                    winner = next(
                        (
                            contestant
                            for contestant in match.get("contestants", [])
                            if contestant.get("id") == match.get("winner_id")
                        ),
                        None,
                    )
                    match_title = (
                        f"Match {match.get('match_number')} • "
                        + "  VS  ".join(contestant_labels)
                    )
                    with st.container(border=True):
                        st.markdown(f"**{match_title}**")
                        schedule_text = " • ".join(
                            value
                            for value in (
                                match.get("date", ""),
                                match.get("time", ""),
                            )
                            if value
                        )
                        if schedule_text:
                            st.caption(f"📅 {schedule_text}")
                        if match.get("bye"):
                            st.info(
                                f"Qualification automatique : {contestant_labels[0]}"
                            )
                        elif winner:
                            st.success(f"Gagnant : {winner.get('label', '')}")
                        else:
                            st.caption("Résultat en attente")

                        if (
                            is_tournament_operator
                            and tournament["status"]
                            in {"validated", "in_progress", "finished"}
                            and not match.get("bye")
                        ):
                            winner_options = {
                                contestant.get("id"): contestant.get("label", "")
                                for contestant in match.get("contestants", [])
                            }
                            current_winner_id = match.get("winner_id")
                            winner_ids = list(winner_options)
                            winner_index = (
                                winner_ids.index(current_winner_id)
                                if current_winner_id in winner_ids
                                else 0
                            )
                            with st.form(
                                f"winner_form_{selected_tournament_id}_{match['id']}"
                            ):
                                selected_winner = st.selectbox(
                                    "Créateur ou duo gagnant",
                                    winner_ids,
                                    index=winner_index,
                                    format_func=lambda winner_id: winner_options[winner_id],
                                )
                                save_winner = st.form_submit_button(
                                    "Enregistrer le gagnant",
                                    use_container_width=True,
                                )
                            if save_winner:
                                try:
                                    status, champion = set_match_winner(
                                        database_url,
                                        selected_tournament_id,
                                        match["id"],
                                        selected_winner,
                                        current_user_email,
                                        current_user_role,
                                    )
                                    if champion:
                                        st.balloons()
                                    st.rerun()
                                except Exception as error:
                                    st.error(f"Résultat non enregistré : {error}")

                        if current_user_role == "admin" and not match.get("bye"):
                            with st.form(
                                f"schedule_form_{selected_tournament_id}_{match['id']}"
                            ):
                                scheduled_date = st.text_input(
                                    "Date du match",
                                    value=match.get("date", ""),
                                    placeholder="Exemple : 12/09/2026",
                                )
                                scheduled_time = st.text_input(
                                    "Heure du match",
                                    value=match.get("time", ""),
                                    placeholder="Exemple : 21H30",
                                )
                                save_schedule = st.form_submit_button(
                                    "Enregistrer date et heure",
                                    use_container_width=True,
                                )
                            if save_schedule:
                                try:
                                    update_match_schedule(
                                        database_url,
                                        selected_tournament_id,
                                        match["id"],
                                        scheduled_date,
                                        scheduled_time,
                                        current_user_email,
                                        current_user_role,
                                    )
                                    st.rerun()
                                except Exception as error:
                                    st.error(f"Planning non enregistré : {error}")

        if tournament["status"] == "draw_ready" and is_tournament_operator:
            st.warning(
                "Le tirage est uniquement en aperçu. Vérifiez tous les duos "
                "et adversaires avant de le valider."
            )
            confirm_final_draw = st.checkbox(
                "Je confirme définitivement ce tirage",
                key=f"confirm_final_draw_{selected_tournament_id}",
            )
            if st.button(
                "✅ Valider définitivement le tableau",
                key=f"validate_draw_{selected_tournament_id}",
                type="primary",
                use_container_width=True,
                disabled=(
                    not confirm_final_draw or not can_finalize_tournament
                ),
            ):
                try:
                    validate_prepared_draw(
                        database_url,
                        selected_tournament_id,
                        current_user_email,
                        current_user_role,
                    )
                    st.success("Le tableau du tournoi est maintenant validé.")
                    st.rerun()
                except Exception as error:
                    st.error(f"Validation impossible : {error}")

        if tournament["status"] == "finished":
            final_matches = [
                match
                for match in tournament["matches"]
                if match.get("round_number")
                == max(row.get("round_number", 0) for row in tournament["matches"])
            ]
            final_winner = next(
                (
                    contestant.get("label", "")
                    for match in final_matches
                    for contestant in match.get("contestants", [])
                    if contestant.get("id") == match.get("winner_id")
                ),
                "",
            )
            if final_winner:
                st.success(f"🏆 VAINQUEUR DU TOURNOI : {final_winner}")

        if match_rows:
            st.download_button(
                "⬇️ Télécharger le tournoi (Excel)",
                data=dataframes_to_excel(
                    (
                        ("Participants", pd.DataFrame(participant_rows)),
                        ("Tableau tournoi", pd.DataFrame(match_rows)),
                    )
                ),
                file_name=(
                    f"pro_consulting_{safe_export_name(tournament['title'])}.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                key=f"download_tournament_{selected_tournament_id}",
                use_container_width=True,
                on_click="ignore",
            )

    if current_user_role == "admin" and tournament["status"] != "registration":
        with st.expander("🛡️ Corriger ou recommencer", expanded=False):
            st.warning(
                "La réouverture efface uniquement le tirage et les résultats "
                "de ce tournoi. La liste des inscrits est conservée."
            )
            confirm_reopen = st.checkbox(
                "Je confirme la réouverture des inscriptions",
                key=f"confirm_reopen_{selected_tournament_id}",
            )
            if st.button(
                "Rouvrir les inscriptions et annuler le tirage",
                key=f"reopen_tournament_{selected_tournament_id}",
                use_container_width=True,
                disabled=not confirm_reopen,
            ):
                try:
                    reopen_registrations(
                        database_url,
                        selected_tournament_id,
                        current_user_email,
                        current_user_role,
                    )
                    st.rerun()
                except Exception as error:
                    st.error(f"Réouverture impossible : {error}")

    if current_user_role == "admin":
        with st.expander("🗑️ Supprimer définitivement ce tournoi", expanded=False):
            st.error(
                "Cette action supprimera uniquement ce tournoi, ses matchs "
                "et son historique. Elle ne touche à aucune autre donnée du logiciel."
            )
            with st.form(
                f"delete_tournament_form_{selected_tournament_id}"
            ):
                deletion_confirmation_text = st.text_input(
                    "Recopiez exactement le titre du tournoi",
                    placeholder=tournament["title"],
                )
                confirm_tournament_deletion = st.checkbox(
                    "Je confirme la suppression définitive"
                )
                delete_tournament_button = st.form_submit_button(
                    "Supprimer définitivement le tournoi",
                    type="primary",
                    use_container_width=True,
                    disabled=(
                        not confirm_tournament_deletion
                        or deletion_confirmation_text.strip()
                        != tournament["title"].strip()
                    ),
                )
            if delete_tournament_button:
                try:
                    deleted_tournament_title = delete_tournament(
                        database_url,
                        selected_tournament_id,
                        current_user_email,
                        current_user_role,
                    )
                    st.session_state.pop("selected_tournament_id", None)
                    st.session_state.tournament_delete_notice = (
                        f"Le tournoi « {deleted_tournament_title} » a été supprimé."
                    )
                    st.rerun()
                except Exception as error:
                    st.error(f"Suppression impossible : {error}")


elif page == "📥 Import Backstage":
    st.title("📥 Import Backstage")

    restore_notice = st.session_state.pop("backstage_restore_notice", None)
    if restore_notice:
        st.success(restore_notice)
    restore_error = st.session_state.pop("backstage_restore_error", None)
    if restore_error:
        st.warning(restore_error)

    if st.session_state.backstage_data is not None:
        st.success(
            f"Fichier actuellement chargé : "
            f"{st.session_state.backstage_filename}"
        )
        if persistent_settings_available:
            st.caption(
                "☁️ Cet export est enregistré pour votre compte et sera "
                "restauré automatiquement après votre prochaine connexion."
            )

        delete_generation = int(
            st.session_state.get("backstage_delete_generation", 0)
        )
        with st.expander("🗑️ Supprimer l’export actuellement enregistré"):
            st.warning(
                "Cette action supprime uniquement l’export de votre compte. "
                "Elle ne touche jamais aux fichiers des autres directions."
            )
            confirm_backstage_deletion = st.checkbox(
                "Je confirme la suppression de cet export Backstage.",
                key=f"confirm_backstage_deletion_{delete_generation}",
            )
            if st.button(
                "Supprimer définitivement cet export",
                key=f"delete_backstage_import_{delete_generation}",
                disabled=not confirm_backstage_deletion,
                use_container_width=True,
            ):
                if persistent_settings_error:
                    st.error(
                        "La base permanente est momentanément inaccessible. "
                        "La suppression est bloquée afin de conserver une "
                        "situation cohérente."
                    )
                else:
                    try:
                        if persistent_settings_available:
                            delete_persistent_scopes(
                                database_url,
                                [backstage_import_scope(current_user_email)],
                            )
                        clear_backstage_session()
                        st.session_state.backstage_delete_generation = (
                            delete_generation + 1
                        )
                        st.session_state.backstage_uploader_generation = int(
                            st.session_state.get(
                                "backstage_uploader_generation",
                                0,
                            )
                        ) + 1
                        st.rerun()
                    except Exception:
                        st.error(
                            "La suppression permanente a échoué. L’export "
                            "actuellement enregistré a été conservé."
                        )

    uploader_generation = int(
        st.session_state.get("backstage_uploader_generation", 0)
    )

    uploaded_file = st.file_uploader(
        "Choisir un nouvel export Backstage",
        type=["xlsx"],
        help=(
            "Après validation, ce fichier remplacera automatiquement "
            "l’ancien export enregistré pour votre compte."
        ),
        key=f"backstage_file_uploader_{uploader_generation}",
    )

    if uploaded_file is not None:
        try:
            uploaded_bytes = uploaded_file.getvalue()
            uploaded_digest = hashlib.sha256(uploaded_bytes).hexdigest()
            raw_dataframe = pd.read_excel(
                BytesIO(uploaded_bytes),
                sheet_name=0,
            )
            prepared_dataframe, detected_columns = prepare_backstage_data(
                raw_dataframe
            )

            if current_user_role == "director":
                import_control = validate_director_import_groups(
                    prepared_dataframe=prepared_dataframe,
                    user_email=current_user_email,
                    database_url=(
                        database_url if persistent_settings_available else None
                    ),
                )
                if not import_control["valid"]:
                    st.error(f"⛔ {import_control['error']}")
                    if import_control["missing_groups"]:
                        st.warning(
                            "Groupes manquants : "
                            + ", ".join(import_control["missing_groups"])
                        )
                    if import_control["unexpected_groups"]:
                        st.warning(
                            "Groupes non autorisés ou en trop : "
                            + ", ".join(
                                import_control["unexpected_groups"]
                            )
                        )
                    if import_control.get("ungrouped_creators", 0):
                        st.warning(
                            f"Créateurs sans groupe : "
                            f"{import_control['ungrouped_creators']}"
                        )
                    st.info(
                        "Aucune donnée de ce fichier n’a été chargée. "
                        "Le dernier export valide reste enregistré. Corrigez "
                        "le nouveau fichier puis recommencez."
                    )
                    st.stop()

            is_new_file = (
                uploaded_digest
                != st.session_state.get("backstage_file_digest")
            )
            st.session_state.backstage_data = prepared_dataframe
            st.session_state.backstage_raw_data = raw_dataframe
            st.session_state.backstage_filename = uploaded_file.name
            st.session_state.backstage_file_digest = uploaded_digest
            st.session_state.detected_columns = detected_columns

            if is_new_file:
                reset_calculations()
                if persistent_settings_available:
                    try:
                        save_persistent_scopes(
                            database_url,
                            {
                                backstage_import_scope(
                                    current_user_email
                                ): {
                                    "filename": uploaded_file.name,
                                    "file_digest": uploaded_digest,
                                    "month": st.session_state.month,
                                    "detected_columns": detected_columns,
                                    "prepared_dataframe_json": (
                                        serialize_backstage_dataframe(
                                            prepared_dataframe
                                        )
                                    ),
                                    "saved_at": datetime.now().isoformat(
                                        timespec="seconds"
                                    ),
                                }
                            },
                            current_user_email,
                        )
                        st.session_state.persistent_backstage_loaded_for = (
                            backstage_import_scope(current_user_email)
                        )
                        st.success(
                            "☁️ Le nouvel export remplace l’ancien et est "
                            "enregistré définitivement pour votre compte."
                        )
                    except Exception:
                        st.error(
                            "Le fichier est utilisable dans cette session, "
                            "mais sa sauvegarde permanente a échoué. "
                            "L’ancien export enregistré n’a pas été supprimé."
                        )
                elif persistent_settings_error:
                    st.error(persistent_settings_error)
                else:
                    st.warning(
                        "La base PostgreSQL n’est pas configurée : ce fichier "
                        "restera uniquement dans cette session."
                    )

            if current_user_role == "director":
                st.success(
                    "L’export Backstage correspond exactement aux groupes "
                    "autorisés pour votre direction."
                )
            else:
                st.success("L’export Backstage a été lu correctement.")
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Créateurs importés", len(prepared_dataframe))
            col2.metric(
                "Diamants générés",
                f"{prepared_dataframe['Diamants'].sum():,.0f}",
            )
            col3.metric(
                "Total heures LIVE",
                f"{prepared_dataframe['Heures LIVE'].sum():,.1f} h",
            )
            col4.metric("Colonnes détectées", len(detected_columns))

            st.subheader("Aperçu des données préparées")
            st.dataframe(
                prepared_dataframe.head(30),
                use_container_width=True,
                hide_index=True,
            )
            show_excel_download(
                prepared_dataframe,
                table_name="donnees_backstage_preparees",
                sheet_name="Backstage préparé",
                key="download_backstage_prepared",
                label="⬇️ Télécharger toutes les données préparées (Excel)",
            )

        except ValueError as error:
            st.error("Certaines colonnes obligatoires sont absentes.")
            st.code(str(error))
        except Exception as error:
            st.error("Une erreur empêche la lecture du fichier.")
            st.code(str(error))


elif page == "⚙️ Paramètres":
    st.title("⚙️ Paramètres mensuels")
    st.info(
        "Les quatre paliers sont indépendants et doivent être "
        "sélectionnés chaque mois."
    )

    if persistent_settings_available:
        st.success(
            "☁️ Sauvegarde permanente active : les paramètres de votre "
            "direction seront retrouvés après une déconnexion ou un "
            "redémarrage de l’application."
        )
    elif persistent_settings_error:
        st.warning(persistent_settings_error)
    else:
        st.warning(
            "La sauvegarde permanente n’est pas encore configurée. "
            "Les paramètres seront conservés uniquement pendant cette "
            "session."
        )

    financial_settings_locked = current_user_role != "admin"

    if financial_settings_locked:
        st.caption(
            "🔒 Les paramètres financiers sont définis par "
            "l’administrateur et ne peuvent pas être modifiés par "
            "un directeur."
        )

    st.subheader("Taux dollar → euro quotidien")
    rate_status_column, rate_refresh_column = st.columns([3, 1])
    force_rate_refresh = rate_refresh_column.button(
        "🔄 Actualiser maintenant",
        key="parameters_refresh_daily_rate",
        use_container_width=True,
    )
    daily_rate_info = resolve_daily_usd_to_eur_rate(
        database_url=database_url if persistent_settings_available else None,
        updated_by=current_user_email,
        force_refresh=force_rate_refresh,
    )
    rate_status_column.success(
        f"Taux du {daily_rate_info['date']} : "
        f"1 $ = {daily_rate_info['rate']:.6f} €"
    )
    rate_status_column.caption(
        f"Source : {daily_rate_info['source']}. Actualisation automatique "
        "à partir du dernier taux officiel BCE disponible."
    )
    if daily_rate_info["warning"]:
        st.warning(daily_rate_info["warning"])

    with st.form("monthly_parameters"):
        left, right = st.columns(2)

        with left:
            month = st.text_input("Mois calculé", st.session_state.month)
            revenue_usd = st.number_input(
                "Chiffre d’affaires Backstage ($)",
                min_value=0.0,
                value=float(st.session_state.revenue_usd),
                step=100.0,
            )
            usd_to_eur = st.number_input(
                "Taux dollar → euro utilisé aujourd’hui (BCE)",
                min_value=0.0,
                value=float(daily_rate_info["rate"]),
                step=0.0001,
                format="%.6f",
                disabled=True,
                help=(
                    "Ce taux est actualisé automatiquement. Il est "
                    "indépendant de la valeur facture par diamant."
                ),
            )
            other_expenses = st.number_input(
                "Autres dépenses (€)",
                min_value=0.0,
                value=float(st.session_state.other_expenses),
                step=10.0,
            )

        with right:
            creator_level = st.selectbox(
                "Palier Créateurs atteint",
                [4, 7, 9, 11, 13, 15],
                index=[4, 7, 9, 11, 13, 15].index(
                    int(st.session_state.creator_level)
                ),
            )
            consultant_level = st.selectbox(
                "Palier Consultants atteint",
                [5, 7, 9, 11, 13],
                index=[5, 7, 9, 11, 13].index(
                    int(st.session_state.consultant_level)
                ),
            )
            manager_level = st.selectbox(
                "Palier Responsables performance atteint",
                [5, 7, 9, 11, 13],
                index=[5, 7, 9, 11, 13].index(
                    int(st.session_state.manager_level)
                ),
            )
            director_level = st.selectbox(
                "Palier Directeur atteint",
                [4, 5, 7, 8, 10, 11, 13],
                index=[4, 5, 7, 8, 10, 11, 13].index(
                    int(st.session_state.director_level)
                ),
            )

        st.subheader("Paramètres financiers")
        fin1, fin2, fin3 = st.columns(3)
        coin_pack_price = fin1.number_input(
            "Prix de 1 000 pièces TikTok (€)",
            min_value=0.0,
            value=float(st.session_state.coin_pack_price),
            step=0.001,
            format="%.3f",
            disabled=financial_settings_locked,
        )
        invoice_rate = fin2.number_input(
            "Valeur facture par diamant (€)",
            min_value=0.0,
            value=float(st.session_state.invoice_rate),
            step=0.0001,
            format="%.4f",
            disabled=financial_settings_locked,
        )
        charges_rate = fin3.number_input(
            "Charges sur le CA (%)",
            min_value=0.0,
            max_value=100.0,
            value=float(st.session_state.charges_rate),
            step=0.1,
            disabled=financial_settings_locked,
        )

        save_parameters = st.form_submit_button(
            "💾 Enregistrer les paramètres",
            use_container_width=True,
        )

    if save_parameters:
        old_levels = (
            st.session_state.creator_level,
            st.session_state.consultant_level,
            st.session_state.manager_level,
            st.session_state.director_level,
        )
        st.session_state.month = month
        st.session_state.revenue_usd = revenue_usd
        st.session_state.usd_to_eur = float(daily_rate_info["rate"])
        st.session_state.other_expenses = other_expenses
        st.session_state.creator_level = creator_level
        st.session_state.consultant_level = consultant_level
        st.session_state.manager_level = manager_level
        st.session_state.director_level = director_level
        # Double protection : même si le formulaire est manipulé côté
        # navigateur, seuls les administrateurs peuvent enregistrer les
        # paramètres financiers.
        if current_user_role == "admin":
            st.session_state.coin_pack_price = coin_pack_price
            st.session_state.invoice_rate = invoice_rate
            st.session_state.charges_rate = charges_rate

        permanent_save_succeeded = False
        permanent_save_error = None

        if persistent_settings_available:
            scopes_to_save = {
                branch_settings_scope(current_user_email): {
                    key: st.session_state[key]
                    for key in BRANCH_PARAMETER_KEYS
                }
            }

            if current_user_role == "admin":
                scopes_to_save["global:financial"] = {
                    key: st.session_state[key]
                    for key in GLOBAL_FINANCIAL_KEYS
                }

            try:
                save_persistent_scopes(
                    database_url,
                    scopes_to_save,
                    current_user_email,
                )
                permanent_save_succeeded = True
            except Exception:
                permanent_save_error = (
                    "L’enregistrement permanent a échoué. Les valeurs "
                    "restent disponibles dans la session actuelle."
                )

        if old_levels != (
            creator_level,
            consultant_level,
            manager_level,
            director_level,
        ):
            reset_calculations()

        if permanent_save_succeeded:
            st.success(
                "Les paramètres ont été enregistrés définitivement."
            )
        elif permanent_save_error:
            st.error(permanent_save_error)
        else:
            st.warning(
                "Les paramètres ont été enregistrés uniquement pour la "
                "session actuelle. Configurez la base permanente pour les "
                "retrouver après une déconnexion."
            )


elif page == "🛡️ Administration":
    if current_user_role not in {"admin", "director"}:
        st.error("Vous n’avez pas accès à cette page.")
        st.stop()

    st.title("🛡️ Administration des exclusions")
    st.write(
        "Ajoutez, modifiez ou supprimez une adresse à tout moment. "
        "Une ligne supprimée est automatiquement réintégrée aux calculs."
    )
    st.info(
        f"Les exclusions enregistrées ici concernent uniquement "
        f"**{current_user_direction}**. Elles n’affectent aucune autre "
        "direction."
    )

    exclusions_save_notice = st.session_state.pop(
        "exclusions_save_notice",
        None,
    )
    if exclusions_save_notice:
        notice_type, notice_message = exclusions_save_notice
        getattr(st, notice_type)(notice_message)

    if persistent_settings_available:
        st.caption(
            "☁️ Sauvegarde permanente active : les exclusions seront "
            "retrouvées après une déconnexion ou un redémarrage."
        )
    elif persistent_settings_error:
        st.warning(persistent_settings_error)
    else:
        st.warning(
            "La sauvegarde permanente n’est pas configurée. Les exclusions "
            "seront conservées uniquement pendant cette session."
        )

    if st.session_state.backstage_data is not None:
        backstage = st.session_state.backstage_data
        detected_rows = []

        if "Agent" in backstage.columns:
            for value in backstage["Agent"].dropna().unique():
                email = normalize_email(value)
                if email and email != "nan":
                    detected_rows.append(
                        {"Adresse détectée": email, "Emplacement": "Agent"}
                    )

        if "Groupe" in backstage.columns:
            for value in backstage["Groupe"].dropna().unique():
                email = normalize_email(value)
                if email and email != "nan":
                    detected_rows.append(
                        {"Adresse détectée": email, "Emplacement": "Groupe"}
                    )

        if detected_rows:
            detected_dataframe = pd.DataFrame(detected_rows).drop_duplicates()
            detected_dataframe = detected_dataframe.sort_values(
                ["Emplacement", "Adresse détectée"]
            )
            with st.expander(
                "🔎 Voir toutes les adresses détectées dans l’export"
            ):
                st.dataframe(
                    detected_dataframe,
                    use_container_width=True,
                    hide_index=True,
                )
                show_excel_download(
                    detected_dataframe,
                    table_name="adresses_detectees",
                    sheet_name="Adresses détectées",
                    key="download_detected_addresses",
                )

    exclusions_dataframe = pd.DataFrame(
        st.session_state.exclusions,
        columns=[
            "Adresse e-mail",
            "Exclure consultants",
            "Exclure responsables performance",
        ],
    )

    edited_exclusions = st.data_editor(
        exclusions_dataframe,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Adresse e-mail": st.column_config.TextColumn(
                "E-mail ou nom du groupe",
                help=(
                    "Saisissez l’adresse exacte de la colonne Agent ou le "
                    "nom exact présent dans la colonne Groupe."
                ),
            ),
            "Exclure consultants": st.column_config.CheckboxColumn(
                "Exclure consultants",
                default=True,
            ),
            "Exclure responsables performance": (
                st.column_config.CheckboxColumn(
                    "Exclure responsables performance",
                    default=True,
                )
            ),
        },
        key=(
            "exclusions_editor_"
            f"{safe_export_name(current_user_email)}"
        ),
    )
    show_excel_download(
        edited_exclusions,
        table_name="exclusions",
        sheet_name="Exclusions",
        key="download_exclusions",
    )
    st.caption(
        "Toute modification, y compris la suppression d’une ligne ou de "
        "la liste complète, est sauvegardée avec le bouton ci-dessous."
    )
    pending_exclusions = clean_exclusions(
        edited_exclusions.to_dict("records")
    )
    if pending_exclusions != st.session_state.exclusions:
        st.warning(
            "Modifications non enregistrées : cliquez sur le bouton de "
            "sauvegarde pour les conserver après la déconnexion."
        )

    if st.button(
        "💾 Enregistrer les ajouts, modifications et suppressions",
        type="primary",
        use_container_width=True,
    ):
        cleaned_exclusions = pending_exclusions
        st.session_state.exclusions = cleaned_exclusions
        st.session_state.pop("consultant_results", None)
        st.session_state.pop("consultant_signature", None)
        st.session_state.pop("consultant_payment_editor", None)
        st.session_state.pop("responsable_results", None)
        st.session_state.pop("responsable_signature", None)
        st.session_state.pop("responsable_payment_editor", None)

        if persistent_settings_available:
            try:
                save_persistent_scopes(
                    database_url,
                    {
                        exclusions_scope(current_user_email): {
                            "rows": cleaned_exclusions,
                            "saved_at": datetime.now().isoformat(
                                timespec="seconds"
                            ),
                        }
                    },
                    current_user_email,
                )
                if cleaned_exclusions:
                    saved_message = (
                        f"{len(cleaned_exclusions)} exclusion(s) "
                        "enregistrée(s) définitivement pour votre "
                        "direction."
                    )
                else:
                    saved_message = (
                        "Toutes les exclusions ont été supprimées et cette "
                        "suppression a été enregistrée définitivement."
                    )
                st.session_state.exclusions_save_notice = (
                    "success",
                    saved_message,
                )
            except Exception:
                st.session_state.exclusions_save_notice = (
                    "error",
                    "L’enregistrement permanent a échoué. Les exclusions "
                    "restent actives uniquement dans la session actuelle.",
                )
        else:
            st.session_state.exclusions_save_notice = (
                "warning",
                f"{len(cleaned_exclusions)} exclusion(s) enregistrée(s) "
                "uniquement pour cette session.",
            )
        st.rerun()

    if st.session_state.exclusions:
        st.caption(
            "Les exclusions sont actives dans les calculs Consultants et "
            "Responsables performance."
        )
    else:
        st.info("Aucune adresse n’est actuellement exclue.")


elif page == "🔐 Accès collaborateurs":
    if current_user_role != "admin":
        st.error("Cette page est réservée à l’administrateur.")
        st.stop()

    render_brand_hero(
        "Accès collaborateurs",
        "Ajoutez les comptes autorisés, choisissez leur rôle et attribuez "
        "leurs groupes. Cette page et ses réglages restent invisibles pour "
        "tous les autres utilisateurs.",
        "PRO CONSULTING • ADMINISTRATION PRIVÉE",
    )
    st.info(
        "Les quatre directeurs déjà enregistrés conservent leurs accès. "
        "Utilisez cette page pour ajouter et gérer les nouveaux "
        "collaborateurs, notamment les responsables performance."
    )

    access_notice = st.session_state.pop(
        "admin_collaborator_access_notice",
        None,
    )
    if access_notice:
        notice_type, notice_message = access_notice
        getattr(st, notice_type)(notice_message)

    if not persistent_settings_available:
        st.error(
            "La base PostgreSQL doit être connectée pour gérer les accès. "
            "Aucun accès temporaire ne peut être créé."
        )
        if persistent_settings_error:
            st.warning(persistent_settings_error)
        st.stop()

    saved_access_payload = {}
    saved_director_payload = {}
    try:
        authorization_payloads = load_authorization_payloads(database_url)
        saved_access_payload = authorization_payloads[
            collaborator_access_scope()
        ]
        saved_director_payload = authorization_payloads[
            director_management_scope()
        ]
    except Exception:
        st.error(
            "Les accès enregistrés ne peuvent pas être chargés. Aucune "
            "modification n’est autorisée tant que la connexion n’est pas "
            "rétablie."
        )
        st.stop()

    existing_access_rows = clean_collaborator_access_records(
        saved_access_payload.get("collaborators", [])
    )
    known_group_values = set()
    if (
        st.session_state.backstage_data is not None
        and "Groupe" in st.session_state.backstage_data.columns
    ):
        known_group_values.update(
            value
            for value in st.session_state.backstage_data["Groupe"]
            .fillna("")
            .astype(str)
            .str.strip()
            .unique()
            if value and value.casefold() != "nan"
        )
    saved_directors = saved_director_payload.get("directors", {})
    if isinstance(saved_directors, dict):
        for director_row in saved_directors.values():
            if isinstance(director_row, dict):
                known_group_values.update(
                    str(group).strip()
                    for group in director_row.get("groups", [])
                    if str(group).strip()
                )
    for collaborator_row in existing_access_rows:
        known_group_values.update(collaborator_row["groups"])
    available_access_groups = sorted(
        known_group_values,
        key=lambda value: value.casefold(),
    )

    if not available_access_groups:
        st.warning(
            "Aucun groupe n’est encore disponible. Importez le fichier "
            "Backstage administrateur ou enregistrez d’abord les groupes "
            "des quatre directions."
        )

    known_authorized_emails = set(BASE_AUTHORIZED_USERS)
    known_authorized_emails.update(
        row["email"] for row in existing_access_rows
    )
    detected_access_emails = set()
    email_pattern = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")
    if st.session_state.backstage_data is not None:
        for email_column in ("Agent", "Groupe"):
            if email_column not in st.session_state.backstage_data.columns:
                continue
            for value in (
                st.session_state.backstage_data[email_column]
                .fillna("")
                .astype(str)
                .str.strip()
                .unique()
            ):
                detected_email = normalize_email(value)
                if email_pattern.fullmatch(detected_email):
                    detected_access_emails.add(detected_email)
    for exclusion_row in st.session_state.get("exclusions", []):
        detected_email = normalize_email(
            exclusion_row.get("Adresse e-mail", "")
        )
        if email_pattern.fullmatch(detected_email):
            detected_access_emails.add(detected_email)

    detected_access_emails.difference_update(known_authorized_emails)
    manual_email_option = "✍️ Ajouter une adresse manuellement"
    selectable_email_options = sorted(detected_access_emails) + [
        manual_email_option
    ]

    st.subheader("Ajouter un collaborateur")
    with st.form("add_collaborator_access_form", clear_on_submit=True):
        new_collaborator_name = st.text_input(
            "Nom affiché",
            placeholder="Exemple : Marie",
        )
        selected_collaborator_email = st.selectbox(
            "Adresse Google autorisée",
            options=selectable_email_options,
            help=(
                "Sélectionnez une adresse détectée dans l’export Backstage "
                "ou choisissez l’option de saisie manuelle."
            ),
            key="selected_existing_collaborator_email",
        )
        manual_collaborator_email = st.text_input(
            "Adresse absente de la liste ? Ajoutez-la manuellement",
            placeholder="nom@gmail.com",
            help=(
                "Ce champ reste toujours modifiable. Lorsqu’il est rempli, "
                "l’adresse saisie remplace celle du menu déroulant."
            ),
        )
        if str(manual_collaborator_email or "").strip():
            new_collaborator_email = manual_collaborator_email
        elif selected_collaborator_email == manual_email_option:
            new_collaborator_email = ""
        else:
            new_collaborator_email = selected_collaborator_email
        add_role_column, add_groups_column = st.columns(2)
        new_collaborator_role_label = add_role_column.selectbox(
            "Rôle",
            options=list(COLLABORATOR_ROLE_LABELS.values()),
        )
        new_collaborator_groups = add_groups_column.multiselect(
            "Groupes attribués",
            options=available_access_groups,
            placeholder="Sélectionnez un ou plusieurs groupes",
        )
        add_collaborator = st.form_submit_button(
            "➕ Autoriser ce collaborateur",
            type="primary",
            use_container_width=True,
        )

    if add_collaborator:
        normalized_new_email = normalize_email(new_collaborator_email)
        role_by_label = {
            label: role
            for role, label in COLLABORATOR_ROLE_LABELS.items()
        }
        validation_error = None
        if not str(new_collaborator_name or "").strip():
            validation_error = "Le nom du collaborateur est obligatoire."
        elif not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", normalized_new_email):
            validation_error = "L’adresse e-mail indiquée n’est pas valide."
        elif normalized_new_email in known_authorized_emails:
            validation_error = "Cette adresse possède déjà un accès."
        elif not new_collaborator_groups:
            validation_error = "Attribuez au moins un groupe."

        if validation_error:
            st.warning(validation_error)
        else:
            updated_access_rows = existing_access_rows + [
                {
                    "email": normalized_new_email,
                    "name": str(new_collaborator_name).strip(),
                    "role": role_by_label[new_collaborator_role_label],
                    "groups": list(new_collaborator_groups),
                    "active": True,
                }
            ]
            try:
                save_persistent_scopes(
                    database_url,
                    {
                        collaborator_access_scope(): {
                            "collaborators": updated_access_rows,
                            "saved_at": datetime.now().isoformat(
                                timespec="seconds"
                            ),
                        }
                    },
                    current_user_email,
                )
                st.session_state.admin_collaborator_access_notice = (
                    "success",
                    "Le collaborateur est autorisé. Il peut maintenant se "
                    "connecter avec cette adresse Google.",
                )
                st.session_state.pop(
                    "selected_existing_collaborator_email",
                    None,
                )
                st.rerun()
            except Exception:
                st.error(
                    "L’accès n’a pas été enregistré. Le collaborateur "
                    "reste bloqué par sécurité."
                )

    st.divider()
    st.subheader("Collaborateurs ajoutés")
    if not existing_access_rows:
        st.caption(
            "Aucun collaborateur supplémentaire n’a encore été ajouté."
        )
    else:
        role_labels = list(COLLABORATOR_ROLE_LABELS.values())
        edited_access_rows = []
        for access_index, access_row in enumerate(existing_access_rows):
            safe_access_key = safe_export_name(access_row["email"])
            with st.expander(
                f"{access_row['name']} — {access_row['email']}",
                expanded=True,
            ):
                edit_name_column, edit_email_column = st.columns(2)
                edited_name = edit_name_column.text_input(
                    "Nom affiché",
                    value=access_row["name"],
                    key=f"collaborator_name_{safe_access_key}_{access_index}",
                )
                edit_email_column.text_input(
                    "Adresse Google",
                    value=access_row["email"],
                    disabled=True,
                    key=f"collaborator_email_{safe_access_key}_{access_index}",
                )
                edit_role_column, edit_groups_column = st.columns(2)
                current_role_label = COLLABORATOR_ROLE_LABELS[
                    access_row["role"]
                ]
                edited_role_label = edit_role_column.selectbox(
                    "Rôle",
                    options=role_labels,
                    index=role_labels.index(current_role_label),
                    key=f"collaborator_role_{safe_access_key}_{access_index}",
                )
                edited_groups = edit_groups_column.multiselect(
                    "Groupes attribués",
                    options=available_access_groups,
                    default=[
                        group
                        for group in access_row["groups"]
                        if group in available_access_groups
                    ],
                    key=f"collaborator_groups_{safe_access_key}_{access_index}",
                )
                edited_active = st.toggle(
                    "Accès actif",
                    value=bool(access_row["active"]),
                    key=f"collaborator_active_{safe_access_key}_{access_index}",
                    help=(
                        "Désactivez cette option pour bloquer immédiatement "
                        "la prochaine ouverture de l’application."
                    ),
                )
            edited_access_rows.append(
                {
                    "email": access_row["email"],
                    "name": str(edited_name or "").strip(),
                    "role": {
                        label: role
                        for role, label in COLLABORATOR_ROLE_LABELS.items()
                    }[edited_role_label],
                    "groups": list(edited_groups),
                    "active": bool(edited_active),
                }
            )

        save_all_collaborator_access = st.button(
            "💾 Enregistrer les rôles et les groupes",
            key="save_all_collaborator_access",
            type="primary",
            use_container_width=True,
        )
        if save_all_collaborator_access:
            invalid_rows = [
                row
                for row in edited_access_rows
                if not row["name"] or (row["active"] and not row["groups"])
            ]
            if invalid_rows:
                st.warning(
                    "Chaque accès actif doit conserver un nom et au moins "
                    "un groupe attribué."
                )
            else:
                try:
                    save_persistent_scopes(
                        database_url,
                        {
                            collaborator_access_scope(): {
                                "collaborators": edited_access_rows,
                                "saved_at": datetime.now().isoformat(
                                    timespec="seconds"
                                ),
                            }
                        },
                        current_user_email,
                    )
                    st.session_state.admin_collaborator_access_notice = (
                        "success",
                        "Les rôles, groupes et statuts ont été enregistrés "
                        "définitivement.",
                    )
                    st.rerun()
                except Exception:
                    st.error(
                        "La modification n’a pas été enregistrée. Les droits "
                        "précédents sont conservés."
                    )

    st.divider()
    st.subheader("Lien unique de connexion")
    st.write(
        "Tous les collaborateurs utilisent la même adresse Streamlit. "
        "Copiez le lien ci-dessous et partagez-le uniquement avec les "
        "personnes autorisées. Leur rôle sera reconnu automatiquement après "
        "la connexion Google."
    )
    try:
        unique_application_url = str(st.context.url or "").strip()
    except (AttributeError, RuntimeError):
        unique_application_url = ""
    if unique_application_url:
        st.code(unique_application_url, language=None)
        st.link_button(
            "🔗 Ouvrir le lien partagé",
            unique_application_url,
            use_container_width=True,
        )
    else:
        st.caption(
            "L’adresse sera affichée automatiquement lorsque cette version "
            "sera ouverte depuis Streamlit."
        )


elif page == "💎 Créateurs":
    st.title("💎 Calcul des créateurs")
    show_estimation_notice()

    if st.session_state.backstage_data is None:
        st.warning("Importez d’abord un export Backstage.")
        st.stop()

    creator_signature = (
        st.session_state.backstage_filename,
        st.session_state.creator_level,
    )

    if (
        "creator_results" not in st.session_state
        or st.session_state.get("creator_signature") != creator_signature
    ):
        creator_results = calculate_creator_rewards(
            st.session_state.backstage_data,
            int(st.session_state.creator_level),
        )
        creator_results["Rémunération calculée 💎"] = creator_results[
            "Rémunération 💎"
        ]
        creator_results["Inclure rémunération créateur"] = "Oui"
        creator_results["Mode paiement"] = "Diamants"
        st.session_state.creator_results = creator_results
        st.session_state.creator_signature = creator_signature

    creator_results = st.session_state.creator_results.copy()

    if "Rémunération calculée 💎" not in creator_results.columns:
        creator_results["Rémunération calculée 💎"] = creator_results[
            "Rémunération 💎"
        ]
    if "Inclure rémunération créateur" not in creator_results.columns:
        creator_results["Inclure rémunération créateur"] = "Oui"

    st.divider()
    payment_table = creator_results[
        [
            "Pseudo",
            "Groupe",
            "Agent",
            "Diamants",
            "Inclure rémunération créateur",
            "Rémunération 💎",
            "Mode paiement",
        ]
    ].copy()

    edited_payment_table = st.data_editor(
        payment_table,
        use_container_width=True,
        hide_index=True,
        disabled=[
            "Pseudo",
            "Groupe",
            "Agent",
            "Diamants",
            "Rémunération 💎",
        ],
        column_config={
            "Inclure rémunération créateur": (
                st.column_config.SelectboxColumn(
                    "Rémunérer le créateur",
                    options=["Oui", "Non"],
                    required=True,
                    help=(
                        "Choisissez Non pour un créateur parti. Cela annule "
                        "uniquement sa rémunération personnelle, sans "
                        "modifier les calculs du consultant ou du responsable."
                    ),
                )
            ),
            "Mode paiement": st.column_config.SelectboxColumn(
                "Mode paiement",
                options=["Diamants", "Facture €"],
                required=True,
            )
        },
        key="creator_payment_editor",
    )
    show_excel_download(
        edited_payment_table,
        table_name="paiements_createurs",
        sheet_name="Paiements créateurs",
        key="download_creator_payments",
    )

    creator_results["Mode paiement"] = edited_payment_table[
        "Mode paiement"
    ].values
    creator_results["Inclure rémunération créateur"] = (
        edited_payment_table["Inclure rémunération créateur"].values
    )
    creator_results["Rémunération 💎"] = creator_results[
        "Rémunération calculée 💎"
    ].where(
        creator_results["Inclure rémunération créateur"] == "Oui",
        0,
    )
    creator_results = financial_columns(creator_results)
    st.session_state.creator_results = creator_results

    diamond_total = creator_results.loc[
        creator_results["Mode paiement"] == "Diamants",
        "Rémunération 💎",
    ].sum()
    metric1, metric2, metric3, metric4, metric5 = st.columns(5)
    metric1.metric("Créateurs importés", len(creator_results))
    metric2.metric(
        "Créateurs rémunérés",
        int((creator_results["Rémunération 💎"] > 0).sum()),
    )
    metric3.metric("Paiements diamants", f"{diamond_total:,.0f} 💎")
    metric4.metric(
        "Paiements factures",
        f"{creator_results['Facture €'].sum():,.0f} €",
    )
    metric5.metric(
        "Comptés pour la hiérarchie",
        int((creator_results["Compté hiérarchie"] == "Oui").sum()),
    )
    show_financial_summary(creator_results)

    with st.expander("Afficher le détail complet des calculs"):
        st.dataframe(
            creator_results,
            use_container_width=True,
            hide_index=True,
        )
        show_excel_download(
            creator_results,
            table_name="detail_createurs",
            sheet_name="Détail créateurs",
            key="download_creator_details",
        )


elif page == "👥 Consultants":
    st.title("👥 Calcul des consultants")
    show_estimation_notice()

    if st.session_state.backstage_data is None:
        st.warning("Importez d’abord un export Backstage.")
        st.stop()

    creator_signature = (
        st.session_state.backstage_filename,
        st.session_state.creator_level,
    )
    if (
        "creator_results" not in st.session_state
        or st.session_state.get("creator_signature") != creator_signature
    ):
        creator_results = calculate_creator_rewards(
            st.session_state.backstage_data,
            int(st.session_state.creator_level),
        )
        creator_results["Mode paiement"] = "Diamants"
        st.session_state.creator_results = creator_results
        st.session_state.creator_signature = creator_signature

    exclusion_signature = tuple(sorted(excluded_emails("consultants")))
    consultant_signature = (
        st.session_state.backstage_filename,
        st.session_state.creator_level,
        st.session_state.consultant_level,
        exclusion_signature,
    )

    if (
        "consultant_results" not in st.session_state
        or st.session_state.get("consultant_signature")
        != consultant_signature
    ):
        consultant_results = calculate_consultant_rewards(
            creator_results=st.session_state.creator_results,
            consultant_level=int(st.session_state.consultant_level),
            minimum_team_diamonds=200_000,
        )
        consultant_results["Mode paiement"] = "Diamants"
        st.session_state.consultant_results = consultant_results
        st.session_state.consultant_signature = consultant_signature

    consultant_results = st.session_state.consultant_results.copy()

    if consultant_results.empty:
        st.warning("Aucun consultant n’a été détecté dans la colonne Agent.")
        st.stop()

    excluded = excluded_emails("consultants")
    consultant_results["Inclure dans le calcul"] = consultant_results[
        "Consultant"
    ].map(lambda value: "Non" if normalize_email(value) in excluded else "Oui")

    excluded_mask = (
        consultant_results["Inclure dans le calcul"] == "Non"
    )
    consultant_results.loc[
        excluded_mask,
        ["Taux", "Rémunération 💎"],
    ] = 0

    st.divider()
    payment_table = consultant_results[
        [
            "Consultant",
            "Inclure dans le calcul",
            "Créateurs rattachés",
            "Créateurs comptés",
            "Diamants éligibles",
            "Seuil atteint",
            "Taux",
            "Rémunération 💎",
            "Mode paiement",
        ]
    ].copy()

    edited_payment_table = st.data_editor(
        payment_table,
        use_container_width=True,
        hide_index=True,
        disabled=[
            "Consultant",
            "Inclure dans le calcul",
            "Créateurs rattachés",
            "Créateurs comptés",
            "Diamants éligibles",
            "Seuil atteint",
            "Taux",
            "Rémunération 💎",
        ],
        column_config={
            "Mode paiement": st.column_config.SelectboxColumn(
                "Mode paiement",
                options=["Diamants", "Facture €"],
                required=True,
            )
        },
        key="consultant_payment_editor",
    )
    show_excel_download(
        edited_payment_table,
        table_name="paiements_consultants",
        sheet_name="Paiements consultants",
        key="download_consultant_payments",
    )

    consultant_results["Mode paiement"] = edited_payment_table[
        "Mode paiement"
    ].values
    consultant_results = financial_columns(consultant_results)
    st.session_state.consultant_results = consultant_results

    diamond_total = consultant_results.loc[
        consultant_results["Mode paiement"] == "Diamants",
        "Rémunération 💎",
    ].sum()
    metric1, metric2, metric3, metric4, metric5 = st.columns(5)
    metric1.metric("Consultants détectés", len(consultant_results))
    metric2.metric(
        "Consultants rémunérés",
        int((consultant_results["Rémunération 💎"] > 0).sum()),
    )
    metric3.metric(
        "Diamants éligibles",
        f"{consultant_results.loc[~excluded_mask, 'Diamants éligibles'].sum():,.0f} 💎",
    )
    metric4.metric("Paiements diamants", f"{diamond_total:,.0f} 💎")
    metric5.metric(
        "Paiements factures",
        f"{consultant_results['Facture €'].sum():,.0f} €",
    )
    show_financial_summary(consultant_results)

    with st.expander("Afficher le détail complet"):
        st.dataframe(
            consultant_results,
            use_container_width=True,
            hide_index=True,
        )
        show_excel_download(
            consultant_results,
            table_name="detail_consultants",
            sheet_name="Détail consultants",
            key="download_consultant_details",
        )


elif page == "📈 Responsables performance":
    st.title("📈 Calcul des responsables performance")
    show_estimation_notice()

    if st.session_state.backstage_data is None:
        st.warning("Importez d’abord un export Backstage.")
        st.stop()

    creator_signature = (
        st.session_state.backstage_filename,
        st.session_state.creator_level,
    )
    if (
        "creator_results" not in st.session_state
        or st.session_state.get("creator_signature") != creator_signature
    ):
        creator_results = calculate_creator_rewards(
            st.session_state.backstage_data,
            int(st.session_state.creator_level),
        )
        creator_results["Mode paiement"] = "Diamants"
        st.session_state.creator_results = creator_results
        st.session_state.creator_signature = creator_signature

    exclusion_signature = tuple(sorted(excluded_emails("responsables")))
    responsable_signature = (
        st.session_state.backstage_filename,
        st.session_state.creator_level,
        st.session_state.manager_level,
        exclusion_signature,
    )

    if (
        "responsable_results" not in st.session_state
        or st.session_state.get("responsable_signature")
        != responsable_signature
    ):
        responsable_results = calculate_responsable_rewards(
            creator_results=st.session_state.creator_results,
            responsable_level=int(st.session_state.manager_level),
            minimum_group_diamonds=600_000,
        )
        responsable_results["Mode paiement"] = "Diamants"
        st.session_state.responsable_results = responsable_results
        st.session_state.responsable_signature = responsable_signature

    responsable_results = st.session_state.responsable_results.copy()

    if responsable_results.empty:
        st.warning(
            "Aucun responsable performance n’a été détecté dans la "
            "colonne Groupe."
        )
        st.stop()

    excluded = excluded_emails("responsables")
    responsable_results["Inclure dans le calcul"] = responsable_results[
        "Responsable performance"
    ].map(lambda value: "Non" if normalize_email(value) in excluded else "Oui")

    excluded_mask = responsable_results["Inclure dans le calcul"] == "Non"
    responsable_results.loc[
        excluded_mask,
        ["Taux", "Rémunération 💎"],
    ] = 0

    st.divider()
    payment_table = responsable_results[
        [
            "Responsable performance",
            "Inclure dans le calcul",
            "Créateurs rattachés",
            "Créateurs comptés",
            "Diamants éligibles",
            "Seuil atteint",
            "Taux",
            "Rémunération 💎",
            "Mode paiement",
        ]
    ].copy()

    edited_payment_table = st.data_editor(
        payment_table,
        use_container_width=True,
        hide_index=True,
        disabled=[
            "Responsable performance",
            "Inclure dans le calcul",
            "Créateurs rattachés",
            "Créateurs comptés",
            "Diamants éligibles",
            "Seuil atteint",
            "Taux",
            "Rémunération 💎",
        ],
        column_config={
            "Mode paiement": st.column_config.SelectboxColumn(
                "Mode paiement",
                options=["Diamants", "Facture €"],
                required=True,
            )
        },
        key="responsable_payment_editor",
    )
    show_excel_download(
        edited_payment_table,
        table_name="paiements_responsables",
        sheet_name="Paiements responsables",
        key="download_responsable_payments",
    )

    responsable_results["Mode paiement"] = edited_payment_table[
        "Mode paiement"
    ].values
    responsable_results = financial_columns(responsable_results)
    st.session_state.responsable_results = responsable_results

    diamond_total = responsable_results.loc[
        responsable_results["Mode paiement"] == "Diamants",
        "Rémunération 💎",
    ].sum()
    metric1, metric2, metric3, metric4, metric5 = st.columns(5)
    metric1.metric("Responsables détectés", len(responsable_results))
    metric2.metric(
        "Responsables rémunérés",
        int((responsable_results["Rémunération 💎"] > 0).sum()),
    )
    metric3.metric(
        "Diamants éligibles",
        f"{responsable_results.loc[~excluded_mask, 'Diamants éligibles'].sum():,.0f} 💎",
    )
    metric4.metric("Paiements diamants", f"{diamond_total:,.0f} 💎")
    metric5.metric(
        "Paiements factures",
        f"{responsable_results['Facture €'].sum():,.0f} €",
    )
    show_financial_summary(responsable_results)

    with st.expander("Afficher le détail complet"):
        st.dataframe(
            responsable_results,
            use_container_width=True,
            hide_index=True,
        )
        show_excel_download(
            responsable_results,
            table_name="detail_responsables",
            sheet_name="Détail responsables",
            key="download_responsable_details",
        )


elif page == "🎁 Suivi récompenses":
    st.title("🎁 Suivi collectif des récompenses")
    st.info(
        "Ce tableau est commun à l’administration, aux directions et aux "
        "responsables performance autorisés. "
        "Les créateurs et leurs récompenses sont repris automatiquement ; "
        "les responsables ne peuvent modifier que les créateurs de leurs "
        "groupes. Un créateur "
        "payé en Facture € apparaît automatiquement à 0 💎 dans ce suivi."
    )
    if current_user_role in {"admin", "director"}:
        st.caption(
            "L’administration et les directeurs peuvent modifier les champs "
            "de suivi de toutes les lignes. Les montants automatiques restent "
            "protégés ; les décisions Validée / Refusée sont réservées à "
            "FONDATEUR ADMIN."
        )
    tracking_reset_notice = st.session_state.pop(
        "reward_tracking_reset_notice",
        None,
    )
    if tracking_reset_notice:
        st.success(tracking_reset_notice)

    refresh_tracking = st.button(
        "🔄 Actualiser les données partagées",
        key="refresh_collective_reward_tracking",
        use_container_width=True,
        help=(
            "Recharge immédiatement la dernière sauvegarde effectuée par "
            "un collaborateur autorisé."
        ),
    )
    if refresh_tracking:
        st.session_state.pop("reward_tracking_loaded_scope", None)
        st.session_state.pop("reward_tracking_editor", None)
        st.session_state.pop("manager_reward_tracking_editor", None)
        st.session_state.pop("reward_tracking_table", None)
        st.session_state.pop("reward_tracking_baseline_rows", None)
        st.session_state.pop("reward_tracking_active_month", None)
        st.session_state.pop("reward_tracking_last_saved_at", None)
        st.session_state.pop("reward_tracking_last_saved_by", None)
        st.session_state.pop("reward_tracking_pending_new_month", None)

    if not persistent_settings_available:
        st.error(
            "⛔ Le partage entre les comptes est indisponible tant que la "
            "base PostgreSQL n'est pas connectée dans les Secrets "
            "Streamlit. L'enregistrement collectif est désactivé pour "
            "éviter de faire croire qu'une saisie locale est partagée."
        )

    creator_results_for_tracking = pd.DataFrame(
        columns=["Pseudo", "Groupe", "Rémunération 💎"]
    )
    if st.session_state.backstage_data is not None:
        creator_signature = (
            st.session_state.backstage_filename,
            st.session_state.creator_level,
        )
        if (
            "creator_results" not in st.session_state
            or st.session_state.get("creator_signature")
            != creator_signature
        ):
            creator_results = calculate_creator_rewards(
                st.session_state.backstage_data,
                int(st.session_state.creator_level),
            )
            creator_results["Rémunération calculée 💎"] = creator_results[
                "Rémunération 💎"
            ]
            creator_results["Inclure rémunération créateur"] = "Oui"
            creator_results["Mode paiement"] = "Diamants"
            st.session_state.creator_results = creator_results
            st.session_state.creator_signature = creator_signature

        creator_results_for_tracking = (
            st.session_state.creator_results.copy()
        )

    tracking_scope = reward_tracking_scope()
    if (
        st.session_state.get("reward_tracking_loaded_scope")
        != tracking_scope
    ):
        saved_tracking_payload = {}
        saved_tracking_rows = []
        if persistent_settings_available:
            try:
                saved_tracking_payload = load_persistent_scope(
                    database_url,
                    tracking_scope,
                )
                saved_tracking_rows = saved_tracking_payload.get("rows", [])

                # Migration transparente depuis l'ancien stockage séparé
                # par mois. Elle ne s'exécute qu'une seule fois, lorsque le
                # registre collectif unique est encore vide.
                if not saved_tracking_payload:
                    legacy_payload = load_persistent_scope(
                        database_url,
                        legacy_reward_tracking_scope(st.session_state.month),
                    )
                    legacy_rows = legacy_payload.get("rows", [])
                    if legacy_rows:
                        saved_tracking_payload = {
                            "month": legacy_payload.get(
                                "month",
                                st.session_state.month,
                            ),
                            "rows": clean_reward_tracking_rows(legacy_rows),
                            "saved_at": legacy_payload.get("saved_at"),
                            "saved_by": legacy_payload.get("saved_by"),
                        }
                        save_persistent_scopes(
                            database_url,
                            {tracking_scope: saved_tracking_payload},
                            current_user_email,
                        )
                        saved_tracking_rows = saved_tracking_payload["rows"]
            except Exception:
                st.warning(
                    "Le suivi collectif enregistré ne peut pas être chargé "
                    "pour le moment."
                )

        saved_tracking_month = str(
            saved_tracking_payload.get("month") or ""
        ).strip()
        requested_tracking_month = str(st.session_state.month or "").strip()
        pending_new_month = bool(
            saved_tracking_payload
            and st.session_state.backstage_data is not None
            and saved_tracking_month
            and requested_tracking_month
            and saved_tracking_month.casefold()
            != requested_tracking_month.casefold()
        )
        if pending_new_month:
            # Le nouvel export ne doit jamais recuperer les dates, heures,
            # decisions ou montants manuels du mois precedent.
            saved_tracking_rows = []

        st.session_state.reward_tracking_table = (
            build_reward_tracking_table(
                creator_results_for_tracking,
                saved_tracking_rows,
            )
        )
        st.session_state.reward_tracking_baseline_rows = (
            st.session_state.reward_tracking_table.to_dict("records")
        )
        st.session_state.reward_tracking_active_month = (
            requested_tracking_month
            if pending_new_month
            else saved_tracking_payload.get("month") or st.session_state.month
        )
        st.session_state.reward_tracking_last_saved_at = (
            saved_tracking_payload.get("saved_at")
        )
        st.session_state.reward_tracking_last_saved_by = (
            saved_tracking_payload.get("saved_by")
        )
        st.session_state.reward_tracking_pending_new_month = pending_new_month
        st.session_state.reward_tracking_loaded_scope = tracking_scope
        st.session_state.pop("reward_tracking_editor", None)
        st.session_state.pop("manager_reward_tracking_editor", None)
    else:
        st.session_state.reward_tracking_table = (
            build_reward_tracking_table(
                creator_results_for_tracking,
                st.session_state.get(
                    "reward_tracking_table",
                    pd.DataFrame(columns=REWARD_TRACKING_COLUMNS),
                ).to_dict("records"),
            )
        )

    active_tracking_month = st.session_state.get(
        "reward_tracking_active_month",
        st.session_state.month,
    )
    if persistent_settings_available:
        st.success(
            "☁️ Registre collectif unique actif : tous les collaborateurs "
            "autorisés utilisent le même tableau."
        )
    if st.session_state.get("reward_tracking_pending_new_month"):
        st.info(
            f"🗓️ Nouveau mois détecté : {active_tracking_month}. "
            "Le suivi précédent sera archivé lors du premier "
            "enregistrement confirmé ; ses saisies ne seront pas reprises."
        )
    last_tracking_save = st.session_state.get(
        "reward_tracking_last_saved_at"
    )
    last_tracking_author = normalize_email(
        st.session_state.get("reward_tracking_last_saved_by")
    )
    if last_tracking_save:
        last_tracking_author_name = AUTHORIZED_USERS.get(
            last_tracking_author,
            {},
        ).get("name", last_tracking_author or "un utilisateur autorisé")
        st.caption(
            f"Mois collectif actif : {active_tracking_month} • "
            f"Dernière sauvegarde : {last_tracking_save} • "
            f"Par : {last_tracking_author_name}"
        )
    else:
        st.caption(f"Mois collectif actif : {active_tracking_month}")

    tracking_table = st.session_state.reward_tracking_table.copy()
    if tracking_table.empty:
        st.warning(
            "Le suivi collectif est vide. Importez un export Backstage pour "
            "ajouter automatiquement les créateurs."
        )
        st.stop()

    tracking_column_config = {
            "Date": st.column_config.TextColumn(
                "Date",
                help="À compléter manuellement, par exemple 05/08/2026.",
            ),
            "Heure": st.column_config.TextColumn(
                "Heure",
                help="À compléter manuellement, par exemple 18:30.",
            ),
            "Type d’événement": st.column_config.SelectboxColumn(
                "Type d’événement",
                options=["", "Live", "Match"],
                required=False,
            ),
            "Créateur": st.column_config.TextColumn("Créateur"),
            "Groupe": st.column_config.TextColumn(
                "Groupe",
                help="Groupe utilisé pour sécuriser les droits de saisie.",
            ),
            "Récompense créateur": st.column_config.NumberColumn(
                "Récompense créateur",
                min_value=0,
                step=100,
                format="%d 💎",
            ),
            "Rémunération consultant / responsable": (
                st.column_config.NumberColumn(
                    "Rémunération consultant / responsable",
                    min_value=0,
                    step=100,
                    format="%d 💎",
                    help="Montant à saisir manuellement.",
                )
            ),
            "Total récompense": st.column_config.NumberColumn(
                "Total récompense",
                min_value=0,
                step=100,
                format="%d 💎",
                help="Calcul automatique et non modifiable.",
            ),
            "Récompense validée": st.column_config.CheckboxColumn(
                "✅ Validée",
                width="small",
                help=(
                    "Seul FONDATEUR ADMIN peut confirmer l’envoi. La ligne "
                    "devient verte."
                ),
                default=False,
            ),
            "Récompense refusée": st.column_config.CheckboxColumn(
                "⛔ Refusée",
                width="small",
                help=(
                    "Seul FONDATEUR ADMIN peut refuser la récompense. La "
                    "ligne devient rouge."
                ),
                default=False,
            ),
        }

    rows_to_save = []
    editable_groups_for_save = None
    allow_new_tracking_rows = True

    if current_user_role == "performance_manager":
        st.subheader("Vue collective complète")
        st.dataframe(
            tracking_table.style.apply(style_reward_status_rows, axis=1),
            use_container_width=True,
            hide_index=True,
            column_config=tracking_column_config,
        )

        manager_group_keys = {
            normalize_group_name(group)
            for group in current_user_groups
            if normalize_group_name(group)
        }
        st.subheader("Créateurs modifiables — mes groupes")
        if current_user_groups:
            st.caption("Groupes attribués : " + ", ".join(current_user_groups))
        manager_editable_table = tracking_table[
            tracking_table["Groupe"]
            .map(normalize_group_name)
            .isin(manager_group_keys)
        ].copy().reset_index(drop=True)

        if manager_editable_table.empty:
            st.warning(
                "Aucun créateur du suivi collectif ne correspond encore à "
                "vos groupes. Demandez à l’administrateur de vérifier votre "
                "affectation ou aux directions d’enregistrer le suivi avec "
                "la colonne Groupe."
            )
            edited_manager_tracking_table = manager_editable_table
        else:
            edited_manager_tracking_table = st.data_editor(
                manager_editable_table.style.apply(
                    style_reward_status_rows,
                    axis=1,
                ),
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                disabled=[
                    "Créateur",
                    "Groupe",
                    "Récompense créateur",
                    "Total récompense",
                    "Récompense validée",
                    "Récompense refusée",
                ],
                column_config=tracking_column_config,
                key="manager_reward_tracking_editor",
            )

        rows_to_save = clean_reward_tracking_rows(
            edited_manager_tracking_table.to_dict("records")
        )
        manager_updates_by_creator = {
            reward_creator_key(row["Créateur"]): row
            for row in rows_to_save
        }
        cleaned_tracking_rows = []
        for collective_row in clean_reward_tracking_rows(
            tracking_table.to_dict("records")
        ):
            cleaned_tracking_rows.append(
                manager_updates_by_creator.get(
                    reward_creator_key(collective_row["Créateur"]),
                    collective_row,
                )
            )
        editable_groups_for_save = current_user_groups
        allow_new_tracking_rows = False
    else:
        disabled_tracking_columns = [
            "Créateur",
            "Groupe",
            "Récompense créateur",
            "Total récompense",
        ]
        if current_user_role != "admin":
            disabled_tracking_columns.extend(
                ["Récompense validée", "Récompense refusée"]
            )

        edited_tracking_table = st.data_editor(
            tracking_table.style.apply(style_reward_status_rows, axis=1),
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            disabled=disabled_tracking_columns,
            column_config=tracking_column_config,
            key="reward_tracking_editor",
            on_change=synchronize_reward_tracking_editor,
        )
        synchronized_tracking_table = st.session_state.get(
            "reward_tracking_table",
            edited_tracking_table,
        )
        cleaned_tracking_rows = clean_reward_tracking_rows(
            synchronized_tracking_table.to_dict("records")
        )
        rows_to_save = cleaned_tracking_rows

    cleaned_tracking_table = pd.DataFrame(
        cleaned_tracking_rows,
        columns=REWARD_TRACKING_COLUMNS,
    )
    st.session_state.reward_tracking_table = cleaned_tracking_table

    decided_tracking_table = cleaned_tracking_table[
        cleaned_tracking_table["Récompense validée"]
        | cleaned_tracking_table["Récompense refusée"]
    ].copy()
    if (
        current_user_role in {"admin", "director"}
        and not decided_tracking_table.empty
    ):
        with st.expander(
            "🎨 Aperçu intégral des lignes validées et refusées",
            expanded=True,
        ):
            st.dataframe(
                decided_tracking_table.style.apply(
                    style_reward_status_rows,
                    axis=1,
                ),
                use_container_width=True,
                hide_index=True,
                column_config=tracking_column_config,
            )

    total_creator_rewards = int(
        cleaned_tracking_table["Récompense créateur"].sum()
    )
    total_hierarchy_rewards = int(
        cleaned_tracking_table[
            "Rémunération consultant / responsable"
        ].sum()
    )
    total_rewards = int(
        cleaned_tracking_table["Total récompense"].sum()
    )
    metric1, metric2, metric3, metric4 = st.columns(4)
    metric1.metric("Créateurs suivis", len(cleaned_tracking_table))
    metric2.metric(
        "Récompenses créateurs",
        f"{total_creator_rewards:,.0f} 💎",
    )
    metric3.metric(
        "Consultants / responsables",
        f"{total_hierarchy_rewards:,.0f} 💎",
    )
    metric4.metric("Total à envoyer", f"{total_rewards:,.0f} 💎")

    save_column, download_column = st.columns(2)
    if save_column.button(
        "💾 Enregistrer le suivi collectif",
        key="save_collective_reward_tracking",
        type="primary",
        use_container_width=True,
        disabled=(
            not persistent_settings_available
            or (
                current_user_role == "performance_manager"
                and not rows_to_save
            )
        ),
    ):
        if persistent_settings_available:
            try:
                local_creator_names = (
                    creator_results_for_tracking["Pseudo"]
                    .fillna("")
                    .astype(str)
                    .tolist()
                    if "Pseudo" in creator_results_for_tracking.columns
                    else []
                )
                confirmed_payload, started_new_month = (
                    save_collective_reward_tracking(
                        database_url=database_url,
                        scope=tracking_scope,
                        local_rows=rows_to_save,
                        baseline_rows=st.session_state.get(
                            "reward_tracking_baseline_rows",
                            [],
                        ),
                        local_creators=local_creator_names,
                        requested_month=active_tracking_month,
                        updated_by=current_user_email,
                        can_update_reward_status=(
                            current_user_role == "admin"
                        ),
                        editable_groups=editable_groups_for_save,
                        allow_new_rows=allow_new_tracking_rows,
                    )
                )
                merged_tracking_rows = confirmed_payload.get("rows", [])
                saved_at = confirmed_payload.get("saved_at")
                shared_active_month = confirmed_payload.get("month")
                st.session_state.reward_tracking_table = (
                    build_reward_tracking_table(
                        creator_results_for_tracking,
                        merged_tracking_rows,
                    )
                )
                st.session_state.reward_tracking_baseline_rows = (
                    st.session_state.reward_tracking_table.to_dict("records")
                )
                st.session_state.reward_tracking_active_month = (
                    shared_active_month
                )
                st.session_state.reward_tracking_last_saved_at = saved_at
                st.session_state.reward_tracking_last_saved_by = (
                    current_user_email
                )
                st.session_state.reward_tracking_pending_new_month = False
                if started_new_month:
                    st.success(
                        f"Le suivi {shared_active_month} est enregistré et "
                        "confirmé. Le mois précédent a été archivé sans "
                        "mélange de données."
                    )
                else:
                    st.success(
                        "Le suivi collectif est enregistré, fusionné et "
                        "confirmé par la base PostgreSQL."
                    )
            except Exception as error:
                st.error(
                    "La sauvegarde permanente a échoué et n'a pas été "
                    "annoncée comme réussie. Ne fermez pas la session avant "
                    "d’avoir téléchargé le fichier Excel. "
                    f"Détail technique : {type(error).__name__}."
                )
        else:
            if persistent_settings_error:
                st.error(persistent_settings_error)
            else:
                st.error(
                    "La section [database] avec son URL PostgreSQL est "
                    "absente des Secrets Streamlit : le suivi ne peut pas "
                    "être sauvegardé après déconnexion."
                )

    download_column.download_button(
        label="⬇️ Télécharger pour Excel / Google Sheets",
        data=reward_tracking_to_excel(cleaned_tracking_table),
        file_name=export_filename("suivi_collectif_recompenses"),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        key="download_collective_reward_tracking",
        use_container_width=True,
        on_click="ignore",
    )
    st.caption(
        "Dans le fichier téléchargé, le choix Live/Match est conservé et "
        "la colonne Total récompense contient une formule automatique. Le "
        "fichier .xlsx peut être importé directement dans Google Sheets."
    )

    if current_user_role in {"admin", "director"}:
        tracking_delete_generation = int(
            st.session_state.get("reward_tracking_delete_generation", 0)
        )
        tracking_month_key = safe_export_name(active_tracking_month)
        with st.expander(
            f"🗑️ Réinitialiser le suivi collectif — {active_tracking_month}"
        ):
            st.warning(
                "Cette suppression est collective : les informations "
                "manuelles de ce mois seront effacées pour tous. Les "
                "récompenses automatiques seront recalculées à partir du "
                "fichier Backstage actuellement chargé."
            )
            confirm_tracking_deletion = st.checkbox(
                "Je confirme la réinitialisation du suivi collectif de ce mois.",
                key=(
                    "confirm_reward_tracking_deletion_"
                    f"{tracking_month_key}_{tracking_delete_generation}"
                ),
            )
            if st.button(
                "Réinitialiser définitivement ce mois",
                key=(
                    "delete_reward_tracking_"
                    f"{tracking_month_key}_{tracking_delete_generation}"
                ),
                disabled=not confirm_tracking_deletion,
                use_container_width=True,
            ):
                if persistent_settings_available:
                    try:
                        reset_saved_at = datetime.now().isoformat(
                            timespec="seconds"
                        )
                        save_persistent_scopes(
                            database_url,
                            {
                                tracking_scope: {
                                    "month": st.session_state.month,
                                    "rows": [],
                                    "saved_at": reset_saved_at,
                                    "saved_by": current_user_email,
                                }
                            },
                            current_user_email,
                        )
                        st.session_state.pop(
                            "reward_tracking_loaded_scope",
                            None,
                        )
                        st.session_state.pop("reward_tracking_editor", None)
                        st.session_state.pop(
                            "manager_reward_tracking_editor",
                            None,
                        )
                        st.session_state.pop("reward_tracking_table", None)
                        st.session_state.pop(
                            "reward_tracking_baseline_rows",
                            None,
                        )
                        st.session_state.pop(
                            "reward_tracking_active_month",
                            None,
                        )
                        st.session_state.pop(
                            "reward_tracking_last_saved_at",
                            None,
                        )
                        st.session_state.pop(
                            "reward_tracking_last_saved_by",
                            None,
                        )
                        st.session_state.reward_tracking_delete_generation = (
                            tracking_delete_generation + 1
                        )
                        st.session_state.reward_tracking_reset_notice = (
                            "Le suivi collectif du mois a été réinitialisé."
                        )
                        st.rerun()
                    except Exception:
                        st.error(
                            "La réinitialisation a échoué. Les informations "
                            "enregistrées ont été conservées."
                        )
                elif persistent_settings_error:
                    st.error(persistent_settings_error)
                else:
                    st.error(
                        "La base PostgreSQL n’est pas configurée : aucune "
                        "suppression collective permanente n’est possible."
                    )


elif page == "🏢 Directeur de branche":
    st.title("🏢 Calcul du directeur de branche")
    show_estimation_notice()

    if st.session_state.backstage_data is None:
        st.warning("Importez d’abord un export Backstage.")
        st.stop()

    creator_signature = (
        st.session_state.backstage_filename,
        st.session_state.creator_level,
    )
    if (
        "creator_results" not in st.session_state
        or st.session_state.get("creator_signature") != creator_signature
    ):
        creator_results = calculate_creator_rewards(
            st.session_state.backstage_data,
            int(st.session_state.creator_level),
        )
        creator_results["Mode paiement"] = "Diamants"
        st.session_state.creator_results = creator_results
        st.session_state.creator_signature = creator_signature

    all_creator_results = st.session_state.creator_results.copy()
    available_groups = sorted(
        value
        for value in all_creator_results["Groupe"]
        .fillna("")
        .astype(str)
        .str.strip()
        .unique()
        if value and value.lower() != "nan"
    )

    if current_user_role == "admin":
        st.info(
            "Cet espace privé vous permet d’affecter les groupes aux "
            "quatre directeurs et d’afficher leurs factures en une seule "
            "fois. Ces affectations sécurisent aussi leurs imports : un "
            "groupe manquant ou non autorisé bloque entièrement le fichier. "
            "Les directeurs ne voient pas cette vue d’ensemble."
        )

        if not st.session_state.get("admin_director_management_loaded"):
            saved_director_config = {}
            if persistent_settings_available:
                try:
                    saved_payload = load_persistent_scope(
                        database_url,
                        director_management_scope(),
                    )
                    saved_director_config = saved_payload.get(
                        "directors",
                        {},
                    )
                except Exception:
                    st.warning(
                        "Les affectations enregistrées n’ont pas pu être "
                        "chargées. Vous pouvez poursuivre dans cette session."
                    )

            st.session_state.admin_director_management_config = (
                clean_director_management_config(
                    saved_director_config,
                    available_groups,
                )
            )
            st.session_state.admin_director_management_loaded = True

        saved_director_config = clean_director_management_config(
            st.session_state.get(
                "admin_director_management_config",
                {},
            ),
            available_groups,
        )

        st.subheader("Taux de conversion quotidien")
        rate_column, refresh_column = st.columns([3, 1])
        refresh_director_rate = refresh_column.button(
            "🔄 Actualiser le taux",
            key="admin_director_refresh_daily_rate",
            use_container_width=True,
        )
        director_rate_info = resolve_daily_usd_to_eur_rate(
            database_url=(
                database_url if persistent_settings_available else None
            ),
            updated_by=current_user_email,
            force_refresh=refresh_director_rate,
        )
        rate_column.success(
            f"Taux du {director_rate_info['date']} : "
            f"1 $ = {director_rate_info['rate']:.6f} € "
            f"— {director_rate_info['source']}"
        )
        if director_rate_info["warning"]:
            st.warning(director_rate_info["warning"])

        st.subheader("Affectation des groupes et chiffres d’affaires")
        edited_director_config = {}

        for profile in DIRECTOR_MANAGEMENT_PROFILES:
            email = profile["email"]
            saved_row = saved_director_config[email]
            key_suffix = re.sub(r"[^a-z0-9]+", "_", profile["name"].lower())

            with st.expander(
                f"{profile['name']} — {profile['direction']}",
                expanded=True,
            ):
                selected_director_groups = st.multiselect(
                    "Groupes en gestion",
                    options=available_groups,
                    default=saved_row["groups"],
                    key=f"admin_director_groups_{key_suffix}",
                )
                revenue_column, expenses_column = st.columns(2)
                director_revenue_usd = revenue_column.number_input(
                    "Chiffre d’affaires de la branche ($)",
                    min_value=0.0,
                    value=float(saved_row["revenue_usd"]),
                    step=100.0,
                    key=f"admin_director_revenue_{key_suffix}",
                )
                director_other_expenses = expenses_column.number_input(
                    "Autres dépenses de la branche (€)",
                    min_value=0.0,
                    value=float(saved_row["other_expenses"]),
                    step=10.0,
                    key=f"admin_director_expenses_{key_suffix}",
                )

            edited_director_config[email] = {
                "groups": list(selected_director_groups),
                "revenue_usd": float(director_revenue_usd),
                "other_expenses": float(director_other_expenses),
            }

        group_owners = {}
        for profile in DIRECTOR_MANAGEMENT_PROFILES:
            for group in edited_director_config[profile["email"]]["groups"]:
                group_owners.setdefault(group, []).append(profile["name"])

        duplicate_groups = {
            group: owners
            for group, owners in group_owners.items()
            if len(owners) > 1
        }
        if duplicate_groups:
            duplicate_details = "; ".join(
                f"{group} : {', '.join(owners)}"
                for group, owners in sorted(duplicate_groups.items())
            )
            st.error(
                "Un groupe ne peut appartenir qu’à un seul directeur. "
                f"Corrigez les doublons suivants : {duplicate_details}."
            )

        save_director_config = st.button(
            "💾 Enregistrer les quatre directions",
            key="save_admin_director_management",
            type="primary",
            use_container_width=True,
            disabled=bool(duplicate_groups),
        )
        if save_director_config:
            st.session_state.admin_director_management_config = (
                edited_director_config
            )
            if persistent_settings_available:
                try:
                    save_persistent_scopes(
                        database_url,
                        {
                            director_management_scope(): {
                                "directors": edited_director_config,
                                "saved_at": datetime.now().isoformat(
                                    timespec="seconds"
                                ),
                            }
                        },
                        current_user_email,
                    )
                    st.success(
                        "Les groupes, chiffres d’affaires et dépenses des "
                        "quatre directions sont enregistrés définitivement."
                    )
                except Exception:
                    st.error(
                        "Les valeurs restent utilisables dans cette session, "
                        "mais leur sauvegarde permanente a échoué."
                    )
            else:
                if persistent_settings_error:
                    st.error(persistent_settings_error)
                else:
                    st.warning(
                        "La section [database] avec son URL PostgreSQL est "
                        "absente des Secrets Streamlit. Les valeurs restent "
                        "uniquement dans cette session."
                    )

        if duplicate_groups:
            st.stop()

        director_overview = build_director_overview(
            all_creator_results=all_creator_results,
            director_config=edited_director_config,
            usd_to_eur=director_rate_info["rate"],
        )
        st.divider()
        st.subheader("Vue d’ensemble des quatre directeurs")
        overview1, overview2, overview3, overview4 = st.columns(4)
        overview1.metric(
            "Groupes attribués",
            len(group_owners),
        )
        overview2.metric(
            "CA total saisi",
            f"{director_overview['CA branche ($)'].sum():,.2f} $",
        )
        overview3.metric(
            "CA total converti",
            f"{director_overview['CA converti (€)'].sum():,.2f} €",
        )
        overview4.metric(
            "Total factures directeurs",
            f"{director_overview['Montant facture directeur (€)'].sum():,.2f} €",
        )

        st.dataframe(
            director_overview,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Diamants générés": st.column_config.NumberColumn(
                    format="%.0f 💎"
                ),
                "CA branche ($)": st.column_config.NumberColumn(
                    format="%.2f $"
                ),
                "CA converti (€)": st.column_config.NumberColumn(
                    format="%.2f €"
                ),
                "Charges (€)": st.column_config.NumberColumn(
                    format="%.2f €"
                ),
                "Coût créateurs (€)": st.column_config.NumberColumn(
                    format="%.2f €"
                ),
                "Coût consultants (€)": st.column_config.NumberColumn(
                    format="%.2f €"
                ),
                "Coût responsables (€)": st.column_config.NumberColumn(
                    format="%.2f €"
                ),
                "Autres dépenses (€)": st.column_config.NumberColumn(
                    format="%.2f €"
                ),
                "Bénéfice avant directeur (€)": (
                    st.column_config.NumberColumn(format="%.2f €")
                ),
                "Taux directeur": st.column_config.NumberColumn(
                    format="%.1f %%"
                ),
                "Montant facture directeur (€)": (
                    st.column_config.NumberColumn(format="%.2f €")
                ),
                "Bénéfice restant (€)": st.column_config.NumberColumn(
                    format="%.2f €"
                ),
            },
        )
        show_excel_download(
            director_overview,
            table_name="factures_quatre_directeurs",
            sheet_name="Factures directeurs",
            key="download_four_directors_overview",
            label="⬇️ Télécharger la vue des quatre directeurs (Excel)",
        )
        st.caption(
            "Les montants sont recalculés avec le taux BCE affiché et le "
            "palier Directeur enregistré. Un directeur sans groupe attribué "
            "reste à 0 € afin d’éviter une facture sans équipe rattachée."
        )
        st.stop()

    st.subheader("Composition de la branche")
    selected_groups = st.multiselect(
        "Cochez les responsables performance appartenant à cette branche",
        options=available_groups,
        default=available_groups,
        help=(
            "Si l’export contient uniquement la branche du directeur, "
            "laissez tous les responsables cochés."
        ),
        key="director_selected_groups",
    )

    if not selected_groups:
        st.warning("Sélectionnez au moins un responsable performance.")
        st.stop()

    st.subheader("Taux de conversion dollar → euro")
    rate_col1, rate_col2 = st.columns([3, 1])
    use_ecb_rate = rate_col1.checkbox(
        "Utiliser automatiquement le dernier taux officiel BCE disponible",
        value=bool(st.session_state.use_ecb_rate),
        key="director_use_ecb_rate",
    )
    st.session_state.use_ecb_rate = use_ecb_rate

    force_rate_refresh = rate_col2.button(
        "🔄 Actualiser le taux",
        use_container_width=True,
        disabled=not use_ecb_rate,
    )

    active_usd_to_eur = float(st.session_state.usd_to_eur)
    rate_source = "Taux manuel de secours"
    rate_date = datetime.now().strftime("%Y-%m-%d")

    if use_ecb_rate:
        daily_rate_info = resolve_daily_usd_to_eur_rate(
            database_url=(
                database_url if persistent_settings_available else None
            ),
            updated_by=current_user_email,
            force_refresh=force_rate_refresh,
        )
        active_usd_to_eur = float(daily_rate_info["rate"])
        rate_date = daily_rate_info["date"]
        rate_source = daily_rate_info["source"]
        st.success(
            f"Taux du {rate_date} : "
            f"1 $ = {active_usd_to_eur:.6f} €"
        )
        if daily_rate_info["warning"]:
            st.warning(daily_rate_info["warning"])
    else:
        active_usd_to_eur = st.number_input(
            "Taux manuel : valeur de 1 dollar en euros",
            min_value=0.0,
            value=float(st.session_state.usd_to_eur),
            step=0.0001,
            format="%.6f",
            key="director_manual_rate",
        )

    st.caption(
        f"Source utilisée : {rate_source}. Le taux BCE est un taux de "
        "référence quotidien et peut différer du taux réellement appliqué "
        "par la banque."
    )

    input1, input2 = st.columns(2)
    revenue_usd = input1.number_input(
        "Chiffre d’affaires Backstage de la branche ($)",
        min_value=0.0,
        value=float(st.session_state.director_branch_revenue_usd),
        step=100.0,
        key="director_revenue_input",
    )
    branch_other_expenses = input2.number_input(
        "Autres dépenses propres à la branche (€)",
        min_value=0.0,
        value=float(st.session_state.director_branch_other_expenses),
        step=10.0,
        key="director_expenses_input",
    )
    st.session_state.director_branch_revenue_usd = revenue_usd
    st.session_state.director_branch_other_expenses = branch_other_expenses

    branch_creators = all_creator_results[
        all_creator_results["Groupe"].isin(selected_groups)
    ].copy()
    if "Mode paiement" not in branch_creators.columns:
        branch_creators["Mode paiement"] = "Diamants"
    branch_creators = financial_columns(branch_creators)

    branch_consultants = calculate_consultant_rewards(
        creator_results=branch_creators,
        consultant_level=int(st.session_state.consultant_level),
        minimum_team_diamonds=200_000,
    )
    if not branch_consultants.empty:
        branch_consultants["Mode paiement"] = "Diamants"
        if "consultant_results" in st.session_state:
            saved_consultant_modes = (
                st.session_state.consultant_results
                .drop_duplicates("Consultant")
                .set_index("Consultant")["Mode paiement"]
                .to_dict()
            )
            branch_consultants["Mode paiement"] = branch_consultants[
                "Consultant"
            ].map(saved_consultant_modes).fillna("Diamants")
        consultant_exclusions = excluded_emails("consultants")
        consultant_excluded_mask = branch_consultants["Consultant"].map(
            lambda value: normalize_email(value) in consultant_exclusions
        )
        branch_consultants.loc[
            consultant_excluded_mask,
            ["Taux", "Rémunération 💎"],
        ] = 0
        branch_consultants = financial_columns(branch_consultants)

    branch_responsables = calculate_responsable_rewards(
        creator_results=branch_creators,
        responsable_level=int(st.session_state.manager_level),
        minimum_group_diamonds=600_000,
    )
    if not branch_responsables.empty:
        branch_responsables["Mode paiement"] = "Diamants"
        if "responsable_results" in st.session_state:
            saved_responsable_modes = (
                st.session_state.responsable_results
                .drop_duplicates("Responsable performance")
                .set_index("Responsable performance")["Mode paiement"]
                .to_dict()
            )
            branch_responsables["Mode paiement"] = branch_responsables[
                "Responsable performance"
            ].map(saved_responsable_modes).fillna("Diamants")
        responsable_exclusions = excluded_emails("responsables")
        responsable_excluded_mask = branch_responsables[
            "Responsable performance"
        ].map(lambda value: normalize_email(value) in responsable_exclusions)
        branch_responsables.loc[
            responsable_excluded_mask,
            ["Taux", "Rémunération 💎"],
        ] = 0
        branch_responsables = financial_columns(branch_responsables)

    revenue_eur = revenue_usd * active_usd_to_eur
    charges_eur = revenue_eur * float(st.session_state.charges_rate) / 100
    creator_cost = float(branch_creators["Total déduction €"].sum())
    consultant_cost = (
        float(branch_consultants["Total déduction €"].sum())
        if not branch_consultants.empty
        else 0.0
    )
    responsable_cost = (
        float(branch_responsables["Total déduction €"].sum())
        if not branch_responsables.empty
        else 0.0
    )
    net_profit_before_director = max(
        0.0,
        revenue_eur
        - charges_eur
        - creator_cost
        - consultant_cost
        - responsable_cost
        - branch_other_expenses,
    )
    director_rate = DIRECTOR_RATES.get(
        int(st.session_state.director_level),
        0.0,
    )
    director_reward = net_profit_before_director * director_rate
    remaining_after_director = (
        net_profit_before_director - director_reward
    )

    st.divider()
    st.subheader("Résultat de la branche")
    result1, result2, result3, result4 = st.columns(4)
    result1.metric("CA converti", f"{revenue_eur:,.2f} €")
    result2.metric(
        f"Charges ({st.session_state.charges_rate:.1f} %)",
        f"− {charges_eur:,.2f} €",
    )
    result3.metric(
        "Bénéfice avant directeur",
        f"{net_profit_before_director:,.2f} €",
    )
    result4.metric(
        f"Rémunération directeur ({director_rate * 100:.0f} %)",
        f"{director_reward:,.2f} €",
    )

    summary = pd.DataFrame(
        [
            ("Chiffre d’affaires converti", revenue_eur),
            ("Charges sur le CA", -charges_eur),
            ("Rémunérations créateurs", -creator_cost),
            ("Rémunérations consultants", -consultant_cost),
            ("Rémunérations responsables performance", -responsable_cost),
            ("Autres dépenses de la branche", -branch_other_expenses),
            ("Bénéfice avant directeur", net_profit_before_director),
            ("Rémunération du directeur", -director_reward),
            ("Bénéfice restant", remaining_after_director),
        ],
        columns=["Élément", "Montant €"],
    )
    st.dataframe(
        summary,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Montant €": st.column_config.NumberColumn(format="%.2f €")
        },
    )
    show_excel_download(
        summary,
        table_name="synthese_directeur",
        sheet_name="Synthèse directeur",
        key="download_director_summary",
    )

    show_workbook_download(
        [
            ("Synthèse", summary),
            ("Créateurs", branch_creators),
            ("Consultants", branch_consultants),
            ("Responsables", branch_responsables),
        ],
        table_name="resultats_complets_branche",
        key="download_complete_branch_workbook",
        label="📦 Télécharger tous les résultats de la branche (Excel)",
    )

    with st.expander("Voir les groupes sélectionnés"):
        st.write(selected_groups)
        st.caption(
            f"{len(branch_creators)} créateur(s), "
            f"{len(branch_consultants)} consultant(s) et "
            f"{len(branch_responsables)} responsable(s) pris en compte."
        )


elif page == "💰 Bénéfice agence":
    if current_user_role != "admin":
        st.error("Cette page est réservée à l’administrateur.")
        st.stop()

    st.title("💰 Bénéfice agence")
    show_estimation_notice()
    st.info(
        "Cette page calcule votre bénéfice après les charges, toutes les "
        "rémunérations, les autres dépenses et la rémunération globale "
        "des directeurs. Elle est visible uniquement par l’administrateur."
    )

    if st.session_state.backstage_data is None:
        st.warning("Importez d’abord l’export Backstage complet de l’agence.")
        st.stop()

    st.subheader("Chiffre d’affaires et taux quotidien")
    revenue_column, refresh_column = st.columns([3, 1])
    agency_revenue_usd = revenue_column.number_input(
        "Chiffre d’affaires total de l’agence ($)",
        min_value=0.0,
        value=float(st.session_state.revenue_usd),
        step=100.0,
        key="agency_revenue_input",
    )
    force_agency_rate_refresh = refresh_column.button(
        "🔄 Actualiser le taux",
        key="agency_refresh_daily_rate",
        use_container_width=True,
    )

    agency_rate_info = resolve_daily_usd_to_eur_rate(
        database_url=database_url if persistent_settings_available else None,
        updated_by=current_user_email,
        force_refresh=force_agency_rate_refresh,
    )
    st.success(
        f"Taux du {agency_rate_info['date']} : "
        f"1 $ = {agency_rate_info['rate']:.6f} € "
        f"— {agency_rate_info['source']}"
    )
    if agency_rate_info["warning"]:
        st.warning(agency_rate_info["warning"])

    if st.button(
        "💾 Enregistrer le chiffre d’affaires agence",
        key="save_agency_revenue",
        type="primary",
        use_container_width=True,
    ):
        st.session_state.revenue_usd = float(agency_revenue_usd)

        if persistent_settings_available:
            try:
                save_persistent_scopes(
                    database_url,
                    {
                        branch_settings_scope(current_user_email): {
                            key: st.session_state[key]
                            for key in BRANCH_PARAMETER_KEYS
                        }
                    },
                    current_user_email,
                )
                st.success(
                    "Le chiffre d’affaires agence a été enregistré "
                    "définitivement."
                )
            except Exception:
                st.error(
                    "Le chiffre d’affaires reste utilisable dans cette "
                    "session, mais sa sauvegarde permanente a échoué."
                )
        else:
            if persistent_settings_error:
                st.error(persistent_settings_error)
            else:
                st.warning(
                    "La section [database] avec son URL PostgreSQL est "
                    "absente des Secrets Streamlit. Le chiffre d’affaires "
                    "reste uniquement dans cette session."
                )

    creator_signature = (
        st.session_state.backstage_filename,
        st.session_state.creator_level,
    )
    if (
        "creator_results" not in st.session_state
        or st.session_state.get("creator_signature") != creator_signature
    ):
        creator_results = calculate_creator_rewards(
            st.session_state.backstage_data,
            int(st.session_state.creator_level),
        )
        creator_results["Mode paiement"] = "Diamants"
        st.session_state.creator_results = creator_results
        st.session_state.creator_signature = creator_signature

    agency_creators = st.session_state.creator_results.copy()
    if "Mode paiement" not in agency_creators.columns:
        agency_creators["Mode paiement"] = "Diamants"
    agency_creators = financial_columns(agency_creators)

    agency_consultants = calculate_consultant_rewards(
        creator_results=agency_creators,
        consultant_level=int(st.session_state.consultant_level),
        minimum_team_diamonds=200_000,
    )
    if not agency_consultants.empty:
        agency_consultants["Mode paiement"] = "Diamants"
        if "consultant_results" in st.session_state:
            saved_consultant_modes = (
                st.session_state.consultant_results
                .drop_duplicates("Consultant")
                .set_index("Consultant")["Mode paiement"]
                .to_dict()
            )
            agency_consultants["Mode paiement"] = agency_consultants[
                "Consultant"
            ].map(saved_consultant_modes).fillna("Diamants")

        consultant_exclusions = excluded_emails("consultants")
        consultant_excluded_mask = agency_consultants["Consultant"].map(
            lambda value: normalize_email(value) in consultant_exclusions
        )
        agency_consultants.loc[
            consultant_excluded_mask,
            ["Taux", "Rémunération 💎"],
        ] = 0
        agency_consultants = financial_columns(agency_consultants)

    agency_responsables = calculate_responsable_rewards(
        creator_results=agency_creators,
        responsable_level=int(st.session_state.manager_level),
        minimum_group_diamonds=600_000,
    )
    if not agency_responsables.empty:
        agency_responsables["Mode paiement"] = "Diamants"
        if "responsable_results" in st.session_state:
            saved_responsable_modes = (
                st.session_state.responsable_results
                .drop_duplicates("Responsable performance")
                .set_index("Responsable performance")["Mode paiement"]
                .to_dict()
            )
            agency_responsables["Mode paiement"] = agency_responsables[
                "Responsable performance"
            ].map(saved_responsable_modes).fillna("Diamants")

        responsable_exclusions = excluded_emails("responsables")
        responsable_excluded_mask = agency_responsables[
            "Responsable performance"
        ].map(lambda value: normalize_email(value) in responsable_exclusions)
        agency_responsables.loc[
            responsable_excluded_mask,
            ["Taux", "Rémunération 💎"],
        ] = 0
        agency_responsables = financial_columns(agency_responsables)

    creator_cost = float(agency_creators["Total déduction €"].sum())
    consultant_cost = (
        float(agency_consultants["Total déduction €"].sum())
        if not agency_consultants.empty
        else 0.0
    )
    responsable_cost = (
        float(agency_responsables["Total déduction €"].sum())
        if not agency_responsables.empty
        else 0.0
    )
    agency_other_expenses = float(st.session_state.other_expenses)
    director_rate = DIRECTOR_RATES.get(
        int(st.session_state.director_level),
        0.0,
    )

    available_agency_groups = sorted(
        value
        for value in agency_creators["Groupe"]
        .fillna("")
        .astype(str)
        .str.strip()
        .unique()
        if value and value.lower() != "nan"
    )
    if "admin_director_management_config" not in st.session_state:
        saved_director_payload = {}
        if persistent_settings_available:
            try:
                saved_director_payload = load_persistent_scope(
                    database_url,
                    director_management_scope(),
                ).get("directors", {})
            except Exception:
                st.warning(
                    "La répartition enregistrée des directeurs n’a pas pu "
                    "être chargée. L’estimation globale est utilisée."
                )
        st.session_state.admin_director_management_config = (
            clean_director_management_config(
                saved_director_payload,
                available_agency_groups,
            )
        )

    benefit_director_config = clean_director_management_config(
        st.session_state.admin_director_management_config,
        available_agency_groups,
    )
    benefit_group_owners = {}
    for profile in DIRECTOR_MANAGEMENT_PROFILES:
        for group in benefit_director_config[profile["email"]]["groups"]:
            benefit_group_owners.setdefault(group, []).append(
                profile["name"]
            )
    benefit_has_duplicate_groups = any(
        len(owners) > 1 for owners in benefit_group_owners.values()
    )
    benefit_has_director_assignments = bool(benefit_group_owners) and not (
        benefit_has_duplicate_groups
    )

    agency_calculation = calculate_agency_profit(
        revenue_usd=agency_revenue_usd,
        usd_to_eur=agency_rate_info["rate"],
        charges_rate=st.session_state.charges_rate,
        creator_cost=creator_cost,
        consultant_cost=consultant_cost,
        responsable_cost=responsable_cost,
        other_expenses=agency_other_expenses,
        director_rate=director_rate,
    )
    revenue_eur = agency_calculation["revenue_eur"]
    charges_eur = agency_calculation["charges_eur"]
    result_before_directors = agency_calculation[
        "result_before_directors"
    ]
    agency_director_overview = pd.DataFrame()
    if benefit_has_director_assignments:
        agency_director_overview = build_director_overview(
            all_creator_results=agency_creators,
            director_config=benefit_director_config,
            usd_to_eur=agency_rate_info["rate"],
        )
        directors_reward = float(
            agency_director_overview[
                "Montant facture directeur (€)"
            ].sum()
        )
        agency_profit = result_before_directors - directors_reward
        director_invoice_source = "Factures calculées des quatre directions"
    else:
        directors_reward = agency_calculation["directors_reward"]
        agency_profit = agency_calculation["agency_profit"]
        director_invoice_source = "Estimation globale du palier Directeur"

    if benefit_has_duplicate_groups:
        st.warning(
            "Un groupe est attribué à plusieurs directeurs dans une ancienne "
            "configuration. L’estimation globale est utilisée jusqu’à la "
            "correction dans l’onglet Directeur de branche."
        )

    agency_payment_totals = combined_payment_totals(
        agency_creators,
        agency_consultants,
        agency_responsables,
    )
    total_invoices_with_directors = (
        agency_payment_totals["invoices"] + directors_reward
    )

    st.divider()
    st.subheader("Votre résultat agence")
    metric1, metric2, metric3, metric4, metric5 = st.columns(5)
    metric1.metric("CA saisi", f"{agency_revenue_usd:,.2f} $")
    metric2.metric("CA converti", f"{revenue_eur:,.2f} €")
    metric3.metric(
        f"Charges ({st.session_state.charges_rate:.1f} %)",
        f"− {charges_eur:,.2f} €",
    )
    metric4.metric(
        f"Directeurs ({director_rate * 100:.0f} %)",
        f"− {directors_reward:,.2f} €",
    )
    metric5.metric("Votre bénéfice", f"{agency_profit:,.2f} €")

    st.subheader("Paiements à préparer")
    payment1, payment2, payment3 = st.columns(3)
    payment1.metric(
        "Diamants à recharger",
        f"{agency_payment_totals['diamonds']:,.0f} 💎",
    )
    payment2.metric(
        "Montant de la recharge",
        f"{agency_payment_totals['diamond_cost']:,.2f} €",
    )
    payment3.metric(
        "Total des factures",
        f"{total_invoices_with_directors:,.2f} €",
    )
    st.caption(
        "Le total des factures additionne les factures des créateurs, "
        "consultants et responsables, ainsi que les factures directeurs. "
        f"Source directeurs : {director_invoice_source}."
    )

    if agency_profit < 0:
        st.error(
            "Le résultat calculé est négatif : les charges et dépenses "
            "dépassent le chiffre d’affaires converti."
        )
    else:
        st.success(
            f"Bénéfice agence estimé pour {st.session_state.month} : "
            f"{agency_profit:,.2f} €"
        )

    st.caption(
        "Lorsque les groupes des quatre directeurs sont enregistrés, leurs "
        "factures individuelles remplacent l’estimation globale dans ce "
        "calcul. Le taux BCE ne modifie jamais la valeur facture par diamant."
    )

    agency_summary = pd.DataFrame(
        [
            ("Chiffre d’affaires converti", revenue_eur),
            ("Charges sur le CA", -charges_eur),
            ("Rémunérations créateurs", -creator_cost),
            ("Rémunérations consultants", -consultant_cost),
            (
                "Rémunérations responsables performance",
                -responsable_cost,
            ),
            ("Autres dépenses agence", -agency_other_expenses),
            ("Résultat avant directeurs", result_before_directors),
            (director_invoice_source, -directors_reward),
            ("Votre bénéfice agence", agency_profit),
        ],
        columns=["Élément", "Montant €"],
    )
    st.dataframe(
        agency_summary,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Montant €": st.column_config.NumberColumn(format="%.2f €")
        },
    )
    show_excel_download(
        agency_summary,
        table_name="benefice_agence",
        sheet_name="Bénéfice agence",
        key="download_agency_profit_summary",
    )

    agency_parameters = pd.DataFrame(
        [
            ("Mois", st.session_state.month),
            ("Chiffre d’affaires saisi ($)", agency_revenue_usd),
            ("Taux USD vers EUR", agency_rate_info["rate"]),
            ("Date du taux BCE", agency_rate_info["date"]),
            ("Source du taux", agency_rate_info["source"]),
            ("Palier Directeur (%)", director_rate * 100),
            ("Source factures directeurs", director_invoice_source),
            ("Charges (%)", st.session_state.charges_rate),
            ("Valeur facture par diamant (€)", st.session_state.invoice_rate),
        ],
        columns=["Paramètre", "Valeur"],
    )
    agency_payments = pd.DataFrame(
        [
            ("Diamants à recharger", agency_payment_totals["diamonds"]),
            (
                "Montant de la recharge (€)",
                agency_payment_totals["diamond_cost"],
            ),
            (
                "Factures créateurs, consultants et responsables (€)",
                agency_payment_totals["invoices"],
            ),
            ("Factures directeurs (€)", directors_reward),
            ("Total des factures (€)", total_invoices_with_directors),
        ],
        columns=["Paiement", "Total"],
    )
    if agency_director_overview.empty:
        agency_director_invoices = pd.DataFrame(
            [
                {
                    "Directeur": "Estimation globale",
                    "Montant facture directeur (€)": directors_reward,
                }
            ]
        )
    else:
        agency_director_invoices = agency_director_overview[
            [
                "Directeur",
                "Direction",
                "Groupes gérés",
                "CA branche ($)",
                "Montant facture directeur (€)",
            ]
        ].copy()
    show_workbook_download(
        [
            ("Bénéfice agence", agency_summary),
            ("Paiements", agency_payments),
            ("Factures directeurs", agency_director_invoices),
            ("Paramètres", agency_parameters),
            ("Créateurs", agency_creators),
            ("Consultants", agency_consultants),
            ("Responsables", agency_responsables),
        ],
        table_name="calcul_complet_benefice_agence",
        key="download_complete_agency_workbook",
        label="📦 Télécharger le calcul complet de l’agence (Excel)",
    )
